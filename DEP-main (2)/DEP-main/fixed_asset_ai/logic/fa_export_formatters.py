"""
Fixed Asset AI - Export Formatters Module

Excel formatting, styling, and conditional formatting utilities
for professional export files.

Author: Fixed Asset AI Team
"""

from __future__ import annotations

from typing import List, Optional
import pandas as pd


def _apply_professional_formatting(ws, df: pd.DataFrame):
    """
    Apply professional Excel formatting to worksheet.

    Includes:
    - Header row styling (bold, colored background)
    - Auto-column width
    - Cell borders
    - Currency and date formatting

    Args:
        ws: openpyxl worksheet object
        df: DataFrame used to create the worksheet
    """
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # Header row formatting
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Auto-size columns based on content
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass

        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

    # Apply thin borders to all cells
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Limit formatting to first 10,000 rows for performance
    max_rows = min(ws.max_row, 10000)

    for row in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Format currency columns
    currency_cols = [
        "Tax Cost", "Depreciable Basis", "Tax Sec 179 Expensed",
        "Bonus Amount", "Tax Cur Depreciation", "Tax Prior Depreciation",
        "De Minimis Expensed", "Section 179 Allowed", "Section 179 Carryforward",
        "Capital Gain", "Capital Loss", "§1245 Recapture (Ordinary Income)",
        "§1250 Recapture (Ordinary Income)", "Unrecaptured §1250 Gain (25%)",
        "Total Year 1 Deduction", "Gross Proceeds", "Book Cost", "MI Cost",
        "NBV", "NBV_Computed",
    ]

    _apply_number_format(ws, df, currency_cols, '$#,##0.00')

    # Format percentage columns
    pct_cols = ["MaterialityScore", "Bonus % Applied"]
    _apply_number_format(ws, df, pct_cols, '0.00')

    # Format date columns
    date_cols = ["Date In Service", "Acquisition Date", "Date Disposed", "Disposal Date"]
    _apply_number_format(ws, df, date_cols, 'M/D/YYYY')


def _apply_number_format(ws, df: pd.DataFrame, columns: List[str], format_str: str):
    """
    Apply number format to specific columns.

    Args:
        ws: openpyxl worksheet
        df: DataFrame with column names
        columns: List of column names to format
        format_str: Excel number format string
    """
    max_rows = min(ws.max_row + 1, 10001)

    for col_name in columns:
        if col_name in df.columns:
            col_idx = list(df.columns).index(col_name) + 1

            # Handle column letters beyond Z (AA, AB, etc.)
            if col_idx <= 26:
                col_letter = chr(64 + col_idx)
            else:
                col_letter = chr(64 + col_idx // 26) + chr(64 + col_idx % 26)

            for row in range(2, max_rows):
                try:
                    cell = ws[f'{col_letter}{row}']
                    cell.number_format = format_str
                except (KeyError, AttributeError):
                    pass


def _apply_conditional_formatting(ws, df: pd.DataFrame):
    """
    Apply conditional formatting to highlight important items.

    Highlights:
    - Red: Critical issues (NBV_Reco = CHECK, ConfidenceGrade = D)
    - Orange: High priority items
    - Yellow: Medium priority / warnings
    - Green: OK status

    Args:
        ws: openpyxl worksheet
        df: DataFrame with column names
    """
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule

    # Define fill colors
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    orange_fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

    # NBV_Reco column
    if "NBV_Reco" in df.columns:
        col_letter = _get_column_letter(df, "NBV_Reco")
        if col_letter:
            # Red for CHECK
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}10000',
                CellIsRule(operator='equal', formula=['"CHECK"'], fill=red_fill)
            )
            # Green for OK
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}10000',
                CellIsRule(operator='equal', formula=['"OK"'], fill=green_fill)
            )

    # ReviewPriority column
    if "ReviewPriority" in df.columns:
        col_letter = _get_column_letter(df, "ReviewPriority")
        if col_letter:
            # Orange for High
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}10000',
                CellIsRule(operator='equal', formula=['"High"'], fill=orange_fill)
            )
            # Yellow for Medium
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}10000',
                CellIsRule(operator='equal', formula=['"Medium"'], fill=yellow_fill)
            )
            # Green for Low (optional)
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}10000',
                CellIsRule(operator='equal', formula=['"Low"'], fill=green_fill)
            )

    # ConfidenceGrade column
    if "ConfidenceGrade" in df.columns:
        col_letter = _get_column_letter(df, "ConfidenceGrade")
        if col_letter:
            # Red for D grade
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}10000',
                CellIsRule(operator='equal', formula=['"D"'], fill=red_fill)
            )
            # Yellow for C grade
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}10000',
                CellIsRule(operator='equal', formula=['"C"'], fill=yellow_fill)
            )
            # Green for A grade
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}10000',
                CellIsRule(operator='equal', formula=['"A"'], fill=green_fill)
            )


def _get_column_letter(df: pd.DataFrame, column_name: str) -> Optional[str]:
    """
    Get Excel column letter for a DataFrame column.

    Args:
        df: DataFrame
        column_name: Column name to find

    Returns:
        Column letter (A, B, ..., AA, AB, ...) or None if not found
    """
    if column_name not in df.columns:
        return None

    col_idx = list(df.columns).index(column_name) + 1

    if col_idx <= 26:
        return chr(64 + col_idx)
    else:
        return chr(64 + col_idx // 26) + chr(64 + col_idx % 26)


def format_summary_sheet(ws, summary_df: pd.DataFrame):
    """
    Apply formatting to summary sheet.

    Args:
        ws: openpyxl worksheet
        summary_df: Summary DataFrame
    """
    from openpyxl.styles import Font, PatternFill

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25

    # Bold section headers (all caps rows)
    bold_font = Font(bold=True, size=12)
    for row_num in range(1, ws.max_row + 1):
        cell_value = str(ws[f'A{row_num}'].value or "")
        if cell_value and cell_value.isupper() and not cell_value.startswith(" "):
            ws[f'A{row_num}'].font = bold_font

    # Highlight total deduction row
    total_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    large_bold = Font(bold=True, size=14)

    for row_num in range(1, ws.max_row + 1):
        cell_value = str(ws[f'A{row_num}'].value or "")
        if "TOTAL YEAR 1 DEDUCTION" in cell_value:
            ws[f'A{row_num}'].fill = total_fill
            ws[f'B{row_num}'].fill = total_fill
            ws[f'A{row_num}'].font = large_bold
            ws[f'B{row_num}'].font = large_bold


def create_excel_table_style(ws, df: pd.DataFrame, table_name: str = "AssetTable"):
    """
    Apply Excel table style for better filtering and sorting.

    Args:
        ws: openpyxl worksheet
        df: DataFrame
        table_name: Name for the Excel table
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Define table range
    if ws.max_row > 1:
        end_col = chr(64 + min(len(df.columns), 26)) if len(df.columns) <= 26 else "Z"
        table_range = f"A1:{end_col}{ws.max_row}"

        # Create table
        table = Table(displayName=table_name, ref=table_range)

        # Apply style
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        try:
            ws.add_table(table)
        except ValueError:
            # Table already exists or invalid range
            pass
