import pandas as pd
from openpyxl import Workbook
from io import BytesIO
import hashlib
import datetime
import re


# =====================================================================
# UTILITIES
# =====================================================================

def _ensure_cols(df: pd.DataFrame, cols_with_defaults: dict) -> pd.DataFrame:
    """Add missing columns with default values."""
    for col, default in cols_with_defaults.items():
        if col not in df.columns:
            df[col] = default
    return df


def _normalize_text(x: str) -> str:
    """Normalize string for consistent matching."""
    if not isinstance(x, str):
        return ""
    return x.strip().lower()


# =====================================================================
# DESCRIPTION & CLIENT CATEGORY TYPO CORRECTION
# =====================================================================

DESC_TYPOES = {
    r"\bsoftwar\b": "software",
    r"\bcemra\b": "camera",
    r"\bcamra\b": "camera",
    r"\bmonitr\b": "monitor",
    r"\btractr\b": "tractor",
    r"\bvehicl\b": "vehicle",
}

CLIENT_CAT_FIX = {
    "softwar": "software",
    "sofware": "software",
    "vehicl": "vehicle",
    "equipmnt": "equipment",
    "equpment": "equipment",
    "furnitre": "furniture",
}


def _fix_description(desc: str):
    """Correct common description typos."""
    if not isinstance(desc, str):
        return desc, False, ""

    original = desc
    fixed = desc
    flag = False

    for typo, correct in DESC_TYPOES.items():
        if re.search(typo, fixed, flags=re.IGNORECASE):
            fixed = re.sub(typo, correct, fixed, flags=re.IGNORECASE)
            flag = True

    note = f"Corrected description: '{original}' → '{fixed}'" if flag else ""
    return fixed, flag, note


def _fix_client_category(cat: str):
    """Correct common category typos."""
    if not isinstance(cat, str):
        return cat, False, ""

    original = _normalize_text(cat)
    if original in CLIENT_CAT_FIX:
        corrected = CLIENT_CAT_FIX[original]
        return corrected, True, f"Corrected category: '{cat}' → '{corrected}'"

    return cat, False, ""


# =====================================================================
# NBV RECONCILIATION
# =====================================================================

def _compute_nbv_reco(df: pd.DataFrame, tolerance: float = 5.0) -> pd.DataFrame:
    df = df.copy()

    df = _ensure_cols(df, {"Cost": None, "Accum Dep": None, "NBV": None})

    df["NBV_Derived"] = pd.NA
    mask = df["Cost"].notna() & df["Accum Dep"].notna()
    df.loc[mask, "NBV_Derived"] = df.loc[mask, "Cost"] - df.loc[mask, "Accum Dep"]

    df["NBV_Diff"] = 0.0
    mask2 = df["NBV"].notna() & df["NBV_Derived"].notna()
    df.loc[mask2, "NBV_Diff"] = df.loc[mask2, "NBV"] - df.loc[mask2, "NBV_Derived"]

    df["NBV_Reco"] = "OK"
    df.loc[mask & df["NBV"].isna(), "NBV_Reco"] = "CHECK"
    df.loc[df["NBV_Diff"].abs() > tolerance, "NBV_Reco"] = "CHECK"

    return df


# =====================================================================
# MATERIALITY ENGINE
# =====================================================================

def _compute_materiality(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df = _ensure_cols(df, {"Cost": 0.0, "NBV": 0.0})

    base = df["Cost"].abs() + df["NBV"].abs()
    max_val = base.max() or 1.0
    df["MaterialityScore"] = (base / max_val) * 100.0

    def _priority(v):
        if v >= 70: return "High"
        if v >= 40: return "Medium"
        return "Low"

    df["ReviewPriority"] = df["MaterialityScore"].apply(_priority)
    return df


# =====================================================================
# QIP DETECTOR
# =====================================================================

def _qip_flag_row(row) -> str:
    desc = str(row.get("Description", "")).lower()
    cat  = str(row.get("Final Category Used", "")).lower()
    life = row.get("Final Life Used", None)

    qip_words = [
        "improvement","reno","remodel","build-out",
        "fit-out","tenant","ti","interior"
    ]

    looks_interior = any(w in desc for w in qip_words)
    is_nonres = ("nonresidential" in cat or "building" in desc or "warehouse" in desc)

    if is_nonres and looks_interior and life in (15, 39, 27.5):
        return "POSSIBLE_QIP"

    return "NO"


# =====================================================================
# BONUS REASONING
# =====================================================================

def _bonus_reason_row(row) -> str:
    bonus = bool(row.get("Bonus Eligible", False))
    ads = bool(row.get("ADS Property", False))
    life = row.get("Final Life Used", None)
    txn = str(row.get("Transaction Type", ""))

    if txn == "Existing":
        return "Existing asset – no new bonus."
    if ads:
        return "ADS property – bonus not allowed."
    if not bonus:
        return "Not bonus-eligible (life > 20 or excluded)."
    if life in (3, 5, 7, 15):
        return "Bonus-eligible personal property/QIP."

    return "Bonus flag set but life > 15 – CPA confirm."

# =====================================================================
# ISSUE SEVERITY ENGINE
# =====================================================================

def _determine_issue_severity(row) -> str:
    txn = str(row.get("Transaction Type", ""))
    nbv = row.get("NBV")
    cost = row.get("Cost")
    reco = str(row.get("NBV_Reco"))
    src = str(row.get("Source"))
    conf = row.get("Rule Confidence")

    if txn in ("Disposal", "Transfer") and (cost or 0) > 0 and (nbv in (None, 0)):
        return "ERROR"

    if reco == "CHECK":
        return "WARN"

    if conf is not None and conf < 0.50:
        return "WARN"

    if src.startswith("gpt_error"):
        return "WARN"

    return "OK"


# =====================================================================
# RPA ROUTING
# =====================================================================

def _rpa_action(row) -> str:
    txn = str(row.get("Transaction Type", ""))
    if txn == "Addition": return "ADD"
    if txn == "Transfer": return "TRANSFER"
    if txn == "Disposal": return "DISPOSE"
    return "NO_ACTION"


# =====================================================================
# BONUS & §179 LIMITATION ENGINE
# =====================================================================

def apply_179_bonus_limits(
    review_df: pd.DataFrame,
    sec179_limit: float,
    sec179_phaseout_threshold: float,
    total_other_179_cost: float,
    business_income_limit: float
) -> pd.DataFrame:

    df = review_df.copy()
    df = _ensure_cols(df, {
        "Transaction Type": "", "Cost": 0.0, "Section 179 Eligible": False,
        "Bonus Eligible": False, "ADS Property": False, "PIS Date": pd.NaT
    })

    df["Cost_num"] = df["Cost"].fillna(0).astype(float)

    is_add = df["Transaction Type"] == "Addition"
    is_179 = df["Section 179 Eligible"] == True

    total_179_this_file = df.loc[is_add & is_179, "Cost_num"].sum()
    total_pool = total_179_this_file + total_other_179_cost

    phaseout = max(0.0, total_pool - sec179_phaseout_threshold)
    effective_limit = max(0.0, sec179_limit - phaseout)
    effective_cap = min(effective_limit, business_income_limit)

    df["Suggested179Amt"] = 0.0

    if effective_cap > 0:
        candidates = df.loc[is_add & is_179].copy()
        candidates = candidates.sort_values(
            ["MaterialityScore", "Cost_num"], ascending=[False, False]
        )
        remaining = effective_cap
        alloc = {}

        for _, row in candidates.iterrows():
            aid = row["Asset ID"]
            amt = min(row["Cost_num"], remaining)
            alloc[aid] = amt
            remaining -= amt
            if remaining <= 0:
                break

        df.loc[is_add & is_179, "Suggested179Amt"] = (
            df.loc[is_add & is_179, "Asset ID"].map(alloc).fillna(0.0)
        )

    # Bonus %
    def _bonus_pct(row):
        if not row.get("Bonus Eligible", False): return 0
        if row.get("ADS Property", False): return 0
        pis = row.get("PIS Date")
        try:
            y = pis.year
        except:
            return 0
        if y >= 2023: return 80   # Adjust as IRS phases down
        return 100

    df["SuggestedBonusPct"] = df.apply(_bonus_pct, axis=1)
    df["SuggestedBonusAmt"] = (df["Cost_num"] * df["SuggestedBonusPct"] / 100).round(2)
    df.drop(columns=["Cost_num"], inplace=True)

    return df


# =====================================================================
# CLASSIFICATION EXPLANATION ENGINE
# =====================================================================

def _classification_explanation(row):
    """Explain MACRS classification following IRS logic."""
    cat = row.get("Final Category Used", "")
    life = row.get("Final Life Used", "")
    src = row.get("Source", "")

    base = f"Classified as {cat} ({life}-year) because "

    if src.startswith("gpt"):
        return base + "GPT MACRS reasoning using IRS mapping tables."
    if src == "rules":
        lc = cat.lower()
        if "land (non" in lc:
            return base + "land is non-depreciable under §167."
        if "improvement" in lc:
            return base + "IRS MACRS Table B-1 lists land improvements as 15-year property."
        if "qualified improvement" in lc:
            return base + "QIP is 15-year interior non-structural under §168(k)(3)."
        if "real property" in lc:
            return base + "nonresidential real property depreciates over 39-year MM."
        if "residential rental" in lc:
            return base + "residential rental is 27.5-year MM."
        if "machinery" in lc or "equipment" in lc:
            return base + "general machinery is 7-year GDS property."
        if "vehicle" in lc:
            return base + "vehicles fall under 5-year MACRS GDS."
        if "computer" in lc:
            return base + "computers are 5-year MACRS GDS."
        if "furniture" in lc:
            return base + "office furniture is 7-year MACRS GDS."
    return base + "fallback personal property rule."


# =====================================================================
# MACRS REASON CODE + CONFIDENCE GRADE
# =====================================================================

def _macrs_reason_code(row) -> str:
    cat = str(row.get("Final Category Used", "")).lower()
    if "land (non" in cat: return "L0"
    if "land improvement" in cat: return "LI15"
    if "qualified improvement" in cat: return "QIP15"
    if "real property" in cat: return "RP39"
    if "residential rental" in cat: return "RR27"
    if "vehicle" in cat: return "V5"
    if "computer" in cat: return "C5"
    if "furniture" in cat: return "F7"
    if "equipment" in cat or "machinery" in cat: return "M7"
    return "PP7"


def _confidence_grade(row) -> str:
    conf = row.get("Rule Confidence", None)
    try: c = float(conf)
    except: return "Unknown"
    if c >= 0.90: return "A"
    if c >= 0.75: return "B"
    if c >= 0.60: return "C"
    return "D"


# =====================================================================
# AUDIT FIELDS ENGINE (INCLUDING TYPO DETECTION)
# =====================================================================

def _audit_fields(row) -> dict:
    """Full audit block with typo flags, override history, integrity hash."""
    audit = {}

    # Source
    src = str(row.get("Source", ""))
    if src == "rules": audit["AuditSource"] = "Rule Engine"
    elif src.startswith("gpt"): audit["AuditSource"] = "GPT Classifier"
    else: audit["AuditSource"] = "Client / Fallback"

    # Rule trigger summary
    cat = str(row.get("Final Category Used", "")).lower()
    if "land (non" in cat: audit["AuditRuleTriggers"] = "Land → Non-Depreciable"
    elif "land improvement" in cat: audit["AuditRuleTriggers"] = "Land Improvement (15-yr)"
    elif "qualified improvement" in cat: audit["AuditRuleTriggers"] = "QIP → 15-year"
    elif "real property" in cat: audit["AuditRuleTriggers"] = "Nonresidential Real Property (39-yr)"
    elif "residential rental" in cat: audit["AuditRuleTriggers"] = "Residential Rental (27.5-yr)"
    elif "vehicle" in cat: audit["AuditRuleTriggers"] = "Vehicle (5-yr)"
    elif "equipment" in cat or "machinery" in cat: audit["AuditRuleTriggers"] = "7-year Machinery"
    else: audit["AuditRuleTriggers"] = "Personal Property Fallback"

    # Warnings
    warnings = []
    if row.get("IssueSeverity") == "ERROR":
        warnings.append("Error-level depreciation/transaction issue")
    if row.get("IssueSeverity") == "WARN":
        warnings.append("Validation warning")
    if row.get("QIP_Flag") == "POSSIBLE_QIP":
        warnings.append("Possible QIP flagged")
    if row.get("NBV_Reco") == "CHECK":
        warnings.append("NBV out of balance")
    if row.get("Desc_TypoFlag") == "YES":
        warnings.append("Description corrected for typos")
    if row.get("Cat_TypoFlag") == "YES":
        warnings.append("Client category corrected")

    audit["AuditWarnings"] = "; ".join(warnings) if warnings else "None"

    # Override history
    override_list = []
    if row.get("CPA_Override_Category"):
        override_list.append("Category overridden by CPA")
    if row.get("CPA_Apply_Bonus"):
        override_list.append(f"Bonus election={row.get('CPA_Apply_Bonus')}")
    if row.get("CPA_Apply_179"):
        override_list.append(f"179 election={row.get('CPA_Apply_179')}, Amt={row.get('CPA_179_Amount')}")
    if row.get("CPA_Notes"):
        override_list.append("CPA Notes entered")

    audit["AuditOverrideHistory"] = "; ".join(override_list) if override_list else "None"

    # Hash for classification integrity
    hash_input = (
        f"{row.get('Asset ID','')}|"
        f"{row.get('Final Category Used','')}|"
        f"{row.get('Final Life Used','')}|"
        f"{row.get('Final Method Used','')}|"
        f"{row.get('Final Conv Used','')}"
    )
    audit["ClassificationHash"] = hashlib.sha256(hash_input.encode("utf-8")).hexdigest()

    audit["AuditTimestamp"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Typo summary
    audit["AuditTypoSummary"] = row.get("Desc_TypoNote","") + " " + row.get("Cat_TypoNote","")

    return audit


# =====================================================================
# MAIN CPA REVIEW BUILDER
# =====================================================================

def build_cpa_review(df: pd.DataFrame,
                     sec179_limit=1280000,
                     sec179_phaseout_threshold=2550000,
                     total_other_179_cost=0,
                     business_income_limit=999999999):

    review_df = df.copy()

    # Fix typos before classification export
    review_df["Desc_Corrected"], review_df["Desc_TypoFlag"], review_df["Desc_TypoNote"] = zip(
        *review_df["Description"].apply(_fix_description)
    )
    review_df["Client_Cat_Corrected"], review_df["Cat_TypoFlag"], review_df["Cat_TypoNote"] = zip(
        *review_df["Client Category"].apply(_fix_client_category)
    )

    # Use corrected descriptions for downstream logic
    review_df["Description"] = review_df["Desc_Corrected"]
    review_df["Client Category"] = review_df["Client_Cat_Corrected"]

    # Apply NBV/Materiality
    review_df = _compute_nbv_reco(review_df)
    review_df = _compute_materiality(review_df)

    # Apply Bonus/179 limitations
    review_df = apply_179_bonus_limits(
        review_df,
        sec179_limit,
        sec179_phaseout_threshold,
        total_other_179_cost,
        business_income_limit
    )

    # Derived fields
    review_df["QIP_Flag"] = review_df.apply(_qip_flag_row, axis=1)
    review_df["BonusReason"] = review_df.apply(_bonus_reason_row, axis=1)
    review_df["IssueSeverity"] = review_df.apply(_determine_issue_severity, axis=1)
    review_df["RPA_Action"] = review_df.apply(_rpa_action, axis=1)

    # Classification explanation / MACRS metadata
    review_df["ClassificationExplanation"] = review_df.apply(_classification_explanation, axis=1)
    review_df["MACRS_Reason_Code"] = review_df.apply(_macrs_reason_code, axis=1)
    review_df["ConfidenceGrade"] = review_df.apply(_confidence_grade, axis=1)

    # Audit block
    audit_df = review_df.apply(_audit_fields, axis=1, result_type="expand")
    review_df = pd.concat([review_df, audit_df], axis=1)

    # Override scaffolding
    overrides = {
        "CPA_Apply_Bonus": "",
        "CPA_Apply_179": "",
        "CPA_179_Amount": "",
        "CPA_Reviewed": "",
        "CPA_Override_Category": "",
        "CPA_Notes": "",
    }
    review_df = _ensure_cols(review_df, overrides)

    # Final ordering
    ordered = [
        "Asset ID","Description","Transaction Type","Client Category",
        "Acquisition Date","PIS Date",

        "Cost","Accum Dep","NBV", "NBV_Derived","NBV_Diff","NBV_Reco",

        "Rule Category","Rule Life","Rule Method","Rule Conv",
        "Final Category Used","Final Life Used","Final Method Used","Final Conv Used",

        "Bonus Eligible","Section 179 Eligible","ADS Property",
        "SuggestedBonusPct","SuggestedBonusAmt","Suggested179Amt",

        "CPA_Apply_Bonus","CPA_Apply_179","CPA_179_Amount",

        "IssueSeverity","MaterialityScore","ReviewPriority",
        "QIP_Flag","BonusReason","RPA_Action",

        "MACRS_Reason_Code","ConfidenceGrade","ClassificationExplanation",

        # typo audit
        "Desc_TypoFlag","Desc_TypoNote","Cat_TypoFlag","Cat_TypoNote",

        # audit block
        "AuditSource","AuditRuleTriggers","AuditWarnings",
        "AuditNBVCheck","AuditOverrideHistory",
        "AuditTypoSummary","AuditTimestamp","ClassificationHash",

        "CPA_Reviewed","CPA_Override_Category","CPA_Notes",
    ]

    review_df = _ensure_cols(review_df, {c: "" for c in ordered})
    review_df = review_df[ordered]

    return review_df


# =====================================================================
# EXCEL EXPORT
# =====================================================================

def export_cpa_excel(review_df: pd.DataFrame) -> bytes:
    safe_df = review_df.copy()

    # Convert dates to strings
    for col in safe_df.columns:
        if pd.api.types.is_datetime64_any_dtype(safe_df[col]):
            safe_df[col] = safe_df[col].apply(
                lambda x: x.strftime("%m/%d/%Y") if not pd.isna(x) else ""
            )

    def excel_safe(x):
        if x is None: return ""
        if isinstance(x, float) and pd.isna(x): return ""
        if isinstance(x, (list, dict)): return str(x)
        if pd.isna(x): return ""
        return x

    safe_df = safe_df.applymap(excel_safe)

    wb = Workbook()
    ws = wb.active
    ws.title = "CPA Review"

    for ci, col in enumerate(safe_df.columns, start=1):
        ws.cell(row=1, column=ci, value=str(col))

    for ri, (_, row) in enumerate(safe_df.iterrows(), start=2):
        for ci, col in enumerate(safe_df.columns, start=1):
            ws.cell(row=ri, column=ci, value=row[col])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

