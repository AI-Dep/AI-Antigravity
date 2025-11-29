# logic/fa_export.py
"""
Fixed Asset AI - Main Export Builder

Core export functionality for generating FA CS import files.
This module has been refactored with supporting modules:
- fa_export_validation.py: Data quality, NBV reconciliation, materiality
- fa_export_audit.py: Classification explanations, audit trail
- fa_export_vehicles.py: Luxury auto rules, vehicle detection
- fa_export_formatters.py: Excel formatting utilities
- fa_cs_mappings.py: FA CS wizard category mappings

Author: Fixed Asset AI Team
"""

from io import BytesIO, StringIO
from datetime import datetime, date
from typing import Optional, Dict
import hashlib
import re

import pandas as pd

# ==============================================================================
# NUMERICAL TOLERANCE CONSTANTS
# ==============================================================================
# Use tolerance-based comparison for floating point values to avoid rounding errors
FLOAT_TOLERANCE = 0.0001  # Tolerance for float comparisons (0.01 cents)

from .parse_utils import parse_date, parse_number
from .tax_year_config import (
    get_bonus_percentage,
    get_section_179_limits,
    get_luxury_auto_limits,
    get_heavy_suv_179_limit,
    validate_tax_year_config,
)
from .strategy_config import get_strategy, AGGRESSIVE, BALANCED, CONSERVATIVE
from .data_validator import validate_asset_data
from .ads_system import should_use_ads, apply_ads_to_asset
from .listed_property import (
    is_listed_property,
    get_business_use_percentage,
    validate_business_use_for_incentives,
)
from .recapture import (
    calculate_section_1245_recapture,
    calculate_section_1250_recapture,
    determine_recapture_type,
)
from .convention_rules import (
    detect_mid_quarter_convention,
    get_quarter,
)
from .macrs_tables import calculate_macrs_depreciation, calculate_disposal_year_depreciation, get_macrs_table
from .section_179_carryforward import (
    apply_section_179_carryforward_to_dataframe,
    generate_section_179_report,
)
from .export_qa_validator import validate_fixed_asset_cs_export
from .transaction_classifier import (
    classify_all_transactions,
    validate_transaction_classification,
)

# Import from refactored modules
from .fa_export_validation import (
    _fix_description,
    _fix_client_category,
    _compute_nbv_reco,
    _compute_materiality,
    _is_disposal,
    _is_transfer,
    _is_current_year,
    _format_date_for_fa_cs,
    _pick,
)
from .fa_export_audit import (
    _classification_explanation,
    _macrs_reason_code,
    _confidence_grade,
    _audit_fields,
)
from .fa_export_vehicles import (
    _is_passenger_auto,
    _is_heavy_suv,
    _apply_luxury_auto_caps,
    _determine_asset_type,
)
from .fa_cs_mappings import _get_fa_cs_wizard_category
from .fa_export_formatters import (
    _apply_professional_formatting,
    _apply_conditional_formatting,
    format_summary_sheet,
)


# --------------------------------------------------------------
# TYPO CORRECTION & DATA QUALITY
# --------------------------------------------------------------

DESC_TYPOES = {
    r"\bsoftwar\b": "software",
    r"\bcemra\b": "camera",
    r"\bcamra\b": "camera",
    r"\bmonitr\b": "monitor",
    r"\btractr\b": "tractor",
    r"\bvehicl\b": "vehicle",
}

CLIENT_CAT_FIX = {
    "softwar": "software",
    "sofware": "software",
    "vehicl": "vehicle",
    "equipmnt": "equipment",
    "equpment": "equipment",
    "furnitre": "furniture",
}


def _normalize_text(x: str) -> str:
    """Normalize string for consistent matching."""
    if not isinstance(x, str):
        return ""
    return x.strip().lower()


def _format_date_for_fa_cs(date_series: pd.Series) -> pd.Series:
    """
    Format dates for FA CS import (M/D/YYYY format without leading zeros).

    Platform-independent solution that works on both Windows and Unix.
    Windows doesn't support the %-m directive, so we use custom formatting.

    Args:
        date_series: Pandas Series of dates

    Returns:
        Series of formatted date strings (e.g., "1/15/2024", not "01/15/2024")
    """
    def format_single_date(dt):
        if pd.isna(dt):
            return ""
        # Convert to datetime if needed
        if not isinstance(dt, (datetime, date, pd.Timestamp)):
            dt = pd.to_datetime(dt)
        # Format without leading zeros: M/D/YYYY
        return f"{dt.month}/{dt.day}/{dt.year}"

    return date_series.apply(format_single_date)


def _fix_description(desc: str):
    """Correct common description typos."""
    if not isinstance(desc, str):
        return desc, False, ""

    original = desc
    fixed = desc
    flag = False

    for typo, correct in DESC_TYPOES.items():
        if re.search(typo, fixed, flags=re.IGNORECASE):
            fixed = re.sub(typo, correct, fixed, flags=re.IGNORECASE)
            flag = True

    note = f"Corrected description: '{original}' → '{fixed}'" if flag else ""
    return fixed, flag, note


def _fix_client_category(cat: str):
    """Correct common category typos."""
    if not isinstance(cat, str):
        return cat, False, ""

    original = _normalize_text(cat)
    if original in CLIENT_CAT_FIX:
        corrected = CLIENT_CAT_FIX[original]
        return corrected, True, f"Corrected category: '{cat}' → '{corrected}'"

    return cat, False, ""


# --------------------------------------------------------------
# NBV RECONCILIATION
# --------------------------------------------------------------

def _compute_nbv_reco(df: pd.DataFrame, tolerance: float = 5.0) -> pd.DataFrame:
    """
    Validate Net Book Value = Cost - Accumulated Depreciation.

    Flags assets with NBV discrepancies > tolerance for CPA review.
    """
    df = df.copy()

    # Determine which accumulated depreciation column to use
    # Check for various column names (input may have different naming)
    accum_dep_col = None
    for col_name in ["Accumulated Depreciation", "Tax Prior Depreciation", "AccumDepr"]:
        if col_name in df.columns:
            accum_dep_col = col_name
            break

    # Ensure Cost column exists and convert to numeric
    if "Cost" not in df.columns:
        df["Cost"] = 0.0
    else:
        df["Cost"] = pd.to_numeric(df["Cost"], errors='coerce').fillna(0.0)

    # Ensure NBV column exists (preserve existing values)
    if "NBV" not in df.columns:
        df["NBV"] = None
    else:
        # Convert to numeric but preserve None/NaN
        df["NBV"] = pd.to_numeric(df["NBV"], errors='coerce')

    # Handle accumulated depreciation
    if accum_dep_col:
        df["_AccumDep"] = pd.to_numeric(df[accum_dep_col], errors='coerce').fillna(0.0)
    else:
        df["_AccumDep"] = 0.0

    # Calculate derived NBV = Cost - Accumulated Depreciation
    df["NBV_Derived"] = 0.0
    mask = df["Cost"].notna() & df["_AccumDep"].notna()
    df.loc[mask, "NBV_Derived"] = df.loc[mask, "Cost"] - df.loc[mask, "_AccumDep"]

    # FILL IN MISSING NBV FROM DERIVED VALUE
    # If client didn't provide NBV but we have Cost and Accumulated Depreciation,
    # calculate and fill in NBV automatically. This is standard accounting:
    # NBV = Cost - Accumulated Depreciation
    missing_nbv_mask = mask & df["NBV"].isna()
    if missing_nbv_mask.any():
        df.loc[missing_nbv_mask, "NBV"] = df.loc[missing_nbv_mask, "NBV_Derived"]

    # Calculate difference between provided NBV and derived NBV
    df["NBV_Diff"] = 0.0
    mask2 = df["NBV"].notna() & df["NBV_Derived"].notna()
    if mask2.any():
        df.loc[mask2, "NBV_Diff"] = df.loc[mask2, "NBV"] - df.loc[mask2, "NBV_Derived"]

    # Set reconciliation status
    df["NBV_Reco"] = "OK"
    # Flag if NBV difference exceeds tolerance (client-provided NBV doesn't match calculated)
    df.loc[df["NBV_Diff"].abs() > tolerance, "NBV_Reco"] = "CHECK"
    # Note: We no longer flag missing NBV as "CHECK" since we auto-fill it from derived value

    # Clean up temporary column
    df = df.drop(columns=["_AccumDep"])

    return df


# --------------------------------------------------------------
# MATERIALITY SCORING
# --------------------------------------------------------------

def _compute_materiality(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate materiality score (0-100) for CPA review prioritization.

    Higher scores indicate more material assets requiring closer review.
    """
    df = df.copy()

    # Ensure Cost column exists and is numeric
    if "Cost" not in df.columns:
        df["Cost"] = 0.0
    else:
        df["Cost"] = pd.to_numeric(df["Cost"], errors='coerce').fillna(0.0)

    base = df["Cost"].abs()
    max_val = base.max() or 1.0
    df["MaterialityScore"] = (base / max_val) * 100.0

    def _priority(v):
        if v >= 70: return "High"
        if v >= 40: return "Medium"
        return "Low"

    df["ReviewPriority"] = df["MaterialityScore"].apply(_priority)
    return df


# --------------------------------------------------------------
# FA CS WIZARD CATEGORY MAPPING
# --------------------------------------------------------------
# Maps our asset classifications to exact FA CS wizard dropdown options
# Used for UiPath RPA automation to select correct asset type

def _get_fa_cs_wizard_category(row) -> str:
    """
    Map our Final Category + Description to FA CS wizard dropdown option.
    Returns the EXACT text that appears in FA CS wizard dropdown.
    """
    final_cat = str(row.get("Final Category", "")).lower()
    desc = str(row.get("Description", "")).lower()
    recovery = row.get("Recovery Period", row.get("MACRS Life", 0))

    # Handle disposals and transfers - no wizard needed
    if _is_disposal(row) or _is_transfer(row):
        return ""

    # ===== 5-YEAR PROPERTY =====
    # Computers
    if any(kw in desc for kw in ['computer', 'laptop', 'desktop', 'monitor', 'server', 'workstation', 'pc ', 'imac', 'macbook']):
        return "Computer, monitor, laptop, PDA or peripheral equip (incld aux mach undr ctrl of cpu/don't incld comp primrly for proc or prod ctrl,switchg,chanellg & POS)"
    if 'printer' in desc:
        return "Printer"
    if 'scanner' in desc:
        return "Scanner"
    if 'copier' in desc or 'copy machine' in desc:
        return "Copier"
    if 'fax' in desc:
        return "Fax machine"
    if any(kw in desc for kw in ['phone', 'telephone', 'cell', 'iphone', 'android', 'radio']):
        return "Telephone, cell phone or 2-way radio"

    # Vehicles
    # NOTE: Changed 'auto' to 'auto '/' auto'/'automobile' - 'auto' alone matches 'automatic', 'automation'
    if any(kw in desc for kw in ['auto ', ' auto', 'automobile', 'car ', ' car', 'sedan', 'coupe', 'suv', 'vehicle']):
        return "Auto"
    if any(kw in desc for kw in ['truck', 'pickup', 'f-150', 'f150', 'f-250', 'f250', 'silverado', 'ram ', 'tundra']):
        # Check weight if mentioned
        if any(kw in desc for kw in ['heavy', '13000', '13,000', 'hdtv']):
            return "Truck - heavy general purpose (unloaded weight >= 13,000 lbs.)"
        return "Truck - light general purpose (unloaded weight < 13,000 lbs)"
    if 'trailer' in desc:
        return "Trailer or trailer mounted container"
    if 'tractor' in desc:
        return "Tractor over the road"
    # NOTE: Changed from 'bus' to specific terms - 'bus' matches 'business'
    if any(kw in desc for kw in ['bus ', ' bus', 'buses', 'shuttle bus', 'school bus', 'transit bus']):
        return "Bus"

    # Rentals
    if 'appliance' in desc and 'rental' in final_cat:
        return "Appliance - rental"
    if 'carpet' in desc and 'rental' in final_cat:
        return "Carpeting - rental"

    # ===== 7-YEAR PROPERTY =====
    # Furniture
    if any(kw in desc for kw in ['furniture', 'desk', 'chair', 'table', 'cabinet', 'bookcase', 'credenza', 'sofa', 'couch']):
        if 'rental' in final_cat:
            return "Furniture or fixture - rental (includes - desks, files, safes - no structural components)"
        return "Furniture or fixture - nonrentals (includes - desks, files, safes - no structural components)"

    # Office equipment
    if any(kw in desc for kw in ['office equipment', 'office equip']):
        return "Office equipment (does not include communication equip in other classes)"

    # Machinery
    if any(kw in desc for kw in ['machine', 'machinery', 'equipment', 'equip']):
        return "Machinery, equipment or fixture (an asset whose life is affected by activity's economic type)"

    # Other 7-year
    if 'answering machine' in desc:
        return "Answering machine"
    if 'shredder' in desc:
        return "Shredder (includes shredders for DVDs, CDs, floppy disks, credit cards and paper)"
    if 'projector' in desc:
        return "Projector"
    if any(kw in desc for kw in ['camera', 'camcorder']):
        return "Camera or camcorder (includes digital or film cameras)"
    if any(kw in desc for kw in ['calculator', 'typewriter', 'adding machine']):
        return "Data handling equipment (includes only- typewriters, calculators, adding & accounting machines & duplicating equipment)"

    # ===== 15-YEAR PROPERTY =====
    if any(kw in desc for kw in ['land improvement', 'parking lot', 'fence', 'sidewalk', 'road', 'landscap', 'paving']):
        return "Land improvement (imprvmts directly related to land - sidewalks,roads,fences,bridges,landscapg,shrubbery,radio & tv transmittg towers)"
    if 'qip' in final_cat or 'qualified improvement' in final_cat:
        return "Improvement property (qualified)"
    if 'restaurant' in desc and 'improvement' in desc:
        return "Restaurant building improvement"
    if 'retail' in desc and 'improvement' in desc:
        return "Retail improvement"
    if any(kw in desc for kw in ['service station', 'car wash', 'carwash']):
        return "Service station building (including land improvements & carwash buildings)"

    # ===== 27.5-YEAR PROPERTY =====
    if 'residential' in final_cat or 'rental property' in desc or 'apartment' in desc:
        return "Real property (residential rental)"

    # ===== 39-YEAR PROPERTY =====
    if any(kw in desc for kw in ['building', 'warehouse', 'office building', 'commercial']):
        return "Real property (nonresidential)"
    if 'leasehold' in desc:
        if 'qualified' in final_cat or 'qip' in final_cat:
            return "Leasehold improvement (qualified)"
        return "Leasehold improvement (not qualified)"

    # ===== NON-DEPRECIABLE / SPECIAL =====
    if 'land' in final_cat and 'improvement' not in final_cat:
        return "Land"
    if 'software' in desc:
        return "Computer software"
    if 'goodwill' in desc or ('intangible' in final_cat and '197' in final_cat):
        return "Intangible asset (IRS Code Sec 197 - goodwill & other intangibles)"
    if 'startup' in desc or 'start-up' in desc or '195' in final_cat:
        return "Intangible asset (IRS Code Sec 195 - start-up expenses)"
    if 'organizational' in desc or 'org expense' in desc:
        if '1065' in desc or 'partnership' in desc:
            return "Intangible asset (IRS Code Sec 709 - organizational expenses (1065 only))"
        return "Intangible asset (IRS Code Sec 248 - organizational expenses (1120, 990))"
    if 'intangible' in final_cat:
        return "Intangible asset"

    # ===== TRANSPORTATION =====
    if 'air transport' in desc or 'aircraft' in desc:
        if 'commercial' in desc:
            return "Air transport (includes assets (except helicopters) used in commercial and contract carrying of passengers and freight by air)"
        return "Airplane (including helicopters (doesn't include airplanes used in commercial or contract carrying of passengers or freight))"
    if 'airplane' in desc or 'helicopter' in desc:
        return "Airplane (including helicopters (doesn't include airplanes used in commercial or contract carrying of passengers or freight))"
    if 'railroad' in desc or 'locomotive' in desc:
        return "Railroad cars & locomotives (except those owned by railroad transportation companies)"
    # NOTE: Changed 'ship' to ' ship' and 'ship ' - 'ship' alone matches 'shipment', 'shipping'
    if any(kw in desc for kw in ['boat', ' ship', 'ship ', 'vessel', 'water transport', 'yacht', 'barge', 'ferry']):
        return "Water transportation equipment (except those used in marine construction)"
    if 'billboard' in desc:
        return "Billboard"

    # ===== FALLBACK BY RECOVERY PERIOD =====
    try:
        life = float(recovery)
        if life == 5:
            return "Machinery, equipment or fixture (an asset whose life is affected by activity's economic type)"
        elif life == 7:
            return "Machinery, equipment or fixture (an asset whose life is affected by activity's economic type)"
        elif life == 15:
            return "Land improvement (imprvmts directly related to land - sidewalks,roads,fences,bridges,landscapg,shrubbery,radio & tv transmittg towers)"
        elif life == 27.5:
            return "Real property (residential rental)"
        elif life == 39:
            return "Real property (nonresidential)"
    except (ValueError, TypeError):
        pass

    # Default fallback
    return "Machinery, equipment or fixture (an asset whose life is affected by activity's economic type)"


# --------------------------------------------------------------
# CLASSIFICATION EXPLANATION
# --------------------------------------------------------------

def _classification_explanation(row):
    """Explain MACRS classification following IRS logic."""
    cat = row.get("Final Category", "")
    life = row.get("Recovery Period", "")
    src = str(row.get("Source", ""))  # Convert to string to handle NaN/float values

    base = f"Classified as {cat} ({life}-year) because "

    if src.startswith("gpt"):
        return base + "GPT MACRS reasoning using IRS mapping tables."
    if src == "rules":
        lc = str(cat).lower()
        if "land (non" in lc:
            return base + "land is non-depreciable under §167."
        if "improvement" in lc:
            return base + "IRS MACRS Table B-1 lists land improvements as 15-year property."
        if "qualified improvement" in lc:
            return base + "QIP is 15-year interior non-structural under §168(k)(3)."
        if "real property" in lc:
            return base + "nonresidential real property depreciates over 39-year MM."
        if "residential rental" in lc:
            return base + "residential rental is 27.5-year MM."
        if "machinery" in lc or "equipment" in lc:
            return base + "general machinery is 7-year GDS property."
        if "vehicle" in lc:
            return base + "vehicles fall under 5-year MACRS GDS."
        if "computer" in lc:
            return base + "computers are 5-year MACRS GDS."
        if "furniture" in lc:
            return base + "office furniture is 7-year MACRS GDS."
    return base + "fallback personal property rule."


def _macrs_reason_code(row) -> str:
    """Generate compact MACRS reason code for audit trail."""
    cat = str(row.get("Final Category", "")).lower()
    if "land (non" in cat: return "L0"
    if "land improvement" in cat: return "LI15"
    if "qualified improvement" in cat: return "QIP15"
    if "real property" in cat: return "RP39"
    if "residential rental" in cat: return "RR27"
    if "vehicle" in cat: return "V5"
    if "computer" in cat: return "C5"
    if "furniture" in cat: return "F7"
    if "equipment" in cat or "machinery" in cat: return "M7"
    return "PP7"


def _confidence_grade(row) -> str:
    """Convert confidence score to letter grade.

    Checks multiple possible confidence column names for compatibility
    with different classification sources (rule engine, GPT, etc.).
    """
    # Check multiple possible confidence column names
    conf = None
    for col_name in ["Rule Confidence", "Confidence", "Classification Confidence", "GPT Confidence"]:
        val = row.get(col_name, None)
        if val is not None and pd.notna(val):
            conf = val
            break

    if conf is None:
        return "Unknown"

    try:
        c = float(conf)
        # Normalize if stored as percentage (e.g., 85 instead of 0.85)
        if c > 1:
            c = c / 100.0
    except (ValueError, TypeError):
        return "Unknown"

    if c >= 0.90: return "A"
    if c >= 0.75: return "B"
    if c >= 0.60: return "C"
    return "D"


# --------------------------------------------------------------
# AUDIT TRAIL WITH SHA256 INTEGRITY HASH
# --------------------------------------------------------------

def _audit_fields(row) -> dict:
    """
    Generate comprehensive audit trail with classification integrity hash.

    Includes:
    - Source tracking (Rule Engine vs GPT Classifier)
    - Rule triggers that determined classification
    - Warnings and validation issues
    - SHA256 hash of classification parameters for integrity verification
    """
    audit = {}

    # Source
    src = str(row.get("Source", ""))
    if src == "rules":
        audit["AuditSource"] = "Rule Engine"
    elif src.startswith("gpt"):
        audit["AuditSource"] = "GPT Classifier"
    else:
        audit["AuditSource"] = "Client / Fallback"

    # Rule trigger summary
    cat = str(row.get("Final Category", "")).lower()
    if "land (non" in cat:
        audit["AuditRuleTriggers"] = "Land → Non-Depreciable"
    elif "land improvement" in cat:
        audit["AuditRuleTriggers"] = "Land Improvement (15-yr)"
    elif "qualified improvement" in cat:
        audit["AuditRuleTriggers"] = "QIP → 15-year"
    elif "real property" in cat:
        audit["AuditRuleTriggers"] = "Nonresidential Real Property (39-yr)"
    elif "residential rental" in cat:
        audit["AuditRuleTriggers"] = "Residential Rental (27.5-yr)"
    elif "vehicle" in cat:
        audit["AuditRuleTriggers"] = "Vehicle (5-yr)"
    elif "equipment" in cat or "machinery" in cat:
        audit["AuditRuleTriggers"] = "7-year Machinery"
    else:
        audit["AuditRuleTriggers"] = "Personal Property Fallback"

    # Warnings
    warnings = []
    if row.get("NBV_Reco") == "CHECK":
        warnings.append("NBV out of balance")
    if row.get("Desc_TypoFlag") == "YES":
        warnings.append("Description corrected for typos")
    if row.get("Cat_TypoFlag") == "YES":
        warnings.append("Client category corrected")
    if row.get("Uses ADS"):
        warnings.append("ADS required per IRC §168(g)")

    audit["AuditWarnings"] = "; ".join(warnings) if warnings else "None"

    # Hash for classification integrity
    hash_input = (
        f"{row.get('Asset #','')}|"
        f"{row.get('Final Category','')}|"
        f"{row.get('Tax Life','')}|"
        f"{row.get('Tax Method','')}|"
        f"{row.get('Convention','')}"
    )
    audit["ClassificationHash"] = hashlib.sha256(hash_input.encode("utf-8")).hexdigest()

    audit["AuditTimestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return audit


# --------------------------------------------------------------
# IRC §280F LUXURY AUTO DEPRECIATION LIMITS
# NOTE: Now loaded from tax_year_config.py for flexibility
# --------------------------------------------------------------


def _is_passenger_auto(row) -> bool:
    """
    Check if asset is a passenger automobile subject to §280F limits.

    Returns True for passenger vehicles, False for heavy trucks/SUVs >6,000 lbs.
    """
    final_class = str(row.get("Final Category", "")).lower()
    desc = str(row.get("Description", "")).lower()

    # Check if classified as passenger auto
    if "passenger" in final_class and "auto" in final_class:
        return True

    # Also check description for typical passenger vehicles
    # NOTE: Use space-bounded terms for short words to avoid false matches
    # - "car" alone matches "card", "carpet" - use "car ", " car"
    # - "van" alone matches "advantage", "canvas" - use " van", "van "
    passenger_indicators = ["car ", " car", "sedan", "suv", "crossover", " van", "van ", "minivan"]
    heavy_indicators = ["heavy duty", "f-250", "f-350", "2500", "3500", "commercial truck"]

    # If heavy truck indicators, not subject to passenger limits
    if any(h in desc for h in heavy_indicators):
        return False

    # If passenger indicators, subject to limits
    if any(p in desc for p in passenger_indicators):
        return True

    return False


def _is_heavy_suv(row) -> bool:
    """
    Check if vehicle is a heavy SUV/truck (>6,000 lbs GVWR).

    Heavy SUVs are:
    - NOT subject to IRC §280F luxury auto limits
    - BUT subject to special Section 179 limit (currently $28,900 for 2024)

    Per IRC §179(b)(5), heavy SUVs have a reduced Section 179 limit.

    Returns:
        True if heavy SUV/truck (>6,000 lbs)
        False otherwise
    """
    final_class = str(row.get("Final Category", "")).lower()
    desc = str(row.get("Description", "")).lower()

    # Must be in Trucks & Trailers category or have truck/suv indicators
    is_truck_category = "truck" in final_class or "trailer" in final_class

    # Check for heavy SUV/truck indicators in description
    heavy_suv_indicators = [
        # Explicit weight indicators
        "gvwr 6", "6000 lbs", "6,000 lbs", "6000lbs", "6,000lbs",
        "gross vehicle weight", "over 6000", "over 6,000",

        # Common heavy SUV models
        "suburban", "tahoe", "yukon", "expedition", "navigator",
        "escalade", "land cruiser", "sequoia", "armada",
        "gx 460", "lx 570", "lx 600", "gx 550",
        "qx80", "range rover", "land rover defender",

        # Heavy duty trucks
        "f-250", "f-350", "f250", "f350",
        "2500", "3500", "2500hd", "3500hd",
        "heavy duty", "hd", "crew cab diesel",
        "dually", "super duty",

        # Commercial trucks
        "commercial truck", "work truck", "cargo van",
        "sprinter", "transit 250", "transit 350",
        "promaster 2500", "promaster 3500",
    ]

    # Check if description contains heavy SUV/truck indicators
    has_heavy_indicator = any(ind in desc for ind in heavy_suv_indicators)

    # Light truck/SUV indicators (NOT heavy)
    light_indicators = [
        "f-150", "f150", "1500", "tacoma", "ranger",
        "colorado", "canyon", "ridgeline", "maverick",
        "cr-v", "rav4", "highlander", "pilot", "pathfinder"
    ]

    has_light_indicator = any(ind in desc for ind in light_indicators)

    # Heavy SUV if:
    # 1. In truck category AND has heavy indicators AND NOT light indicators
    # 2. OR has explicit GVWR >6,000 lbs
    if "gvwr" in desc or "gross vehicle weight" in desc:
        return True  # Explicit weight spec

    if is_truck_category and has_heavy_indicator and not has_light_indicator:
        return True

    return False


def _apply_luxury_auto_caps(row, sec179: float, bonus: float, tax_year: int) -> tuple[float, float, str]:
    """
    Apply IRC §280F luxury automobile depreciation limitations.

    IRC §280F imposes strict depreciation limits on passenger automobiles to prevent
    excessive deductions on luxury vehicles. Year 1 limits (2024):
    - With bonus depreciation: $20,200 ($12,200 base + $8,000 bonus)
    - Without bonus: $12,200

    PRIORITY ORDER FOR CAPPING:
    The code prioritizes Bonus Depreciation first, then Section 179:
    1. Apply up to full bonus amount (up to limit)
    2. Apply Section 179 for any remaining limit

    RATIONALE:
    - Bonus depreciation is automatic (no election required)
    - Section 179 requires election and has additional limitations
    - Prioritizing bonus maximizes first-year deduction (bonus is lost if not used)
    - Section 179 can potentially be carried forward if disallowed

    ALTERNATIVE APPROACH (for consideration):
    Some tax software prioritizes Section 179 first on the theory that expensing
    should take precedence over depreciation. However, IRS guidance in Publication 946
    treats them as a combined total for §280F purposes without specifying priority.

    VERIFICATION STATUS:
    ⚠️  This priority order (Bonus first, then Section 179) is a reasonable
    interpretation but should be verified with:
    - IRS Publication 946 (How to Depreciate Property)
    - Form 4562 instructions
    - Tax advisor for specific client situations

    If tax advisor recommends different priority, this function can be easily
    modified to change the order.

    Args:
        row: Asset row with classification data
        sec179: Calculated Section 179 amount (before any caps)
        bonus: Calculated bonus depreciation amount (before any caps)
        tax_year: Current tax year

    Returns:
        Tuple of (capped_sec179, capped_bonus, notes)
        - capped_sec179: Section 179 amount after applying §280F cap
        - capped_bonus: Bonus amount after applying §280F cap
        - notes: Explanation of cap applied (empty string if no cap)

    Example:
        Luxury car cost $50,000, Section 179 = $25,000, Bonus = $20,000
        Total requested = $45,000, but §280F limit = $20,200

        With current priority (Bonus first):
        - Bonus: min($20,000, $20,200) = $20,000 ✓
        - Section 179: min($25,000, $20,200 - $20,000) = $200
        - Total: $20,200 (at limit) ✓

    See Also:
        - IRC §280F(a): Luxury automobile depreciation limits
        - IRS Publication 946, Table 1-1: Dollar limits for passenger automobiles
        - Form 4562, Part V: Listed Property
    """
    if not _is_passenger_auto(row):
        return sec179, bonus, ""

    cost = float(row.get("Cost") or 0.0)
    in_service = row.get("In Service Date")

    # Only apply to current year additions
    if not _is_current_year(in_service, tax_year):
        return sec179, bonus, ""

    # Calculate total year 1 depreciation requested
    total_requested = sec179 + bonus

    # Get limits for the tax year (dynamically loaded from tax_year_config.py)
    limits = get_luxury_auto_limits(tax_year, asset_year=1)

    # Determine applicable limit based on whether bonus is being claimed
    if bonus > 0:
        year_1_limit = limits["year_1_with_bonus"]
        limit_type = "with bonus"
    else:
        year_1_limit = limits["year_1_without_bonus"]
        limit_type = "without bonus"

    # If requested amount exceeds limit, apply cap
    if total_requested > year_1_limit:
        # PRIORITY ORDER: Bonus first, then Section 179
        # (See docstring for rationale and verification status)
        capped_bonus = min(bonus, year_1_limit)
        remaining = year_1_limit - capped_bonus
        capped_sec179 = min(sec179, remaining)

        notes = (
            f"IRC §280F luxury auto limit applied: "
            f"Year 1 {limit_type} = ${year_1_limit:,.0f} "
            f"(requested ${total_requested:,.0f}, "
            f"excess ${total_requested - year_1_limit:,.0f} not allowed)"
        )

        return capped_sec179, capped_bonus, notes

    return sec179, bonus, ""


# --------------------------------------------------------------
# Helpers
# --------------------------------------------------------------

def _pick(col1, col2):
    """Return first non-null value."""
    if pd.notna(col1) and col1 not in ("", None):
        return col1
    return col2


def _is_current_year(date_val: Optional[date], tax_year: int) -> bool:
    """Check if placed-in-service date is in current tax year."""
    if not isinstance(date_val, (datetime, date)):
        return False
    return date_val.year == tax_year


def _is_disposal(row) -> bool:
    """Determine if row represents a disposal."""
    # Check Transaction Type first (more reliable), then Sheet Role
    raw = str(row.get("Transaction Type", "")).lower()
    if raw == "":
        raw = str(row.get("Sheet Role", "")).lower()
    # NOTE: These are checked against Transaction Type/Sheet Role (user-defined),
    # not descriptions, so "sold" is safe here
    return any(x in raw for x in ["dispos", "disposal", "disposed", "sold", "retire"])


def _is_transfer(row) -> bool:
    """Determine if row represents a transfer."""
    # Check Transaction Type first (more reliable), then Sheet Role
    raw = str(row.get("Transaction Type", "")).lower()
    if raw == "":
        raw = str(row.get("Sheet Role", "")).lower()
    return any(x in raw for x in ["transfer", "xfer", "reclass"])


def _determine_asset_type(row) -> str:
    """
    Determine asset type for FA CS folder classification.

    Based on user testing (2025): FA CS puts imported assets into folders.
    "Business" folder is where most business assets should go.
    "Miscellaneous" is the default if Asset Type not recognized.

    Returns:
        "Business" for all regular business assets
        Empty string for disposals/transfers
    """
    # Skip disposals and transfers
    if _is_disposal(row) or _is_transfer(row):
        return ""

    # FA CS expects "Business" for the Business folder
    # Using specific types like "Equipment", "Vehicles" resulted in
    # assets going to "Miscellaneous" folder instead
    return "Business"


def _convert_method_to_fa_cs_format(uses_ads: bool) -> str:
    """
    Convert internal method format to Fixed Asset CS import format.

    Based on user testing feedback (2025-01-20):
    FA CS actually expects just "MACRS" for regular depreciation,
    not "MACRS GDS" or "MACRS ADS".

    Args:
        uses_ads: Whether asset uses Alternative Depreciation System

    Returns:
        "MACRS" (for both GDS and ADS)
    """
    # NOTE: Originally tried "MACRS GDS" and "MACRS ADS" but user testing
    # showed FA CS only accepts "MACRS" during import process
    return "MACRS"


# --------------------------------------------------------------
# Main export builder
# --------------------------------------------------------------

def build_fa(
    df: pd.DataFrame,
    tax_year: int,
    strategy: str,
    taxable_income: float,
    use_acq_if_missing: bool = True,
    de_minimis_limit: float = 0.0,  # Set to 2500 or 5000 to enable de minimis safe harbor
    section_179_carryforward_from_prior_year: float = 0.0,  # Prior year Section 179 carryforward (from Form 4562)
    asset_number_start: int = 1,  # Starting Asset # for FA CS (e.g., 1001 to continue from existing assets)
) -> pd.DataFrame:

    # ----------------------------------------------------------------------
    # CRITICAL: Validate non-empty DataFrame
    # ----------------------------------------------------------------------
    if df is None or df.empty or len(df) == 0:
        raise ValueError(
            "Cannot process empty asset list. Please provide at least one asset record. "
            "If you're importing from Excel, verify the file contains data rows."
        )

    df = df.copy()

    # ----------------------------------------------------------------------
    # DATA VALIDATION - Catch errors before processing
    # ----------------------------------------------------------------------
    validation_errors, should_stop = validate_asset_data(df, tax_year)

    if validation_errors:
        print(f"\n{'='*70}")
        print(f"DATA VALIDATION RESULTS: {len(validation_errors)} issues found")
        print(f"{'='*70}")

        # Group by severity
        critical = [e for e in validation_errors if e.severity == "CRITICAL"]
        errors = [e for e in validation_errors if e.severity == "ERROR"]
        warnings = [e for e in validation_errors if e.severity == "WARNING"]

        if critical:
            print(f"\n🔴 CRITICAL ({len(critical)}):")
            for e in critical[:5]:  # Show first 5
                print(f"  {e}")

        if errors:
            print(f"\n❌ ERRORS ({len(errors)}):")
            for e in errors[:10]:  # Show first 10
                print(f"  {e}")

        if warnings:
            print(f"\n⚠️  WARNINGS ({len(warnings)}):")
            for e in warnings[:10]:  # Show first 10
                print(f"  {e}")

        print(f"\n{'='*70}\n")

        # IMPORTANT: Validation errors are ADVISORY only
        # Processing continues regardless - CPA reviews and decides what to do
        # The system does NOT exclude assets without human CPA approval
        if should_stop:
            # Log warning but DO NOT stop - let CPA decide
            print("⚠️  WARNING: Critical validation errors found but processing continues.")
            print("    CPA must review flagged assets before finalizing export.")
            # Was: raise ValueError(...) - removed to allow CPA review

    # ============================================================================
    # CRITICAL FIX: TRANSACTION TYPE CLASSIFICATION
    # ============================================================================
    # ISSUE: System was defaulting ALL assets to "addition" even if they were
    # placed in service in prior years (e.g., 2020 assets being treated as 2024 additions).
    #
    # This caused:
    # ❌ Section 179 claimed on old assets (NOT ALLOWED per IRC §179)
    # ❌ Bonus depreciation on old assets (NOT ALLOWED per IRC §168(k))
    # ❌ Massive tax deduction overstatement
    # ❌ IRS audit risk
    #
    # SOLUTION: Properly classify based on in-service date vs tax year:
    # ✅ Current Year Addition: In-service date = current tax year
    # ✅ Existing Asset: In-service date < current tax year
    # ✅ Disposal: Transaction type or disposal date indicates sale
    # ✅ Transfer: Transaction type indicates transfer/reclass

    df = classify_all_transactions(df, tax_year, verbose=True)

    # Validate that no existing assets are being treated as additions
    is_valid, classification_errors = validate_transaction_classification(df, tax_year)

    if not is_valid:
        raise ValueError(
            f"CRITICAL: Transaction classification errors detected. "
            f"Found {len(classification_errors)} assets incorrectly classified. "
            f"This would cause incorrect Section 179/Bonus claims. "
            f"See error details above."
        )

    # ----------------------------------------------------------------------
    # Basic cleanup
    # ----------------------------------------------------------------------
    if "Asset ID" not in df.columns:
        df["Asset ID"] = df.index.astype(str)

    # Parse cost
    if "Cost" in df.columns:
        df["Cost"] = df["Cost"].apply(parse_number)
    else:
        df["Cost"] = 0.0

    # Parse dates - safely handle missing columns
    if "Acquisition Date" in df.columns:
        df["Acquisition Date"] = df["Acquisition Date"].apply(parse_date)
    else:
        df["Acquisition Date"] = None

    if "In Service Date" in df.columns:
        df["In Service Date"] = df["In Service Date"].apply(parse_date)
    else:
        df["In Service Date"] = None

    if use_acq_if_missing:
        df["In Service Date"] = df.apply(
            lambda r: r["In Service Date"] or r["Acquisition Date"],
            axis=1
        )

    # ----------------------------------------------------------------------
    # TYPO CORRECTION - Clean up data before classification
    # ----------------------------------------------------------------------
    # Fix common typos in descriptions and client categories
    # Tracks what was corrected for audit trail
    if "Description" in df.columns:
        df["Desc_Corrected"], df["Desc_TypoFlag"], df["Desc_TypoNote"] = zip(
            *df["Description"].apply(_fix_description)
        )
        df["Description"] = df["Desc_Corrected"]

        # Convert boolean flags to YES/NO for clarity
        df["Desc_TypoFlag"] = df["Desc_TypoFlag"].apply(lambda x: "YES" if x else "NO")

        typo_count = (df["Desc_TypoFlag"] == "YES").sum()
        if typo_count > 0:
            print(f"\n✓ Corrected {typo_count} description typos")

    if "Client Category" in df.columns:
        df["Cat_Corrected"], df["Cat_TypoFlag"], df["Cat_TypoNote"] = zip(
            *df["Client Category"].apply(_fix_client_category)
        )
        # Store original before correction
        if "Client Category Original" not in df.columns:
            df["Client Category Original"] = df["Client Category"]
        df["Client Category"] = df["Cat_Corrected"]

        # Convert boolean flags to YES/NO for clarity
        df["Cat_TypoFlag"] = df["Cat_TypoFlag"].apply(lambda x: "YES" if x else "NO")

        cat_typo_count = (df["Cat_TypoFlag"] == "YES").sum()
        if cat_typo_count > 0:
            print(f"✓ Corrected {cat_typo_count} category typos")

    # ============================================================================
    # TIER 3: DE MINIMIS SAFE HARBOR (Rev. Proc. 2015-20)
    # ============================================================================
    # Election to immediately expense items ≤$2,500 (or ≤$5,000 with AFS)
    # - Does NOT count against Section 179 limits
    # - Reduces depreciation tracking burden
    # - Separate from bonus depreciation
    #
    # NOTE: This is an ELECTION that taxpayers can make. We apply it if the
    # de_minimis_limit parameter is set.

    if de_minimis_limit and de_minimis_limit > 0:
        de_minimis_expensed = []

        for idx, row in df.iterrows():
            cost = float(row.get("Cost") or 0.0)
            trans_type = str(row.get("Transaction Type", ""))

            # CRITICAL: Only for CURRENT YEAR additions (not existing assets)
            # De minimis safe harbor only applies to property placed in service in current year
            if "Current Year Addition" in trans_type:
                if 0 < cost <= de_minimis_limit:
                    # Mark as de minimis expensed
                    de_minimis_expensed.append(cost)
                    # Set cost to 0 so it doesn't get depreciation calculated
                    df.at[idx, "De Minimis Expensed"] = cost
                    df.at[idx, "Cost"] = 0.0
                else:
                    de_minimis_expensed.append(0.0)
            else:
                de_minimis_expensed.append(0.0)

        df["De Minimis Expensed"] = de_minimis_expensed

        # Print summary if any de minimis items
        total_de_minimis = sum(de_minimis_expensed)
        count_de_minimis = sum(1 for x in de_minimis_expensed if x > 0)

        if total_de_minimis > 0:
            print("\n" + "=" * 80)
            print(f"DE MINIMIS SAFE HARBOR - Rev. Proc. 2015-20")
            print("=" * 80)
            print(f"Limit: ${de_minimis_limit:,.2f} per item")
            print(f"Items expensed immediately: {count_de_minimis}")
            print(f"Total de minimis expensed: ${total_de_minimis:,.2f}")
            print(f"(Does NOT count against Section 179 limits)")
            print("=" * 80 + "\n")
    else:
        # No de minimis election
        df["De Minimis Expensed"] = 0.0

    # ----------------------------------------------------------------------
    # Depreciation classification only applies to ADDITIONS
    # ----------------------------------------------------------------------

    # Validate tax year configuration and warn if using estimated values
    config_validation = validate_tax_year_config(tax_year)
    for warning in config_validation.warnings:
        print(f"TAX CONFIG WARNING: {warning}")

    # ============================================================================
    # TIER 3: MID-QUARTER CONVENTION DETECTION (IRC §168(d)(3))
    # ============================================================================
    # CRITICAL: If >40% of property placed in service in Q4, must use MQ convention
    # This is one of the most common IRS audit adjustments

    global_convention, mq_details = detect_mid_quarter_convention(df, tax_year, verbose=True)

    # Store mid-quarter test results for later use
    df.attrs["mid_quarter_test"] = mq_details
    df.attrs["global_convention"] = global_convention

    # Apply mid-quarter convention to all assets and add quarter column
    conventions = []
    quarters = []

    for _, row in df.iterrows():
        category = str(row.get("Final Category", ""))
        in_service = row.get("In Service Date")

        # Check if real property (always uses MM)
        is_real = any(x in category for x in ["Residential", "Nonresidential", "Real Property", "Building"])

        if is_real:
            conventions.append("MM")  # Mid-month for real property
            quarters.append(None)  # No quarter needed for MM
        else:
            # Personal property uses global convention (HY or MQ)
            conventions.append(global_convention)

            # If MQ, store which quarter for depreciation calculation
            # Must check for NaT before passing to get_quarter
            if global_convention == "MQ" and in_service is not None and not pd.isna(in_service):
                quarter = get_quarter(in_service)
                quarters.append(quarter)
            else:
                quarters.append(None)

    df["Convention"] = conventions
    df["Quarter"] = quarters  # Used for MQ depreciation calculations

    # Get tax year configuration
    section_179_config = get_section_179_limits(tax_year)
    # Note: bonus_percentage is now calculated per asset based on acquisition/in-service dates
    # to handle OBBB Act 100% bonus eligibility

    # Calculate total Section 179-eligible property cost for phase-out
    # CRITICAL: Only count CURRENT YEAR ADDITIONS (not existing assets)
    total_179_eligible_cost = 0.0
    for _, row in df.iterrows():
        trans_type = str(row.get("Transaction Type", ""))

        # Only current year additions are eligible for Section 179
        if "Current Year Addition" in trans_type:
            cost = float(row.get("Cost") or 0.0)
            is_qip = row.get("qip", False) or "QIP" in str(row.get("Final Category", ""))

            if cost > 0 and not is_qip:
                total_179_eligible_cost += cost

    # Apply IRC §179(b)(2) phase-out
    # Dollar limit is reduced (but not below zero) by the amount by which the cost of
    # section 179 property placed in service exceeds the threshold
    phaseout_reduction = max(0.0, total_179_eligible_cost - section_179_config["phaseout_threshold"])
    section_179_dollar_limit = max(0.0, section_179_config["max_deduction"] - phaseout_reduction)

    # Apply business income limitation
    section_179_effective_limit = min(section_179_dollar_limit, max(float(taxable_income), 0.0))

    section179_amounts = []
    bonus_amounts = []
    bonus_percentages_used = []  # Track which bonus % was applied (for audit trail)
    luxury_auto_notes = []
    ads_flags = []  # Track which assets use ADS
    remaining_179 = section_179_effective_limit

    for _, row in df.iterrows():

        # Skip depreciation for disposals and transfers
        if _is_disposal(row) or _is_transfer(row):
            section179_amounts.append(0.0)
            bonus_amounts.append(0.0)
            bonus_percentages_used.append(0.0)
            luxury_auto_notes.append("")
            ads_flags.append(False)
            continue

        cost = float(row.get("Cost") or 0.0)
        in_service = row.get("In Service Date")
        acquisition = row.get("Acquisition Date")
        is_current = _is_current_year(in_service, tax_year)

        # ============================================================================
        # CRITICAL FIX: Check transaction type for Section 179/Bonus eligibility
        # ============================================================================
        # ONLY "Current Year Addition" assets are eligible for Section 179/Bonus
        # "Existing Asset" (prior year assets) are NOT eligible per IRC §179 and §168(k)
        trans_type = str(row.get("Transaction Type", ""))
        is_current_year_addition = "Current Year Addition" in trans_type

        # ============================================================================
        # TIER 2: ADS (Alternative Depreciation System) Detection
        # ============================================================================
        # CRITICAL: Check if asset requires ADS per IRC §168(g)
        # ADS required for: listed property ≤50% business use, tax-exempt property, etc.

        uses_ads, ads_reason = should_use_ads(row.to_dict())

        # Validate business use for listed property (IRC §280F)
        is_section179_eligible = True
        bonus_eligible = True

        # ============================================================================
        # TIER 3: QIP Section 179 Eligibility (OBBB Act vs Pre-OBBB)
        # ============================================================================
        # CRITICAL TAX COMPLIANCE CHANGE (OBBB Act - July 4, 2025):
        # - Pre-OBBB (before 1/1/2025): QIP NOT eligible for Section 179 per IRC §179(d)(1)
        # - OBBB Act (1/1/2025+): QIP IS eligible for Section 179 (subject to $2.5M limit)
        #
        # NOTE: Buildings, land, and land improvements still NOT eligible

        is_qip = row.get("qip", False) or "QIP" in str(row.get("Final Category", ""))
        final_category = str(row.get("Final Category", ""))

        if is_qip:
            # Check if placed in service after 12/31/2024 (OBBB effective date)
            # Must check for NaT before comparing dates
            if in_service is not None and not pd.isna(in_service):
                # Convert pandas Timestamp to date if necessary
                in_service_date = in_service.date() if hasattr(in_service, 'date') else in_service
                if in_service_date > date(2024, 12, 31):
                    # OBBB Act: QIP IS eligible for Section 179
                    is_section179_eligible = True
                else:
                    # Pre-OBBB: QIP NOT eligible for Section 179
                    is_section179_eligible = False
            else:
                # No in-service date: Pre-OBBB assumption
                is_section179_eligible = False

        # Buildings, land improvements, and land are NEVER eligible (even under OBBB)
        if any(x in final_category for x in ["Nonresidential Real Property", "Residential", "Land"]):
            if "Improvement" not in final_category and "QIP" not in final_category:
                is_section179_eligible = False  # Building or land = no Section 179

        # Check listed property business use requirements
        is_section179_eligible, bonus_eligible, business_use_warnings = validate_business_use_for_incentives(
            row.to_dict(),
            allow_section_179=is_section179_eligible,
            allow_bonus=True
        )

        # Calculate bonus percentage for this specific asset
        # OBBB Act: 100% if acquired AND placed in service after 1/19/2025
        # Otherwise: TCJA phase-down (80% for 2024-2025, etc.)
        asset_bonus_pct = get_bonus_percentage(tax_year, acquisition, in_service)

        sec179 = 0.0
        bonus = 0.0
        auto_note = ""

        # ============================================================================
        # TIER 2: ADS Property - NO Section 179, NO Bonus (IRC §168(g))
        # ============================================================================
        if uses_ads:
            # ADS property is NOT eligible for Section 179 or bonus depreciation
            sec179 = 0.0
            bonus = 0.0
            asset_bonus_pct = 0.0

            # Add ADS note
            auto_note = f"ADS REQUIRED: {ads_reason}. No Section 179/bonus allowed per IRC §168(g)"

        # ============================================================================
        # Standard MACRS with Section 179/Bonus (if eligible)
        # ============================================================================
        # CRITICAL: Only apply to CURRENT YEAR ADDITIONS (not existing assets)
        elif cost > 0 and is_current_year_addition:

            # Get strategy configuration
            strat = get_strategy(strategy)

            if strat.apply_179:
                # Section 179 up to limit (ONLY for eligible property)
                if remaining_179 > 0 and is_section179_eligible:
                    sec179 = min(cost, remaining_179)

                    # ============================================================================
                    # TIER 3: Heavy SUV Special Section 179 Limit (IRC §179(b)(5))
                    # ============================================================================
                    # Heavy SUVs (>6,000 lbs GVWR) are NOT subject to luxury auto caps
                    # BUT have a special reduced Section 179 limit ($28,900 for 2024)
                    if _is_heavy_suv(row):
                        heavy_suv_limit = get_heavy_suv_179_limit(tax_year)
                        if sec179 > heavy_suv_limit:
                            sec179 = heavy_suv_limit
                            if auto_note:
                                auto_note += " | "
                            auto_note = f"Heavy SUV §179 limit applied: ${heavy_suv_limit:,.0f} (IRC §179(b)(5))"

                    remaining_179 -= sec179

            if strat.apply_bonus:
                # Bonus for remainder (ONLY if eligible)
                if bonus_eligible:
                    # CRITICAL: Apply asset-specific bonus percentage
                    # OBBB Act: 100% for property acquired AND placed in service after 1/19/2025
                    # TCJA phase-down: 2024-2025: 80%, 2026: 60%, 2027: 40%, 2028: 20%, 2029+: 0%
                    bonus = max(cost - sec179, 0.0) * asset_bonus_pct
                else:
                    bonus = 0.0
                    asset_bonus_pct = 0.0
            else:
                # No bonus in conservative strategy
                bonus = 0.0
                asset_bonus_pct = 0.0

            # CRITICAL: Apply IRC §280F luxury auto limits (for non-ADS property)
            sec179, bonus, auto_note = _apply_luxury_auto_caps(row, sec179, bonus, tax_year)

        # ============================================================================
        # Compliance Notes
        # ============================================================================

        # Add business use warnings
        for warning in business_use_warnings:
            if auto_note:
                auto_note += " | "
            auto_note += warning

        # Add note if QIP was excluded from Section 179
        if is_qip and cost > 0 and is_current and strategy == "Aggressive (179 + Bonus)" and not uses_ads:
            if auto_note:
                auto_note += " | "
            auto_note += "QIP not eligible for §179 per IRC §179(d)(1)"

        # Add note if OBBB 100% bonus applied
        # Use tolerance-based comparison for float values
        if abs(asset_bonus_pct - 1.0) < FLOAT_TOLERANCE and bonus > 0:
            if auto_note:
                auto_note += " | "
            auto_note += f"OBBB Act: 100% bonus (acquired {acquisition}, in-service {in_service})"

        section179_amounts.append(sec179)
        bonus_amounts.append(bonus)
        bonus_percentages_used.append(asset_bonus_pct)
        luxury_auto_notes.append(auto_note)
        ads_flags.append(uses_ads)

    df["Section 179 Amount"] = section179_amounts
    df["Bonus Amount"] = bonus_amounts
    df["Bonus Percentage Used"] = bonus_percentages_used  # Track OBBB vs TCJA bonus %
    df["Uses ADS"] = ads_flags  # Track which assets use Alternative Depreciation System
    df["Auto Limit Notes"] = luxury_auto_notes

    # ============================================================================
    # PHASE 4: MACRS DEPRECIATION CALCULATION (IRS Publication 946 Tables)
    # ============================================================================
    # For CURRENT YEAR ADDITIONS: Calculate Year 1 MACRS depreciation
    # For EXISTING ASSETS: Calculate current year depreciation based on depreciation year
    # Basis for MACRS = Cost - Section 179 - Bonus Depreciation
    #
    # CRITICAL: For existing assets, we need the ORIGINAL depreciable basis
    # (Cost - Prior Section 179 - Prior Bonus), not the current cost.
    # If prior 179/bonus columns exist, use them; otherwise warn user.

    macrs_year1_depreciation = []
    depreciable_basis_list = []
    existing_asset_basis_warnings = []

    # Check if we have prior depreciation info columns
    has_prior_179 = "Prior Section 179" in df.columns or "Prior Sec 179" in df.columns
    has_prior_bonus = "Prior Bonus" in df.columns or "Prior Bonus Depreciation" in df.columns
    has_original_basis = "Original Depreciable Basis" in df.columns or "Tax Basis" in df.columns

    for idx, row in df.iterrows():
        cost = float(row.get("Cost") or 0.0)
        sec179 = float(row.get("Section 179 Amount") or 0.0)
        bonus = float(row.get("Bonus Amount") or 0.0)
        transaction_type = str(row.get("Transaction Type", ""))
        in_service_date = row.get("In Service Date")
        is_existing_asset = "Existing Asset" in transaction_type

        # ============================================================================
        # CRITICAL FIX: Calculate correct depreciable basis for existing assets
        # ============================================================================
        # For CURRENT YEAR ADDITIONS: depreciable_basis = cost - sec179 - bonus
        # For EXISTING ASSETS: need ORIGINAL depreciable basis from when first placed in service

        if is_existing_asset:
            # Try to get original depreciable basis from input data
            original_basis = None

            # Priority 1: Explicit "Original Depreciable Basis" or "Tax Basis" column
            if has_original_basis:
                original_basis = float(row.get("Original Depreciable Basis") or row.get("Tax Basis") or 0.0)

            # Priority 2: Calculate from Prior Section 179 and Prior Bonus columns
            elif has_prior_179 or has_prior_bonus:
                prior_179 = float(row.get("Prior Section 179") or row.get("Prior Sec 179") or 0.0)
                prior_bonus = float(row.get("Prior Bonus") or row.get("Prior Bonus Depreciation") or 0.0)
                original_basis = max(cost - prior_179 - prior_bonus, 0.0)

            # Priority 3: Infer from accumulated depreciation (if recovery period known)
            # If accumulated depreciation exists, we can estimate original basis
            elif "Accumulated Depreciation" in df.columns:
                accum_depr = float(row.get("Accumulated Depreciation") or 0.0)

                # If accum_depr is close to cost, asset likely took 100% bonus
                # This is an ESTIMATE - flag for review
                if accum_depr > 0:
                    # Assume the asset's depreciable basis was set such that
                    # accumulated MACRS depreciation matches accumulated column
                    # This is imperfect but better than using full cost
                    recovery_period_check = row.get("Recovery Period") or row.get("MACRS Life")

                    if recovery_period_check and in_service_date and not pd.isna(in_service_date):
                        try:
                            if isinstance(in_service_date, date):
                                in_service_year = in_service_date.year
                            elif hasattr(in_service_date, 'year'):
                                in_service_year = in_service_date.year
                            else:
                                in_service_year = int(str(in_service_date)[:4])

                            years_depreciated = tax_year - in_service_year

                            # Get cumulative MACRS percentage for years depreciated
                            try:
                                recovery_period_int = int(float(recovery_period_check))
                                method = row.get("Method", "200DB")
                                convention = row.get("Convention", "HY")
                                table = get_macrs_table(recovery_period_int, method, convention)

                                if years_depreciated > 0 and years_depreciated <= len(table):
                                    cumulative_pct = sum(table[:years_depreciated])
                                    if cumulative_pct > 0:
                                        # Back-calculate original basis: basis = accum_depr / cumulative_pct
                                        estimated_basis = accum_depr / cumulative_pct
                                        # Sanity check: basis should not exceed cost
                                        if estimated_basis <= cost * 1.01:  # Allow 1% tolerance
                                            original_basis = estimated_basis
                                            existing_asset_basis_warnings.append(
                                                f"Asset {row.get('Asset ID', idx)}: Estimated basis ${original_basis:,.2f} "
                                                f"from accumulated depreciation (review recommended)"
                                            )
                            except (ValueError, TypeError, IndexError) as e:
                                # Log but continue - basis estimation is best-effort
                                existing_asset_basis_warnings.append(
                                    f"Asset {row.get('Asset ID', idx)}: Could not estimate basis from depreciation table ({type(e).__name__})"
                                )
                        except (ValueError, TypeError) as e:
                            # Log but continue - date parsing failed
                            existing_asset_basis_warnings.append(
                                f"Asset {row.get('Asset ID', idx)}: Could not parse in-service date for basis estimation ({type(e).__name__})"
                            )

            # Fallback: Use full cost but warn (this may overstate depreciation!)
            if original_basis is None or original_basis <= 0:
                original_basis = cost
                if cost > 0:
                    existing_asset_basis_warnings.append(
                        f"WARNING: Asset {row.get('Asset ID', idx)} (Existing Asset) using full cost ${cost:,.2f} as basis. "
                        f"If this asset took Section 179 or Bonus in prior years, depreciation may be OVERSTATED. "
                        f"Add 'Prior Section 179' and 'Prior Bonus' columns to fix."
                    )

            depreciable_basis = max(original_basis, 0.0)
        else:
            # Current year addition: use current year 179/bonus
            depreciable_basis = max(cost - sec179 - bonus, 0.0)

        depreciable_basis_list.append(depreciable_basis)

        # Skip MACRS calculation if no depreciable basis
        if depreciable_basis <= 0:
            macrs_year1_depreciation.append(0.0)
            continue

        # Get MACRS parameters
        # Support both "Recovery Period" (preferred) and "MACRS Life" (fallback) column names
        recovery_period = row.get("Recovery Period") or row.get("MACRS Life")
        method = row.get("Method", "200DB")
        convention = row.get("Convention", "HY")
        quarter = row.get("Quarter")  # For MQ convention

        # Get month for MM convention (real property)
        month = None
        if convention == "MM" and in_service_date:
            if isinstance(in_service_date, date):
                month = in_service_date.month

        # CRITICAL FIX: Handle disposals with partial year depreciation
        # Per IRS Publication 946, disposal year gets partial depreciation based on convention
        if _is_disposal(row):
            # Calculate depreciation year for disposal
            disposal_date = parse_date(row.get("Disposal Date"))

            if in_service_date and not pd.isna(in_service_date) and recovery_period:
                try:
                    in_service_year = in_service_date.year if hasattr(in_service_date, 'year') else int(str(in_service_date)[:4])
                    recovery_year = tax_year - in_service_year + 1
                    max_year = int(float(recovery_period)) + 1
                    recovery_year = max(1, min(recovery_year, max_year))

                    # Get disposal quarter/month for convention calculations
                    disposal_quarter = None
                    disposal_month = None
                    if disposal_date and not pd.isna(disposal_date):
                        disposal_month = disposal_date.month if hasattr(disposal_date, 'month') else None
                        disposal_quarter = get_quarter(disposal_date)

                    # Calculate disposal year depreciation
                    macrs_dep = calculate_disposal_year_depreciation(
                        basis=depreciable_basis,
                        recovery_period=int(float(recovery_period)),
                        method=method,
                        convention=convention,
                        year_of_recovery=recovery_year,
                        disposal_quarter=disposal_quarter,
                        disposal_month=disposal_month,
                        placed_in_service_quarter=quarter,
                        placed_in_service_month=month
                    )
                except Exception as e:
                    print(f"Warning: Disposal depreciation calculation failed for asset {row.get('Asset ID', idx)}: {e}")
                    macrs_dep = 0.0
            else:
                macrs_dep = 0.0

            macrs_year1_depreciation.append(macrs_dep)
            continue

        # Skip transfers (no depreciation for pure transfers)
        if _is_transfer(row):
            macrs_year1_depreciation.append(0.0)
            continue

        # CRITICAL FIX: Determine depreciation year based on transaction type
        # For existing assets, calculate which year of depreciation schedule we're in
        depreciation_year = 1  # Default for current year additions

        if "Existing Asset" in transaction_type:
            # Calculate depreciation year based on in-service date
            if in_service_date and not pd.isna(in_service_date):
                try:
                    if isinstance(in_service_date, date):
                        in_service_year = in_service_date.year
                    elif hasattr(in_service_date, 'year'):
                        in_service_year = in_service_date.year
                    else:
                        in_service_year = int(str(in_service_date)[:4])

                    # Year of depreciation = current tax year - in service year + 1
                    depreciation_year = tax_year - in_service_year + 1

                    # Clamp to valid range (1 to recovery_period + 1 for final year)
                    if recovery_period:
                        max_year = int(float(recovery_period)) + 1
                        depreciation_year = max(1, min(depreciation_year, max_year))
                except (ValueError, TypeError):
                    depreciation_year = 1  # Fallback

        # Calculate MACRS depreciation for the appropriate year
        try:
            macrs_dep = calculate_macrs_depreciation(
                basis=depreciable_basis,
                recovery_period=recovery_period,
                method=method,
                convention=convention,
                year=depreciation_year,  # Use calculated depreciation year
                quarter=quarter,
                month=month
            )
        except Exception as e:
            # Fallback to 0 if calculation fails
            print(f"Warning: MACRS calculation failed for asset {row.get('Asset ID', idx)}: {e}")
            macrs_dep = 0.0

        macrs_year1_depreciation.append(macrs_dep)

    df["Depreciable Basis"] = depreciable_basis_list
    df["MACRS Year 1 Depreciation"] = macrs_year1_depreciation

    # Print summary of depreciation
    total_sec179 = sum(section179_amounts)
    total_bonus = sum(bonus_amounts)
    total_macrs_current = sum(macrs_year1_depreciation)
    total_current_year_deduction = total_sec179 + total_bonus + total_macrs_current

    print("\n" + "=" * 80)
    print(f"DEPRECIATION SUMMARY - Tax Year {tax_year}")
    print("=" * 80)
    print(f"Section 179 Expensing:           ${total_sec179:>15,.2f}  (current year additions only)")
    print(f"Bonus Depreciation:              ${total_bonus:>15,.2f}  (current year additions only)")
    print(f"MACRS Current Year Depreciation: ${total_macrs_current:>15,.2f}  (all assets - correct year)")
    print(f"{'─' * 80}")
    print(f"Total Current Year Deduction:    ${total_current_year_deduction:>15,.2f}")
    print("=" * 80)
    print("NOTE: For existing assets, MACRS depreciation is calculated based on")
    print("      which year of their depreciation schedule they are in.")
    print("=" * 80 + "\n")

    # Print warnings about existing asset basis if any
    if existing_asset_basis_warnings:
        print("\n" + "!" * 80)
        print("EXISTING ASSET BASIS WARNINGS")
        print("!" * 80)
        for warning in existing_asset_basis_warnings[:10]:  # Show first 10
            print(f"  * {warning}")
        if len(existing_asset_basis_warnings) > 10:
            print(f"  ... and {len(existing_asset_basis_warnings) - 10} more warnings")
        print("\nTO FIX: Add 'Prior Section 179' and 'Prior Bonus' columns to your input data")
        print("        for existing assets that took these deductions in prior years.")
        print("!" * 80 + "\n")

    # ============================================================================
    # PHASE 4: SECTION 179 CARRYFORWARD TRACKING (IRC §179(b)(3))
    # ============================================================================
    # Apply taxable income limitation to Section 179 deduction
    # Any disallowed Section 179 carries forward indefinitely to future years
    #
    # NOTE: carryforward_from_prior_years should come from prior year tax return
    # (Form 4562, Part I, Line 13). Value is passed as function parameter.

    carryforward_from_prior_years = section_179_carryforward_from_prior_year

    if total_sec179 > 0 and taxable_income > 0:
        # Apply income limitation and allocate across assets
        df, sec179_summary = apply_section_179_carryforward_to_dataframe(
            df=df,
            taxable_business_income=taxable_income,
            carryforward_from_prior_years=carryforward_from_prior_years
        )

        # Print Section 179 carryforward report
        sec179_report = generate_section_179_report(sec179_summary)
        print(sec179_report)

        # Store carryforward summary in dataframe metadata for export
        df.attrs["section_179_summary"] = sec179_summary

    elif total_sec179 > 0 and taxable_income <= 0:
        # No taxable income - entire Section 179 must be carried forward
        print("\n" + "=" * 80)
        print("⚠️  CRITICAL: SECTION 179 TAXABLE INCOME LIMITATION")
        print("=" * 80)
        print(f"Section 179 Elected: ${total_sec179:,.2f}")
        print(f"Taxable Business Income: ${taxable_income:,.2f}")
        print("")
        print("ENTIRE Section 179 deduction disallowed due to insufficient taxable income.")
        print(f"Carryforward to next year: ${total_sec179:,.2f}")
        print("")
        print("MUST disclose on Form 4562, Part I, Line 13 of next year's return.")
        print("=" * 80 + "\n")

        # Set all Section 179 to carryforward
        df["Section 179 Allowed"] = 0.0
        df["Section 179 Carryforward"] = df["Section 179 Amount"]

    else:
        # No Section 179 - add empty columns
        df["Section 179 Allowed"] = df["Section 179 Amount"]
        df["Section 179 Carryforward"] = 0.0

    # ============================================================================
    # TIER 2: RECAPTURE CALCULATIONS (IRC §1245 / §1250)
    # ============================================================================
    # Calculate depreciation recapture for disposals

    recapture_1245 = []
    recapture_1250_ordinary = []
    recapture_1250_unrecaptured = []
    recapture_capital_gain = []
    recapture_capital_loss = []
    recapture_adjusted_basis = []

    for idx, row in df.iterrows():
        if not _is_disposal(row):
            # Not a disposal - no recapture
            recapture_1245.append(0.0)
            recapture_1250_ordinary.append(0.0)
            recapture_1250_unrecaptured.append(0.0)
            recapture_capital_gain.append(0.0)
            recapture_capital_loss.append(0.0)
            recapture_adjusted_basis.append(0.0)
            continue

        # Get disposal data
        cost = float(row.get("Cost") or 0.0)

        # Get proceeds from multiple possible column names (supports both formats)
        proceeds = float(row.get("Proceeds") or row.get("proceeds") or 0.0)

        final_category = str(row.get("Final Category", ""))

        # Get depreciation taken (for disposed assets, this would come from historical data)
        # Support both capitalized (from export) and lowercase (from sheet_loader) column names
        accumulated_dep = float(
            row.get("Accumulated Depreciation") or
            row.get("accumulated_depreciation") or 0.0
        )
        sec179_taken = float(
            row.get("Section 179 Taken (Historical)") or
            row.get("section_179_taken") or 0.0
        )
        bonus_taken = float(
            row.get("Bonus Taken (Historical)") or
            row.get("bonus_taken") or 0.0
        )

        # Determine recapture type
        recapture_type = determine_recapture_type(final_category)

        if recapture_type == "none":
            # Land - no depreciation, no recapture
            recapture_1245.append(0.0)
            recapture_1250_ordinary.append(0.0)
            recapture_1250_unrecaptured.append(0.0)
            recapture_capital_gain.append(proceeds - cost if proceeds > cost else 0.0)
            recapture_capital_loss.append(cost - proceeds if cost > proceeds else 0.0)
            recapture_adjusted_basis.append(cost)

        elif recapture_type == "1245":
            # Section 1245 - Personal property
            result = calculate_section_1245_recapture(
                cost=cost,
                accumulated_depreciation=accumulated_dep,
                proceeds=proceeds,
                section_179_taken=sec179_taken,
                bonus_taken=bonus_taken
            )

            recapture_1245.append(result["section_1245_recapture"])
            recapture_1250_ordinary.append(0.0)
            recapture_1250_unrecaptured.append(0.0)
            recapture_capital_gain.append(result["capital_gain"])
            recapture_capital_loss.append(result["capital_loss"])
            recapture_adjusted_basis.append(result["adjusted_basis"])

        elif recapture_type == "1250":
            # Section 1250 - Real property
            result = calculate_section_1250_recapture(
                cost=cost,
                accumulated_depreciation=accumulated_dep,
                proceeds=proceeds,
                accelerated_depreciation=0.0  # Post-1986 real property uses SL only
            )

            recapture_1245.append(0.0)
            recapture_1250_ordinary.append(result["section_1250_recapture"])
            recapture_1250_unrecaptured.append(result["unrecaptured_1250_gain"])
            recapture_capital_gain.append(result["capital_gain"])
            recapture_capital_loss.append(result["capital_loss"])
            recapture_adjusted_basis.append(result["adjusted_basis"])

    # Add recapture columns to dataframe
    df["§1245 Recapture (Ordinary Income)"] = recapture_1245
    df["§1250 Recapture (Ordinary Income)"] = recapture_1250_ordinary
    df["Unrecaptured §1250 Gain (25% rate)"] = recapture_1250_unrecaptured
    df["Capital Gain"] = recapture_capital_gain
    df["Capital Loss"] = recapture_capital_loss
    df["Adjusted Basis at Disposal"] = recapture_adjusted_basis

    # ----------------------------------------------------------------------
    # Build FA CS Export - UPDATED FOR FA CS IMPORT COMPATIBILITY
    # ----------------------------------------------------------------------
    # Based on FA CS import field mapping screenshots (2025-01-20)
    # Column names and formats must match FA CS expectations exactly
    # ----------------------------------------------------------------------
    fa = pd.DataFrame()

    # ============================================================================
    # CORE ASSET INFORMATION
    # ============================================================================
    # CRITICAL: FA CS requires Asset# to be NUMERIC ONLY (no "A-001" format)
    # Excel will strip leading zeros (0001 → 1)
    # Generate sequential numeric IDs starting from asset_number_start
    # Default is 1, but can be set higher (e.g., 1001) to continue from existing FA CS assets
    fa["Asset #"] = range(asset_number_start, asset_number_start + len(df))

    # Keep original Asset ID for reference (if alphanumeric)
    # This allows cross-referencing between FA CS numeric IDs and client's original IDs
    fa["Original Asset ID"] = df["Asset ID"].astype(str)

    fa["Description"] = df["Description"].astype(str)  # FA CS uses "Description" not "Property Description"

    # Format dates as M/D/YYYY strings (not datetime objects with timestamps)
    # FA CS expects clean date format like "1/1/2024" not "2024-01-01 00:00:00"
    # Platform-independent formatting (Windows doesn't support %-m directive)
    fa["Date In Service"] = _format_date_for_fa_cs(pd.to_datetime(df["In Service Date"]))
    fa["Acquisition Date"] = _format_date_for_fa_cs(pd.to_datetime(df["Acquisition Date"]))

    # Disposal fields (for FA CS import mapping)
    # Map from source columns: Disposal Date, Date Disposed, or empty
    if "Disposal Date" in df.columns:
        fa["Date Disposed"] = _format_date_for_fa_cs(pd.to_datetime(df["Disposal Date"], errors='coerce'))
    elif "Date Disposed" in df.columns:
        fa["Date Disposed"] = _format_date_for_fa_cs(pd.to_datetime(df["Date Disposed"], errors='coerce'))
    else:
        fa["Date Disposed"] = ""

    # Gross Proceeds for disposals (map from Proceeds or Sale Price)
    # Use pd.to_numeric to handle any string values
    if "Proceeds" in df.columns:
        fa["Gross Proceeds"] = pd.to_numeric(df["Proceeds"], errors='coerce').fillna(0).round(2)
    elif "Sale Price" in df.columns:
        fa["Gross Proceeds"] = pd.to_numeric(df["Sale Price"], errors='coerce').fillna(0).round(2)
    else:
        fa["Gross Proceeds"] = 0.0

    # ============================================================================
    # TAX (FEDERAL) DEPRECIATION FIELDS
    # ============================================================================
    # FA CS requires "Tax" prefix for federal tax fields

    # Round to 2 decimal places for FA CS compatibility
    # Use pd.to_numeric to handle any string values
    fa["Tax Cost"] = pd.to_numeric(df["Cost"], errors='coerce').fillna(0).round(2)  # FA CS uses "Tax Cost" not "Cost/Basis"

    # Tax Method: Convert to FA CS format
    # FA CS only accepts "MACRS" (not "MACRS GDS" or "200DB") per user testing
    fa["Tax Method"] = df.apply(
        lambda r: _convert_method_to_fa_cs_format(r.get("Uses ADS", False)) if not _is_disposal(r) and not _is_transfer(r) else "",
        axis=1,
    )

    # Tax Life: Just numbers (5, 7, 15, 27.5, 39) - NOT "5-Year MACRS"
    # FA CS expects plain numbers for Tax Life field
    fa["Tax Life"] = df.apply(
        lambda r: r.get("Recovery Period", r.get("MACRS Life", "")) if not _is_disposal(r) and not _is_transfer(r) else "",
        axis=1,
    )

    # Convention: Same as before (HY, MQ, MM)
    fa["Convention"] = df.apply(
        lambda r: r["Convention"] if not _is_disposal(r) and not _is_transfer(r) else "",
        axis=1,
    )

    # ============================================================================
    # SECTION 179 & BONUS DEPRECIATION
    # ============================================================================
    # FA CS uses "Tax Sec 179 Expensed" not "Section 179 Amount"
    # Round to 2 decimal places for FA CS compatibility
    # Use pd.to_numeric to handle any string values
    fa["Tax Sec 179 Expensed"] = pd.to_numeric(df["Section 179 Amount"], errors='coerce').fillna(0).round(2)
    fa["Bonus Amount"] = pd.to_numeric(df["Bonus Amount"], errors='coerce').fillna(0).round(2)
    fa["Bonus % Applied"] = pd.to_numeric(df["Bonus Percentage Used"], errors='coerce').fillna(0).apply(lambda x: f"{x:.0%}" if x > 0 else "")

    # ============================================================================
    # PRIOR & CURRENT DEPRECIATION
    # ============================================================================
    # For existing assets: Tax Prior Depreciation = accumulated depreciation
    # For current year additions: Tax Prior Depreciation = 0
    # Tax Cur Depreciation = current year MACRS depreciation (for all assets)

    # Round to 2 decimal places for FA CS compatibility
    fa["Tax Prior Depreciation"] = df.apply(
        lambda r: round(float(r.get("Accumulated Depreciation") or r.get("accumulated_depreciation") or 0.0), 2)
                  if "Existing Asset" in str(r.get("Transaction Type", "")) else 0.0,
        axis=1
    )

    fa["Tax Cur Depreciation"] = pd.to_numeric(df["MACRS Year 1 Depreciation"], errors='coerce').fillna(0).round(2)

    # ============================================================================
    # BOOK DEPRECIATION FIELDS (for GAAP reporting)
    # ============================================================================
    # FA CS can import Book columns - user testing confirmed this works!
    # Book typically uses Straight Line method with longer lives than Tax

    # Book Cost: Same as Tax Cost for most assets
    fa["Book Cost"] = pd.to_numeric(df["Cost"], errors='coerce').fillna(0).round(2)

    # Book Method: Typically Straight Line (SL) for GAAP
    # FA CS accepts: SL, DB (declining balance), etc.
    fa["Book Method"] = df.apply(
        lambda r: "SL" if not _is_disposal(r) and not _is_transfer(r) else "",
        axis=1,
    )

    # Book Life: Typically longer than Tax life (e.g., 10 years for computers vs 5 for tax)
    # Common Book lives: Computers=10, Furniture=10, Vehicles=5, Buildings=40
    def _get_book_life(row):
        if _is_disposal(row) or _is_transfer(row):
            return ""
        tax_life = row.get("Recovery Period", row.get("MACRS Life", 0))
        try:
            tax_life = float(tax_life)
        except (ValueError, TypeError):
            tax_life = 7  # Default

        # Book life is typically longer than tax life
        if tax_life <= 5:
            return 10  # 5-year tax → 10-year book
        elif tax_life <= 7:
            return 10  # 7-year tax → 10-year book
        elif tax_life <= 15:
            return 15  # 15-year tax → 15-year book
        elif tax_life >= 27.5:
            return 40  # Real property → 40-year book
        else:
            return 10  # Default

    fa["Book Life"] = df.apply(_get_book_life, axis=1)

    # ============================================================================
    # STATE DEPRECIATION COLUMNS (MI - Michigan)
    # ============================================================================
    # FA CS has state-specific books (MI, CA, NY, etc.) separate from generic "State"
    # Importing to generic "State" doesn't populate state-specific reports (e.g., MI Report)
    # Must use state-specific column names like "MI Cost" for proper FA CS assignment
    # TODO: Add state selector in UI for multi-state support

    # MI Cost: Same as Tax Cost
    fa["MI Cost"] = pd.to_numeric(df["Cost"], errors='coerce').fillna(0).round(2)

    # MI Method: Michigan follows federal MACRS
    fa["MI Method"] = df.apply(
        lambda r: _convert_method_to_fa_cs_format(r["Method"])
        if not _is_disposal(r) and not _is_transfer(r)
        else "",
        axis=1,
    )

    # MI Life: Same as Tax Life
    fa["MI Life"] = df.apply(
        lambda r: r.get("Recovery Period", r.get("MACRS Life", ""))
        if not _is_disposal(r) and not _is_transfer(r)
        else "",
        axis=1,
    )

    # ============================================================================
    # TRANSACTION TYPE & SHEET ROLE
    # ============================================================================
    # Transaction Type has been properly classified earlier in build_fa()
    # Based on in-service date vs tax year:
    # - "Current Year Addition": In-service date = current tax year
    # - "Existing Asset": In-service date < current tax year
    # - "Disposal": Disposal indicators
    # - "Transfer": Transfer indicators

    # Transaction Type (from classifier - already set) - keep for internal tracking
    fa["Transaction Type"] = df["Transaction Type"]

    # Sheet Role: FA CS ALWAYS uses "main" for all transaction types
    # Do NOT use "addition", "disposal", "transfer", "existing"
    fa["Sheet Role"] = "main"

    # ============================================================================
    # SUPPLEMENTAL COLUMNS (for analysis, not required by FA CS)
    # ============================================================================

    # Depreciable Basis - shows basis after Section 179 and Bonus
    fa["Depreciable Basis"] = df.get("Depreciable Basis", 0.0)

    # PHASE 4: Section 179 Carryforward (IRC §179(b)(3))
    fa["Section 179 Allowed"] = df.get("Section 179 Allowed", df.get("Section 179 Amount", 0.0))
    fa["Section 179 Carryforward"] = df.get("Section 179 Carryforward", 0.0)

    # TIER 3: De Minimis Safe Harbor
    fa["De Minimis Expensed"] = df.get("De Minimis Expensed", 0.0)

    # TIER 3: Mid-Quarter Convention Quarter
    fa["Quarter (MQ)"] = df.get("Quarter", None)

    # IRC §280F luxury auto limit notes
    fa["Auto Limit Notes"] = df["Auto Limit Notes"]

    # TIER 2: Recapture columns (for disposals)
    fa["§1245 Recapture (Ordinary Income)"] = df["§1245 Recapture (Ordinary Income)"]
    fa["§1250 Recapture (Ordinary Income)"] = df["§1250 Recapture (Ordinary Income)"]
    fa["Unrecaptured §1250 Gain (25%)"] = df["Unrecaptured §1250 Gain (25% rate)"]
    fa["Capital Gain"] = df["Capital Gain"]
    fa["Capital Loss"] = df["Capital Loss"]
    fa["Adjusted Basis at Disposal"] = df["Adjusted Basis at Disposal"]

    # TIER 2: ADS flag
    fa["Uses ADS"] = df["Uses ADS"]

    # Source tracking
    fa["Source"] = df.get("Source", "")

    # Original Category (client-provided)
    fa["Client Category Original"] = df.get("Client Category Original", df.get("Client Category", ""))

    # Final Computed Category (for additions only)
    fa["Final Category"] = df.apply(
        lambda r: r["Final Category"] if not _is_disposal(r) and not _is_transfer(r) else "",
        axis=1,
    )

    # FA CS Wizard Category - EXACT dropdown option text for UiPath RPA automation
    # This column tells UiPath which option to select in FA CS wizard dropdown
    fa["FA_CS_Wizard_Category"] = df.apply(_get_fa_cs_wizard_category, axis=1)

    # Asset Type - General classification for FA CS folder organization
    # Added to help FA CS classify assets into "Business" folder instead of "Miscellaneous"
    fa["Asset Type"] = fa.apply(_determine_asset_type, axis=1)

    # ===========================================================================
    # CPA REVIEW ENHANCEMENTS - NBV, Materiality, Audit Trail
    # ===========================================================================
    # These features help CPAs review and validate the export before filing

    # NBV Reconciliation - validates Net Book Value = Cost - Accumulated Depreciation
    # Only set NBV to None if it doesn't already exist (preserve extracted values)
    if "NBV" not in fa.columns:
        fa["NBV"] = None
    fa = _compute_nbv_reco(fa, tolerance=5.0)

    # Materiality Scoring - prioritize high-value assets for CPA review
    fa = _compute_materiality(fa)

    # Classification Explanations - explain why each asset was classified as it was
    fa["ClassificationExplanation"] = fa.apply(_classification_explanation, axis=1)
    fa["MACRS_Reason_Code"] = fa.apply(_macrs_reason_code, axis=1)
    fa["ConfidenceGrade"] = fa.apply(_confidence_grade, axis=1)

    # Audit Trail - comprehensive audit fields with SHA256 integrity hash
    audit_df = fa.apply(_audit_fields, axis=1, result_type="expand")
    fa = pd.concat([fa, audit_df], axis=1)

    # Typo tracking columns (if they exist)
    if "Desc_TypoFlag" in df.columns:
        fa["Desc_TypoFlag"] = df["Desc_TypoFlag"]
        fa["Desc_TypoNote"] = df.get("Desc_TypoNote", "")
    if "Cat_TypoFlag" in df.columns:
        fa["Cat_TypoFlag"] = df["Cat_TypoFlag"]
        fa["Cat_TypoNote"] = df.get("Cat_TypoNote", "")

    # Print CPA review summary
    high_priority = (fa["ReviewPriority"] == "High").sum()
    nbv_issues = (fa["NBV_Reco"] == "CHECK").sum()

    if high_priority > 0 or nbv_issues > 0:
        print("\n" + "=" * 80)
        print("CPA REVIEW SUMMARY")
        print("=" * 80)
        if high_priority > 0:
            print(f"📊 High Materiality Assets: {high_priority} (Review Priority: High)")
        if nbv_issues > 0:
            print(f"⚠️  NBV Reconciliation Issues: {nbv_issues} (NBV_Reco: CHECK)")
        print("=" * 80 + "\n")

    # ===========================================================================
    # EXPORT QUALITY VALIDATION - Fixed Asset CS & RPA Compatibility
    # ===========================================================================
    # Validate export file quality before returning
    # CRITICAL: This ensures the file is ready for Fixed Asset CS import
    # and RPA automation without errors

    is_valid, validation_errors, summary = validate_fixed_asset_cs_export(
        fa, verbose=True
    )

    if not is_valid:
        print("\n⚠️  WARNING: Export validation found CRITICAL/ERROR issues.")
        print("   Review validation report above before proceeding with RPA.")
        print("   Fix critical issues to prevent automation failures.\n")

    return fa


def build_fa_minimal(
    df: pd.DataFrame,
    tax_year: int,
    asset_number_start: int = 1,  # Starting Asset # for FA CS
) -> pd.DataFrame:
    """
    Build MINIMAL FA CS import file - let FA CS calculate depreciation.

    This version includes ONLY the fields required for FA CS to import and
    calculate depreciation internally. Use this to test whether FA CS:
    1. Accepts minimal imports
    2. Calculates Section 179, Bonus, MACRS internally
    3. Produces same results as manual entry

    Args:
        df: DataFrame with classified assets (must have Transaction Type set)
        tax_year: Current tax year
        asset_number_start: Starting number for Asset # sequence (default 1)

    Returns:
        Minimal FA CS import DataFrame

    Testing Strategy:
        1. Export with this function
        2. Also manually enter same asset in FA CS
        3. Compare FA CS calculated values
        4. If they match, FA CS is recalculating (safe to use minimal export)
        5. If they differ, FA CS uses imported values (need full export)
    """
    fa = pd.DataFrame()

    # ============================================================================
    # MINIMUM REQUIRED FIELDS ONLY
    # ============================================================================
    # CRITICAL: FA CS requires Asset# to be NUMERIC ONLY (no "A-001" format)
    # Excel will strip leading zeros (0001 → 1)
    # Generate sequential numeric IDs starting from asset_number_start
    fa["Asset #"] = range(asset_number_start, asset_number_start + len(df))

    # Keep original Asset ID for reference (if alphanumeric)
    fa["Original Asset ID"] = df["Asset ID"].astype(str)

    fa["Description"] = df["Description"].astype(str)

    # Format dates as M/D/YYYY strings (FA CS expects clean date format)
    # Platform-independent formatting (Windows doesn't support %-m directive)
    fa["Date In Service"] = _format_date_for_fa_cs(pd.to_datetime(df["In Service Date"]))
    fa["Acquisition Date"] = _format_date_for_fa_cs(pd.to_datetime(df["Acquisition Date"]))

    # Round to 2 decimal places for FA CS compatibility
    # Use pd.to_numeric to handle any string values
    fa["Tax Cost"] = pd.to_numeric(df["Cost"], errors='coerce').fillna(0).round(2)

    # Tax Method: Convert to FA CS format
    fa["Tax Method"] = df.apply(
        lambda r: _convert_method_to_fa_cs_format(r.get("Uses ADS", False))
        if not _is_disposal(r) and not _is_transfer(r) else "",
        axis=1,
    )

    # Tax Life: Plain numbers only
    fa["Tax Life"] = df.apply(
        lambda r: r.get("Recovery Period", r.get("MACRS Life", ""))
        if not _is_disposal(r) and not _is_transfer(r) else "",
        axis=1,
    )

    # Convention
    fa["Convention"] = df.apply(
        lambda r: r.get("Convention", "")
        if not _is_disposal(r) and not _is_transfer(r) else "",
        axis=1,
    )

    # Sheet Role - always "main"
    fa["Sheet Role"] = "main"

    # ============================================================================
    # OPTIONAL: Include Transaction Type for tracking (not required by FA CS)
    # ============================================================================
    fa["Transaction Type"] = df.get("Transaction Type", "")

    # ============================================================================
    # NOTE: NO calculated fields included
    # ============================================================================
    # The following are NOT included - let FA CS calculate them:
    # - Tax Sec 179 Expensed (FA CS will calculate based on elections)
    # - Bonus Amount (FA CS will calculate based on rules)
    # - Tax Cur Depreciation (FA CS will calculate MACRS)
    # - Tax Prior Depreciation (FA CS will track for existing assets)
    # - Depreciable Basis (FA CS will calculate)

    print("\n" + "=" * 80)
    print("MINIMAL FA CS EXPORT - Testing Mode")
    print("=" * 80)
    print("This export includes ONLY minimum required fields.")
    print("FA CS will calculate all depreciation internally.")
    print("")
    print("Use this to test whether FA CS:")
    print("  1. Accepts minimal imports")
    print("  2. Calculates Section 179/Bonus/MACRS internally")
    print("  3. Matches manual entry calculations")
    print("")
    print("Fields included:")
    print("  ✓ Asset #, Description, Date In Service, Acquisition Date")
    print("  ✓ Tax Cost, Tax Method, Tax Life, Convention")
    print("  ✓ Sheet Role")
    print("")
    print("Fields NOT included (let FA CS calculate):")
    print("  ✗ Tax Sec 179 Expensed")
    print("  ✗ Bonus Amount")
    print("  ✗ Tax Cur Depreciation")
    print("  ✗ Tax Prior Depreciation")
    print("=" * 80 + "\n")

    return fa


def export_fa_excel(fa_df: pd.DataFrame, audit_info: dict = None) -> bytes:
    """
    Export Fixed Asset data to Excel with streamlined 5-tab format.

    Tabs:
    1. FA_CS_Data - Full data for FA CS manual entry reference
    2. RPA_Input - Minimal fields for UiPath RPA automation
    3. Review - Items needing CPA attention
    4. Summary - Totals, counts, audit info
    5. De_Minimis - Expensed items (only if any exist)

    Args:
        fa_df: Fixed asset dataframe from build_fa()
        audit_info: Optional dict with preparer info and approvals for audit trail

    Returns:
        Excel file bytes
    """
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.formatting.rule import CellIsRule

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ================================================================
        # SEPARATE DE MINIMIS ITEMS FROM DEPRECIABLE ASSETS
        # ================================================================
        # De minimis items are expensed immediately (not depreciated)
        # They should NOT go into FA CS - separate them for documentation
        de_minimis_mask = (
            (fa_df.get("De Minimis Expensed", pd.Series([0] * len(fa_df))) > 0) |
            ((fa_df.get("Tax Cost", pd.Series([0] * len(fa_df))) == 0) &
             (fa_df.get("De Minimis Expensed", pd.Series([0] * len(fa_df))) > 0))
        )
        de_minimis_df = fa_df[de_minimis_mask].copy()
        depreciable_df = fa_df[~de_minimis_mask].copy()

        # ================================================================
        # TAB 1: FA_CS_Data - Main data for FA CS entry
        # ================================================================
        # Contains all fields needed for manual entry into FA CS
        # Excludes de minimis items (they don't go into FA CS)
        fa_data_cols = [
            "Asset #",
            "Description",
            "Date In Service",
            "Tax Cost",
            "Tax Method",
            "Tax Life",
            "Tax Sec 179 Expensed",  # CRITICAL: Only way to set 179!
            "Tax Cur Depreciation",  # BLANK for FA CS to calculate
            "Tax Prior Depreciation",
            "Book Cost",             # Book depreciation - FA CS imports these!
            "Book Method",           # Typically "SL" for GAAP
            "Book Life",             # Typically longer than Tax life
            "MI Cost",               # Michigan state depreciation
            "MI Method",             # MACRS (same as Tax)
            "MI Life",               # Same as Tax Life
            "Date Disposed",         # For disposals
            "Gross Proceeds",        # For disposals
            "FA_CS_Wizard_Category", # For UiPath RPA - wizard dropdown selection
        ]

        # Add columns if they don't exist (on depreciable assets only)
        if "Date Disposed" not in depreciable_df.columns:
            depreciable_df["Date Disposed"] = ""
        if "Gross Proceeds" not in depreciable_df.columns:
            depreciable_df["Gross Proceeds"] = ""

        # CRITICAL: Set Tax Cur Depreciation to blank (not 0) so FA CS calculates it
        # If we put 0, FA CS treats it as an override and won't calculate!
        # Save original values before blanking for use in calculations later
        original_tax_cur_depreciation = fa_df["Tax Cur Depreciation"].copy() if "Tax Cur Depreciation" in fa_df.columns else pd.Series([0] * len(fa_df))
        depreciable_df["Tax Cur Depreciation"] = ""

        # Select only columns that exist (EXCLUDES de minimis items)
        available_fa = [c for c in fa_data_cols if c in depreciable_df.columns]
        fa_data_df = depreciable_df[available_fa].copy()

        # Write FA_CS_Data tab
        fa_data_df.to_excel(writer, sheet_name="FA_CS_Data", index=False)

        # Format the FA_CS_Data worksheet
        ws_fa = writer.sheets["FA_CS_Data"]
        ws_fa.freeze_panes = "A2"
        if ws_fa.max_row > 0:
            ws_fa.auto_filter.ref = ws_fa.dimensions

        _apply_professional_formatting(ws_fa, fa_data_df)

        # ================================================================
        # TAB 2: RPA_Input - Minimal fields for UiPath RPA automation
        # ================================================================
        # Simple, clean data for RPA to:
        # 1. Click "Add Asset" in FA CS
        # 2. Enter these fields
        # 3. Click Wizard → Select FA_CS_Wizard_Category
        # 4. Optionally click "Max 179" if flagged
        rpa_cols = [
            "Asset #",
            "Description",
            "Date In Service",
            "Tax Cost",
            "FA_CS_Wizard_Category",  # Which wizard dropdown to select
        ]

        available_rpa = [c for c in rpa_cols if c in depreciable_df.columns]
        rpa_df = depreciable_df[available_rpa].copy()

        # Add Max_179 flag for RPA (CPA can review and modify before running RPA)
        # Flag assets that are eligible for 179 (current year additions with cost > 0)
        if "Tax Sec 179 Expensed" in depreciable_df.columns:
            rpa_df["Max_179"] = depreciable_df["Tax Sec 179 Expensed"].apply(
                lambda x: "YES" if pd.to_numeric(x, errors='coerce') and pd.to_numeric(x, errors='coerce') > 0 else ""
            )
        else:
            rpa_df["Max_179"] = ""

        rpa_df.to_excel(writer, sheet_name="RPA_Input", index=False)

        # Format RPA_Input worksheet
        ws_rpa = writer.sheets["RPA_Input"]
        ws_rpa.freeze_panes = "A2"
        if ws_rpa.max_row > 0:
            ws_rpa.auto_filter.ref = ws_rpa.dimensions

        _apply_professional_formatting(ws_rpa, rpa_df)

        # ================================================================
        # TAB 3: Review - Items needing CPA attention (combines CPA_Review + Audit)
        # ================================================================
        # Key fields for CPA review: priorities, warnings, classification info

        # Add eligibility reason columns for CPA verification
        def get_179_eligibility(row):
            """Explain why asset is/isn't eligible for Section 179."""
            trans_type = str(row.get("Transaction Type", "")).lower()
            cost = pd.to_numeric(row.get("Tax Cost", 0), errors='coerce') or 0
            sec179 = pd.to_numeric(row.get("Tax Sec 179 Expensed", 0), errors='coerce') or 0

            if "disposal" in trans_type:
                return "No - Disposal (not eligible)"
            if "existing" in trans_type:
                return "No - Existing asset (must be new acquisition)"
            if cost <= 0:
                return "No - Zero/negative cost"
            if sec179 > 0:
                return f"Yes - ${sec179:,.0f} elected"
            # Check if it could have been eligible
            category = str(row.get("Final Category", "")).lower()
            if "land" in category or "real property" in category:
                return "No - Real property (not §179 eligible)"
            return "Eligible but not elected"

        def get_bonus_eligibility(row):
            """Explain why asset is/isn't eligible for Bonus depreciation."""
            trans_type = str(row.get("Transaction Type", "")).lower()
            cost = pd.to_numeric(row.get("Tax Cost", 0), errors='coerce') or 0
            bonus = pd.to_numeric(row.get("Bonus Amount", 0), errors='coerce') or 0

            if "disposal" in trans_type:
                return "No - Disposal"
            if cost <= 0:
                return "No - Zero cost"
            if bonus > 0:
                return f"Yes - ${bonus:,.0f} (after §179)"
            # Check why no bonus
            category = str(row.get("Final Category", "")).lower()
            if "land" in category:
                return "No - Land (not depreciable)"
            if "used" in trans_type.lower():
                return "Check - Used property (verify eligibility)"
            return "No bonus applied"

        fa_df["§179_Eligibility"] = fa_df.apply(get_179_eligibility, axis=1)
        fa_df["Bonus_Eligibility"] = fa_df.apply(get_bonus_eligibility, axis=1)

        review_cols = [
            "Asset #",
            "Description",
            "Tax Cost",
            "Final Category",
            "Transaction Type",
            "ReviewPriority",
            "ConfidenceGrade",
            "NBV_Reco",
            "AuditWarnings",
            "ClassificationExplanation",
            "Tax Sec 179 Expensed",
            "§179_Eligibility",
            "Bonus Amount",
            "Bonus_Eligibility",
        ]

        # Select available columns
        available_review = [c for c in review_cols if c in fa_df.columns]
        review_df = fa_df[available_review].copy()

        # Add calculated total deduction column
        if all(c in review_df.columns for c in ["Tax Sec 179 Expensed", "Bonus Amount"]):
            review_df["Total Year 1 Deduction"] = (
                pd.to_numeric(review_df["Tax Sec 179 Expensed"], errors='coerce').fillna(0) +
                pd.to_numeric(review_df["Bonus Amount"], errors='coerce').fillna(0) +
                pd.to_numeric(original_tax_cur_depreciation, errors='coerce').fillna(0)
            )

        review_df.to_excel(writer, sheet_name="Review", index=False)

        # Format Review worksheet
        ws_review = writer.sheets["Review"]
        ws_review.freeze_panes = "A2"
        if ws_review.max_row > 0:
            ws_review.auto_filter.ref = ws_review.dimensions

        _apply_professional_formatting(ws_review, review_df)
        _apply_conditional_formatting(ws_review, review_df)

        # ================================================================
        # TAB 4: Summary - Totals, verification, sample calculations
        # ================================================================

        # Calculate summary statistics
        # CRITICAL FIX: Convert to numeric before summing to avoid UFuncTypeError
        # Some columns may contain strings (e.g., "" or "N/A") instead of numbers
        def safe_sum(series_or_default):
            """Safely sum a series, converting to numeric first to handle mixed types."""
            if isinstance(series_or_default, pd.Series):
                return pd.to_numeric(series_or_default, errors='coerce').fillna(0).sum()
            return 0

        total_sec179 = safe_sum(fa_df.get("Tax Sec 179 Expensed", pd.Series([0])))
        total_sec179_allowed = safe_sum(fa_df.get("Section 179 Allowed", pd.Series([0])))
        total_sec179_carryforward = safe_sum(fa_df.get("Section 179 Carryforward", pd.Series([0])))
        total_bonus = safe_sum(fa_df.get("Bonus Amount", pd.Series([0])))
        total_macrs = safe_sum(original_tax_cur_depreciation)
        total_de_minimis = safe_sum(fa_df.get("De Minimis Expensed", pd.Series([0])))

        # Count by priority
        high_priority = len(fa_df[fa_df.get("ReviewPriority", "") == "High"]) if "ReviewPriority" in fa_df.columns else 0
        medium_priority = len(fa_df[fa_df.get("ReviewPriority", "") == "Medium"]) if "ReviewPriority" in fa_df.columns else 0
        low_priority = len(fa_df[fa_df.get("ReviewPriority", "") == "Low"]) if "ReviewPriority" in fa_df.columns else 0

        # Count NBV issues
        nbv_issues = len(fa_df[fa_df.get("NBV_Reco", "") == "CHECK"]) if "NBV_Reco" in fa_df.columns else 0

        # Count luxury auto adjustments
        luxury_auto_count = 0
        if "Auto Limit Notes" in fa_df.columns:
            luxury_auto_count = len(fa_df[fa_df["Auto Limit Notes"].astype(str).str.contains("§280F", na=False)])

        # Count transaction types
        total_assets = len(fa_df)
        current_year_additions = 0
        existing_assets = 0
        disposals = 0

        if "Transaction Type" in fa_df.columns:
            current_year_additions = len(fa_df[fa_df["Transaction Type"].astype(str).str.contains("Current Year", na=False)])
            existing_assets = len(fa_df[fa_df["Transaction Type"].astype(str).str.contains("Existing", na=False)])
            disposals = len(fa_df[fa_df["Transaction Type"].astype(str).str.contains("Disposal", na=False)])

        # Create summary dataframe
        summary_data = [
            ["", ""],
            ["FIXED ASSET DEPRECIATION SUMMARY", ""],
            ["", ""],

            ["SECTION 179 EXPENSING", ""],
            ["Total Section 179 Elected", f"${total_sec179:,.2f}"],
            ["Allowed - Current Year", f"${total_sec179_allowed:,.2f}"],
            ["Carryforward - Next Year", f"${total_sec179_carryforward:,.2f}"],
            ["", ""],

            ["BONUS DEPRECIATION", ""],
            ["Total Bonus Depreciation", f"${total_bonus:,.2f}"],
            ["", ""],

            ["MACRS DEPRECIATION", ""],
            ["Total MACRS Year 1", f"${total_macrs:,.2f}"],
            ["", ""],

            ["DE MINIMIS SAFE HARBOR", ""],
            ["Total De Minimis Expensed", f"${total_de_minimis:,.2f}"],
            ["", ""],

            ["TOTAL YEAR 1 DEDUCTION", f"${total_sec179_allowed + total_bonus + total_macrs + total_de_minimis:,.2f}"],
            ["", ""],
            ["", ""],

            ["CPA REVIEW ITEMS", ""],
            ["High Priority Assets", str(high_priority)],
            ["Medium Priority Assets", str(medium_priority)],
            ["Low Priority Assets", str(low_priority)],
            ["", ""],

            ["ISSUES REQUIRING ATTENTION", ""],
            ["NBV Reconciliation Issues", str(nbv_issues)],
            ["Luxury Auto Limit Adjustments", str(luxury_auto_count)],
            ["", ""],

            ["ASSET COUNTS", ""],
            ["Total Assets", str(total_assets)],
            ["Current Year Additions", str(current_year_additions)],
            ["Existing Assets", str(existing_assets)],
            ["Disposals", str(disposals)],
        ]

        # Calculate confidence breakdown percentages
        grade_a = len(fa_df[fa_df.get("ConfidenceGrade", pd.Series(dtype=str)) == "A"]) if "ConfidenceGrade" in fa_df.columns else 0
        grade_b = len(fa_df[fa_df.get("ConfidenceGrade", pd.Series(dtype=str)) == "B"]) if "ConfidenceGrade" in fa_df.columns else 0
        grade_c = len(fa_df[fa_df.get("ConfidenceGrade", pd.Series(dtype=str)) == "C"]) if "ConfidenceGrade" in fa_df.columns else 0
        grade_d = len(fa_df[fa_df.get("ConfidenceGrade", pd.Series(dtype=str)) == "D"]) if "ConfidenceGrade" in fa_df.columns else 0

        pct_a = (grade_a / total_assets * 100) if total_assets > 0 else 0
        pct_b = (grade_b / total_assets * 100) if total_assets > 0 else 0
        pct_c = (grade_c / total_assets * 100) if total_assets > 0 else 0
        pct_d = (grade_d / total_assets * 100) if total_assets > 0 else 0

        summary_data.extend([
            ["", ""],
            ["CLASSIFICATION CONFIDENCE", ""],
            ["Grade A (High)", f"{grade_a} ({pct_a:.0f}%)"],
            ["Grade B (Medium)", f"{grade_b} ({pct_b:.0f}%)"],
            ["Grade C (Low)", f"{grade_c} ({pct_c:.0f}%)"],
            ["Grade D (Manual Review)", f"{grade_d} ({pct_d:.0f}%)"],
        ])

        # Add processing info section (always shown)
        summary_data.extend([
            ["", ""],
            ["", ""],
            ["PROCESSING INFORMATION", ""],
            ["Tool Version", audit_info.get("tool_version", "1.0.0") if audit_info else "1.0.0"],
            ["Processing Date", audit_info.get("processing_timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")) if audit_info else datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Tax Year", audit_info.get("tax_year", "") if audit_info else ""],
            ["Strategy", audit_info.get("strategy", "") if audit_info else ""],
            ["Source File", audit_info.get("source_file", "") if audit_info else ""],
            ["File Checksum", audit_info.get("source_file_checksum", "") if audit_info else ""],
        ])

        # Add audit protection information if provided
        if audit_info:
            preparer = audit_info.get("preparer_name", "")
            reviewer = audit_info.get("reviewer_name", "")
            if preparer or reviewer:
                summary_data.extend([
                    ["", ""],
                    ["AUDIT PROTECTION DOCUMENTATION", ""],
                    ["Preparer", preparer],
                    ["Firm", audit_info.get("preparer_firm", "")],
                    ["Reviewer", reviewer],
                    ["Preparation Date", audit_info.get("preparation_date", "")],
                ])

            # Add confirmations if any are set
            if any([audit_info.get("client_approved"), audit_info.get("reconciled_prior"),
                    audit_info.get("source_docs_on_file"), audit_info.get("reviewed_by_manager")]):
                summary_data.extend([
                    ["", ""],
                    ["CONFIRMATIONS", ""],
                    ["Client Approved", "Yes" if audit_info.get("client_approved") else "No"],
                    ["Reconciled to Prior Year", "Yes" if audit_info.get("reconciled_prior") else "No"],
                    ["Source Documents on File", "Yes" if audit_info.get("source_docs_on_file") else "No"],
                    ["Manager/Partner Review", "Yes" if audit_info.get("reviewed_by_manager") else "No"],
                ])

            # Add pre-export checklist if completed
            if audit_info.get("pre_export_checklist_completed"):
                checklist_items = audit_info.get("checklist_items", {})
                summary_data.extend([
                    ["", ""],
                    ["PRE-EXPORT CHECKLIST (Completed)", ""],
                    ["Reviewed Critical Issues", "Yes" if checklist_items.get("reviewed_critical_issues") else "No"],
                    ["Verified Export Totals", "Yes" if checklist_items.get("verified_export_totals") else "No"],
                    ["Noted 179 Carryforward", "Yes" if checklist_items.get("noted_179_carryforward") else "No"],
                    ["Reviewed High Materiality", "Yes" if checklist_items.get("reviewed_high_materiality") else "No"],
                    ["Ready for Approval", "Yes" if checklist_items.get("ready_for_approval") else "No"],
                ])

        summary_df = pd.DataFrame(summary_data, columns=["Category", "Amount"])
        summary_df.to_excel(writer, sheet_name="Summary", index=False, header=False)

        # Format summary
        ws_summary = writer.sheets["Summary"]
        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 20

        # Bold section headers
        bold_font = Font(bold=True, size=12)
        for row_num in range(1, ws_summary.max_row + 1):
            cell_value = str(ws_summary[f'A{row_num}'].value or "")
            if cell_value and cell_value.isupper() and not cell_value.startswith(" "):
                ws_summary[f'A{row_num}'].font = bold_font

        # Highlight total deduction row
        total_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        for row_num in range(1, ws_summary.max_row + 1):
            cell_value = str(ws_summary[f'A{row_num}'].value or "")
            if "TOTAL YEAR 1 DEDUCTION" in cell_value:
                ws_summary[f'A{row_num}'].fill = total_fill
                ws_summary[f'B{row_num}'].fill = total_fill
                ws_summary[f'A{row_num}'].font = Font(bold=True, size=14)
                ws_summary[f'B{row_num}'].font = Font(bold=True, size=14)

        # ================================================================
        # TAB 5: De_Minimis - Items expensed under safe harbor (only if any)
        # ================================================================
        # De minimis items are expensed immediately (not depreciated)
        # They should NOT be entered into FA CS
        # This tab documents them for the tax return deduction

        if len(de_minimis_df) > 0:
            de_min_cols = [
                "Asset #",
                "Description",
                "Date In Service",
                "De Minimis Expensed",  # Original cost before zeroing
                "Final Category",
                "Transaction Type",
            ]
            available_de_min = [c for c in de_min_cols if c in de_minimis_df.columns]
            de_min_export_df = de_minimis_df[available_de_min].copy()

            # Add note column
            de_min_export_df["Note"] = "Expensed under de minimis safe harbor - DO NOT enter in FA CS"

            de_min_export_df.to_excel(writer, sheet_name="De_Minimis", index=False)

            ws_de_min = writer.sheets["De_Minimis"]
            ws_de_min.freeze_panes = "A2"
            if ws_de_min.max_row > 0:
                ws_de_min.auto_filter.ref = ws_de_min.dimensions

            _apply_professional_formatting(ws_de_min, de_min_export_df)

    output.seek(0)
    return output.getvalue()


def export_asset_number_crossref(fa_df: pd.DataFrame) -> bytes:
    """
    Export Asset # cross-reference table for RPA and reconciliation.

    CRITICAL FOR RPA: This mapping file allows you to:
    1. Look up the FA CS numeric Asset # from the client's original Asset ID
    2. Reconcile FA CS entries back to source data
    3. Validate RPA imports by matching original IDs to FA CS IDs

    The export includes:
    - Asset # (FA CS numeric ID, used for RPA input)
    - Original Asset ID (client's alphanumeric ID from source file)
    - Description (for easy identification)
    - Tax Cost (for reconciliation)
    - Date In Service (for reconciliation)

    Args:
        fa_df: DataFrame from build_fa() with FA CS data

    Returns:
        bytes: Excel file with cross-reference mapping
    """
    from io import BytesIO
    import openpyxl

    output = BytesIO()

    # Build cross-reference DataFrame
    crossref_cols = ["Asset #", "Original Asset ID"]

    # Add optional columns if they exist
    if "Description" in fa_df.columns:
        crossref_cols.append("Description")
    if "Tax Cost" in fa_df.columns:
        crossref_cols.append("Tax Cost")
    if "Date In Service" in fa_df.columns:
        crossref_cols.append("Date In Service")

    crossref_df = fa_df[[c for c in crossref_cols if c in fa_df.columns]].copy()

    # Add helper columns
    crossref_df["Numeric Only"] = crossref_df["Asset #"].apply(
        lambda x: "Yes" if str(x).isdigit() else "No"
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        crossref_df.to_excel(writer, sheet_name="Asset_Crossref", index=False)

        # Add instructions sheet
        instructions = pd.DataFrame({
            "Asset # Cross-Reference Guide": [
                "PURPOSE:",
                "This file maps FA CS numeric Asset #s to original client Asset IDs.",
                "",
                "HOW TO USE:",
                "1. When RPA enters assets, use 'Asset #' column (numeric only).",
                "2. Use 'Original Asset ID' to reconcile back to client source data.",
                "3. Use this file to validate FA CS entries after RPA import.",
                "",
                "COLUMNS:",
                "- Asset #: FA CS numeric ID (1, 2, 3...) - USE THIS FOR RPA",
                "- Original Asset ID: Client's ID (may have letters like 'FA-001')",
                "- Description: Asset description for easy identification",
                "- Tax Cost: Cost basis for reconciliation",
                "- Date In Service: Placed in service date for reconciliation",
                "- Numeric Only: Confirms Asset # is numeric (should all be 'Yes')",
            ]
        })
        instructions.to_excel(writer, sheet_name="Instructions", index=False)

        # Format the crossref sheet
        ws = writer.sheets["Asset_Crossref"]
        ws.freeze_panes = "A2"
        if ws.max_row > 0:
            ws.auto_filter.ref = ws.dimensions

    output.seek(0)
    return output.getvalue()


def build_fa_cs_asset_lookup(fa_cs_export_df: pd.DataFrame) -> Dict[str, int]:
    """
    Build lookup table from FA CS export to map Original Asset ID → FA CS Asset #.

    CRITICAL FOR DISPOSALS/TRANSFERS:
    When processing disposals or transfers, you need to reference the EXISTING
    FA CS Asset # (not generate a new one). This function creates the mapping.

    Usage:
        1. Export existing assets from FA CS (or use prior year's export)
        2. Call this function to build lookup
        3. Use lookup when processing disposals/transfers

    Args:
        fa_cs_export_df: DataFrame exported from FA CS or prior year export
                         Must have "Asset #" and "Original Asset ID" columns

    Returns:
        Dict mapping Original Asset ID (str) → FA CS Asset # (int)

    Example:
        lookup = build_fa_cs_asset_lookup(prior_year_export_df)
        # lookup = {"FA-001": 1, "FA-002": 2, "COMP-100": 3, ...}
    """
    lookup = {}

    if "Asset #" not in fa_cs_export_df.columns:
        raise ValueError("FA CS export must have 'Asset #' column")

    if "Original Asset ID" not in fa_cs_export_df.columns:
        raise ValueError("FA CS export must have 'Original Asset ID' column")

    for _, row in fa_cs_export_df.iterrows():
        asset_num = row["Asset #"]
        original_id = str(row["Original Asset ID"]).strip()

        if original_id and pd.notna(asset_num):
            # Store as string key → int value
            lookup[original_id] = int(asset_num)

    return lookup


def separate_by_transaction_type(fa_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """
    Separate FA CS export into separate DataFrames by transaction type.

    CRITICAL FOR RPA:
    - Additions: Need NEW Asset #s, use standard import/RPA workflow
    - Disposals: Need to UPDATE EXISTING assets, require Asset # lookup
    - Transfers: Need to UPDATE EXISTING assets, require Asset # lookup
    - Existing: Prior year assets, may need different handling

    Args:
        fa_df: Full FA CS export DataFrame

    Returns:
        Dict with keys: "additions", "disposals", "transfers", "existing"
        Each value is a filtered DataFrame
    """
    result = {
        "additions": pd.DataFrame(),
        "disposals": pd.DataFrame(),
        "transfers": pd.DataFrame(),
        "existing": pd.DataFrame(),
    }

    if "Transaction Type" not in fa_df.columns:
        # No transaction type - treat all as additions
        result["additions"] = fa_df.copy()
        return result

    trans_type = fa_df["Transaction Type"].astype(str).str.lower()

    result["additions"] = fa_df[trans_type.str.contains("current year addition", na=False)].copy()
    result["disposals"] = fa_df[trans_type.str.contains("disposal", na=False)].copy()
    result["transfers"] = fa_df[trans_type.str.contains("transfer", na=False)].copy()
    result["existing"] = fa_df[trans_type.str.contains("existing", na=False)].copy()

    return result


def match_disposals_to_fa_cs(
    disposals_df: pd.DataFrame,
    fa_cs_lookup: Dict[str, int]
) -> pd.DataFrame:
    """
    Match disposal records to their existing FA CS Asset #s.

    CRITICAL: Disposals must reference the ORIGINAL FA CS Asset # to update
    the correct asset. This function:
    1. Looks up each disposal's Original Asset ID in the FA CS lookup
    2. Replaces the generated Asset # with the correct FA CS Asset #
    3. Flags any disposals that couldn't be matched

    Args:
        disposals_df: DataFrame of disposal records (from separate_by_transaction_type)
        fa_cs_lookup: Dict from build_fa_cs_asset_lookup()

    Returns:
        DataFrame with:
        - "FA CS Asset #" column (the correct existing Asset #)
        - "Match Status" column ("Matched", "NOT FOUND", etc.)
        - Original columns preserved
    """
    if disposals_df.empty:
        return disposals_df

    result_df = disposals_df.copy()

    fa_cs_asset_nums = []
    match_statuses = []

    for _, row in result_df.iterrows():
        original_id = str(row.get("Original Asset ID", "")).strip()

        if original_id in fa_cs_lookup:
            fa_cs_asset_nums.append(fa_cs_lookup[original_id])
            match_statuses.append("Matched")
        else:
            # Try fuzzy matching (case insensitive, strip spaces)
            found = False
            for lookup_id, asset_num in fa_cs_lookup.items():
                if lookup_id.lower().strip() == original_id.lower():
                    fa_cs_asset_nums.append(asset_num)
                    match_statuses.append(f"Matched (fuzzy: {lookup_id})")
                    found = True
                    break

            if not found:
                fa_cs_asset_nums.append(None)
                match_statuses.append(f"NOT FOUND - '{original_id}' not in FA CS")

    result_df["FA CS Asset #"] = fa_cs_asset_nums
    result_df["Match Status"] = match_statuses

    # Report unmatched
    unmatched = result_df[result_df["FA CS Asset #"].isna()]
    if len(unmatched) > 0:
        print(f"\n⚠️  WARNING: {len(unmatched)} disposal(s) could not be matched to FA CS assets:")
        for _, row in unmatched.head(5).iterrows():
            print(f"   - Original ID: {row.get('Original Asset ID', 'N/A')}, "
                  f"Description: {str(row.get('Description', ''))[:40]}")
        if len(unmatched) > 5:
            print(f"   ... and {len(unmatched) - 5} more")
        print("   FIX: Import FA CS asset list or manually enter Asset #s\n")

    return result_df


def match_transfers_to_fa_cs(
    transfers_df: pd.DataFrame,
    fa_cs_lookup: Dict[str, int]
) -> pd.DataFrame:
    """
    Match transfer records to their existing FA CS Asset #s.

    Same as match_disposals_to_fa_cs but for transfers.
    """
    return match_disposals_to_fa_cs(transfers_df, fa_cs_lookup)


def export_separated_fa_cs(
    fa_df: pd.DataFrame,
    fa_cs_lookup: Optional[Dict[str, int]] = None
) -> bytes:
    """
    Export FA CS data with separate sheets for each transaction type.

    RECOMMENDED FOR RPA:
    This creates an Excel file with separate sheets:
    - "Additions" - New assets, ready for standard RPA import
    - "Disposals" - With FA CS Asset # lookup (if lookup provided)
    - "Transfers" - With FA CS Asset # lookup (if lookup provided)
    - "Existing" - Prior year assets (reference only)
    - "Unmatched" - Disposals/transfers that couldn't be matched
    - "Instructions" - RPA workflow guidance

    Args:
        fa_df: Full FA CS export DataFrame
        fa_cs_lookup: Optional lookup from build_fa_cs_asset_lookup()
                      Required for matching disposals/transfers

    Returns:
        bytes: Excel file with separated sheets
    """
    from io import BytesIO

    output = BytesIO()

    # Separate by transaction type
    separated = separate_by_transaction_type(fa_df)

    # Match disposals and transfers if lookup provided
    if fa_cs_lookup:
        if not separated["disposals"].empty:
            separated["disposals"] = match_disposals_to_fa_cs(
                separated["disposals"], fa_cs_lookup
            )
        if not separated["transfers"].empty:
            separated["transfers"] = match_transfers_to_fa_cs(
                separated["transfers"], fa_cs_lookup
            )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Additions - ready for import
        if not separated["additions"].empty:
            separated["additions"].to_excel(writer, sheet_name="Additions", index=False)
            ws = writer.sheets["Additions"]
            ws.freeze_panes = "A2"

        # Disposals - with matching info
        if not separated["disposals"].empty:
            separated["disposals"].to_excel(writer, sheet_name="Disposals", index=False)
            ws = writer.sheets["Disposals"]
            ws.freeze_panes = "A2"

        # Transfers - with matching info
        if not separated["transfers"].empty:
            separated["transfers"].to_excel(writer, sheet_name="Transfers", index=False)
            ws = writer.sheets["Transfers"]
            ws.freeze_panes = "A2"

        # Existing assets - reference
        if not separated["existing"].empty:
            separated["existing"].to_excel(writer, sheet_name="Existing", index=False)
            ws = writer.sheets["Existing"]
            ws.freeze_panes = "A2"

        # Unmatched items (disposals/transfers that need manual Asset # entry)
        unmatched_dfs = []
        if not separated["disposals"].empty and "Match Status" in separated["disposals"].columns:
            unmatched_disposals = separated["disposals"][
                separated["disposals"]["FA CS Asset #"].isna()
            ]
            if not unmatched_disposals.empty:
                unmatched_dfs.append(unmatched_disposals)

        if not separated["transfers"].empty and "Match Status" in separated["transfers"].columns:
            unmatched_transfers = separated["transfers"][
                separated["transfers"]["FA CS Asset #"].isna()
            ]
            if not unmatched_transfers.empty:
                unmatched_dfs.append(unmatched_transfers)

        if unmatched_dfs:
            unmatched_all = pd.concat(unmatched_dfs, ignore_index=True)
            unmatched_all.to_excel(writer, sheet_name="Unmatched_REVIEW", index=False)
            ws = writer.sheets["Unmatched_REVIEW"]
            ws.freeze_panes = "A2"

        # Instructions sheet
        instructions = pd.DataFrame({
            "RPA Workflow Instructions": [
                "=" * 60,
                "SEPARATED EXPORT FOR RPA AUTOMATION",
                "=" * 60,
                "",
                "SHEET: Additions",
                "  - NEW assets to add to FA CS",
                "  - Asset # is sequential (1, 2, 3...)",
                "  - RPA Workflow: CREATE new asset entry",
                "",
                "SHEET: Disposals",
                "  - Assets being disposed",
                "  - 'FA CS Asset #' = the EXISTING asset to update",
                "  - RPA Workflow: FIND asset by FA CS Asset #, then UPDATE disposal fields",
                "  - If 'Match Status' = 'NOT FOUND', manually look up Asset #",
                "",
                "SHEET: Transfers",
                "  - Assets being transferred/reclassified",
                "  - 'FA CS Asset #' = the EXISTING asset to update",
                "  - RPA Workflow: FIND asset by FA CS Asset #, then UPDATE location/category",
                "",
                "SHEET: Existing",
                "  - Prior year assets (reference only)",
                "  - Usually no RPA action needed",
                "",
                "SHEET: Unmatched_REVIEW",
                "  - Disposals/transfers that couldn't be matched to FA CS",
                "  - MANUAL ACTION: Look up Asset # in FA CS and enter manually",
                "",
                "=" * 60,
                "HOW TO BUILD FA CS LOOKUP:",
                "=" * 60,
                "1. Export current FA CS asset list (with Asset # and Original Asset ID)",
                "2. Use build_fa_cs_asset_lookup(export_df) to create lookup",
                "3. Pass lookup to export_separated_fa_cs()",
                "",
                "Or for first-time use:",
                "1. Run additions first (get Asset #s assigned)",
                "2. Save cross-reference file",
                "3. Use cross-reference for future disposal/transfer matching",
            ]
        })
        instructions.to_excel(writer, sheet_name="Instructions", index=False)

        # Summary sheet
        summary_data = {
            "Category": ["Additions", "Disposals", "Transfers", "Existing", "Unmatched"],
            "Count": [
                len(separated["additions"]),
                len(separated["disposals"]),
                len(separated["transfers"]),
                len(separated["existing"]),
                len(unmatched_all) if unmatched_dfs else 0,
            ],
            "RPA Action": [
                "CREATE new entries",
                "UPDATE existing (need FA CS Asset #)",
                "UPDATE existing (need FA CS Asset #)",
                "Reference only (usually skip)",
                "MANUAL - look up Asset # first",
            ],
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    output.seek(0)
    return output.getvalue()


def export_fa_ascii(fa_df: pd.DataFrame, delimiter: str = "\t") -> bytes:
    """
    Export Fixed Asset data to ASCII (text) format for FA CS import.

    FA CS supports ASCII file import as an alternative to Excel.
    This creates a simple delimited text file with the FA CS import columns.

    Args:
        fa_df: DataFrame with fixed asset data
        delimiter: Column delimiter - tab (\\t) or comma (,). Default is tab.

    Returns:
        bytes: ASCII file content as bytes

    Usage in FA CS:
        File → Import → ASCII File
        - Select delimiter type (Tab or Comma)
        - Map columns to FA CS fields
    """
    import io

    # FA CS Import columns (same as Excel export)
    fa_cs_cols = [
        "Asset #",
        "Description",
        "Date In Service",
        "Tax Cost",
        "Tax Method",
        "Tax Life",
        "Tax Sec 179 Expensed",
        "Tax Cur Depreciation",
        "Tax Prior Depreciation",
        "Book Cost",
        "Book Method",
        "Book Life",
        "MI Cost",
        "MI Method",
        "MI Life",
        "Date Disposed",
        "Gross Proceeds",
        "FA_CS_Wizard_Category",  # For UiPath RPA - wizard dropdown selection
    ]

    # Add columns if they don't exist
    if "Date Disposed" not in fa_df.columns:
        fa_df["Date Disposed"] = ""
    if "Gross Proceeds" not in fa_df.columns:
        fa_df["Gross Proceeds"] = ""

    # CRITICAL: Tax Cur Depreciation must be blank for FA CS to calculate
    fa_df["Tax Cur Depreciation"] = ""

    # Select only columns that exist
    available_cols = [c for c in fa_cs_cols if c in fa_df.columns]
    ascii_df = fa_df[available_cols].copy()

    # Convert to CSV/TSV format
    output = StringIO()

    # Determine separator name for file extension hint
    sep_name = "tab" if delimiter == "\t" else "comma"

    # Write header row
    output.write(delimiter.join(available_cols) + "\n")

    # Write data rows
    for _, row in ascii_df.iterrows():
        values = []
        for col in available_cols:
            val = row[col]
            # Handle None/NaN
            if pd.isna(val):
                val = ""
            else:
                val = str(val)
            # Quote values containing delimiter or newlines
            if delimiter in val or "\n" in val or '"' in val:
                val = '"' + val.replace('"', '""') + '"'
            values.append(val)
        output.write(delimiter.join(values) + "\n")

    return output.getvalue().encode('utf-8')


def export_fa_csv(fa_df: pd.DataFrame) -> bytes:
    """Export Fixed Asset data to CSV format (comma-delimited)."""
    return export_fa_ascii(fa_df, delimiter=",")


def export_fa_tsv(fa_df: pd.DataFrame) -> bytes:
    """Export Fixed Asset data to TSV format (tab-delimited)."""
    return export_fa_ascii(fa_df, delimiter="\t")


def _apply_professional_formatting(ws, df):
    """Apply professional Excel formatting to worksheet."""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # Header row formatting
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass

        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

    # Apply borders
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10000), min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Format currency columns
    currency_cols = [
        "Tax Cost", "Depreciable Basis", "Tax Sec 179 Expensed",
        "Bonus Amount", "Tax Cur Depreciation", "Tax Prior Depreciation",
        "De Minimis Expensed", "Section 179 Allowed", "Section 179 Carryforward",
        "Capital Gain", "Capital Loss", "§1245 Recapture (Ordinary Income)",
        "§1250 Recapture (Ordinary Income)", "Unrecaptured §1250 Gain (25%)",
        "Total Year 1 Deduction", "MaterialityScore"
    ]

    for col_name in currency_cols:
        if col_name in df.columns:
            col_idx = list(df.columns).index(col_name) + 1
            col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)

            for row in range(2, min(ws.max_row + 1, 10001)):
                try:
                    cell = ws[f'{col_letter}{row}']
                    if col_name == "MaterialityScore":
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '$#,##0.00'
                except (KeyError, AttributeError):
                    pass

    # Format date columns
    date_cols = ["Date In Service", "Acquisition Date", "Disposal Date"]

    for col_name in date_cols:
        if col_name in df.columns:
            col_idx = list(df.columns).index(col_name) + 1
            col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)

            for row in range(2, min(ws.max_row + 1, 10001)):
                try:
                    cell = ws[f'{col_letter}{row}']
                    cell.number_format = 'M/D/YYYY'
                except (KeyError, AttributeError):
                    pass


def _apply_conditional_formatting(ws, df):
    """Apply conditional formatting to highlight issues."""
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule

    # Red fill for critical issues
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

    # Orange fill for high priority
    orange_fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')

    # Yellow fill for medium priority
    yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    # Green fill for OK status
    green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

    # Apply to NBV_Reco column
    if "NBV_Reco" in df.columns:
        col_idx = list(df.columns).index("NBV_Reco") + 1
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)

        # Red for CHECK
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}10000',
            CellIsRule(operator='equal', formula=['"CHECK"'], fill=red_fill)
        )

        # Green for OK
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}10000',
            CellIsRule(operator='equal', formula=['"OK"'], fill=green_fill)
        )

    # Apply to ReviewPriority column
    if "ReviewPriority" in df.columns:
        col_idx = list(df.columns).index("ReviewPriority") + 1
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)

        # Orange for High
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}10000',
            CellIsRule(operator='equal', formula=['"High"'], fill=orange_fill)
        )

        # Yellow for Medium
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}10000',
            CellIsRule(operator='equal', formula=['"Medium"'], fill=yellow_fill)
        )

    # Apply to ConfidenceGrade column
    if "ConfidenceGrade" in df.columns:
        col_idx = list(df.columns).index("ConfidenceGrade") + 1
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)

        # Red for D grade
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}10000',
            CellIsRule(operator='equal', formula=['"D"'], fill=red_fill)
        )

        # Yellow for C grade
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}10000',
            CellIsRule(operator='equal', formula=['"C"'], fill=yellow_fill)
        )


# ==============================================================================
# PHASE 4: MULTI-YEAR DEPRECIATION PROJECTION EXPORT
# ==============================================================================

def export_depreciation_projection(
    fa_df: pd.DataFrame,
    tax_year: int,
    projection_years: int = 10,
    output_path: Optional[str] = None
) -> pd.DataFrame:
    """
    Generate and export multi-year depreciation projection.

    Creates a projection showing year-by-year depreciation for all assets
    over the specified number of years. Useful for tax planning and cashflow forecasting.

    Args:
        fa_df: Fixed asset dataframe from build_fa()
        tax_year: Current tax year
        projection_years: Number of years to project (default 10)
        output_path: Optional path to save Excel file

    Returns:
        DataFrame with depreciation projection summary by year
    """
    from .depreciation_projection import (
        project_portfolio_depreciation,
        export_projection_to_excel,
    )

    # Verify required columns exist
    required_cols = [
        "Depreciable Basis",
        "Recovery Period",
        "Method",
        "Convention",
        "Date In Service"
    ]

    missing = [col for col in required_cols if col not in fa_df.columns]
    if missing:
        raise ValueError(
            f"Cannot generate projection - missing columns: {missing}. "
            "Run build_fa() first to calculate MACRS parameters."
        )

    # Prepare dataframe for projection (rename columns to match projection module)
    projection_df = fa_df.copy()
    projection_df["In Service Date"] = projection_df["Date In Service"]

    # Generate projection summary
    print("\n" + "=" * 80)
    print(f"GENERATING {projection_years}-YEAR DEPRECIATION PROJECTION")
    print("=" * 80)

    summary_df = project_portfolio_depreciation(
        df=projection_df,
        current_tax_year=tax_year,
        projection_years=projection_years
    )

    # Print summary
    print(f"\nProjection Summary ({projection_years} years):")
    print("-" * 80)

    for _, row in summary_df.iterrows():
        year = int(row["Tax Year"])
        total_dep = row["Total Depreciation"]
        count = int(row["Assets Depreciating"])

        print(f"  {year}: ${total_dep:>12,.2f}  ({count} assets)")

    total_all_years = pd.to_numeric(summary_df["Total Depreciation"], errors='coerce').fillna(0).sum()
    print("-" * 80)
    print(f"  Total ({projection_years} years): ${total_all_years:>12,.2f}")
    print("=" * 80 + "\n")

    # Export to Excel if path provided
    if output_path:
        export_projection_to_excel(
            df=projection_df,
            current_tax_year=tax_year,
            output_path=output_path,
            projection_years=projection_years
        )

    return summary_df


def generate_comprehensive_depreciation_report(
    fa_df: pd.DataFrame,
    tax_year: int,
    output_dir: str = ".",
    projection_years: int = 10
) -> dict:
    """
    Generate comprehensive depreciation report with all Phase 4 features.

    Creates multiple Excel files:
    1. Fixed Asset Register (standard export)
    2. Multi-Year Depreciation Projection
    3. Section 179 Carryforward Schedule (if applicable)

    Args:
        fa_df: Fixed asset dataframe from build_fa()
        tax_year: Current tax year
        output_dir: Directory to save output files
        projection_years: Number of years to project

    Returns:
        Dict with paths to generated files
    """
    import os
    from .depreciation_projection import export_projection_to_excel
    from .section_179_carryforward import Section179CarryforwardTracker

    output_files = {}

    # 1. Fixed Asset Register
    fa_register_path = os.path.join(output_dir, f"Fixed_Asset_Register_{tax_year}.xlsx")
    fa_bytes = export_fa_excel(fa_df)
    with open(fa_register_path, "wb") as f:
        f.write(fa_bytes)
    output_files["fixed_asset_register"] = fa_register_path
    print(f"✓ Fixed Asset Register saved: {fa_register_path}")

    # 2. Multi-Year Depreciation Projection
    projection_path = os.path.join(output_dir, f"Depreciation_Projection_{tax_year}_{projection_years}yr.xlsx")

    # Prepare dataframe
    projection_df = fa_df.copy()
    projection_df["In Service Date"] = projection_df["Date In Service"]

    export_projection_to_excel(
        df=projection_df,
        current_tax_year=tax_year,
        output_path=projection_path,
        projection_years=projection_years
    )
    output_files["depreciation_projection"] = projection_path

    # 3. Section 179 Carryforward Schedule (if applicable)
    if "Section 179 Carryforward" in fa_df.columns:
        total_carryforward = pd.to_numeric(fa_df["Section 179 Carryforward"], errors='coerce').fillna(0).sum()

        if total_carryforward > 0:
            carryforward_path = os.path.join(output_dir, f"Section_179_Carryforward_{tax_year}.xlsx")

            # Create carryforward dataframe - use correct column names that match fa_export output
            # Build list of columns that exist in the dataframe
            carryforward_cols = []
            col_mapping = {
                "Asset #": ["Asset #", "Asset ID"],  # Try new name first, then old
                "Description": ["Description", "Property Description"],
                "Date In Service": ["Date In Service"],
                "Tax Cost": ["Tax Cost", "Cost/Basis", "Cost"],
                "Tax Sec 179 Expensed": ["Tax Sec 179 Expensed", "Section 179 Amount"],
                "Section 179 Allowed": ["Section 179 Allowed"],
                "Section 179 Carryforward": ["Section 179 Carryforward"]
            }

            for target_col, possible_names in col_mapping.items():
                for name in possible_names:
                    if name in fa_df.columns:
                        carryforward_cols.append(name)
                        break

            carryforward_df = fa_df[pd.to_numeric(fa_df["Section 179 Carryforward"], errors='coerce').fillna(0) > 0][carryforward_cols].copy()

            carryforward_df.to_excel(carryforward_path, index=False, sheet_name="Section 179 Carryforward")
            output_files["section_179_carryforward"] = carryforward_path
            print(f"✓ Section 179 Carryforward Schedule saved: {carryforward_path}")
            print(f"  Total carryforward to {tax_year + 1}: ${total_carryforward:,.2f}")

    print("\n" + "=" * 80)
    print(f"COMPREHENSIVE DEPRECIATION REPORT COMPLETE")
    print("=" * 80)
    print(f"Files generated: {len(output_files)}")
    for report_type, path in output_files.items():
        print(f"  - {report_type}: {path}")
    print("=" * 80 + "\n")

    return output_files


# ==============================================================================
# PRIOR YEAR DATA LOADING UTILITIES
# ==============================================================================

def load_section_179_carryforward_from_file(file_path: str = "prior_year_data.json") -> float:
    """
    Load Section 179 carryforward from prior year data file.

    The file should contain JSON with structure:
    {
        "tax_year": 2023,
        "section_179_carryforward": 12500.00
    }

    Args:
        file_path: Path to prior year data JSON file

    Returns:
        Section 179 carryforward amount (0.0 if file not found or invalid)

    Example:
        >>> carryforward = load_section_179_carryforward_from_file("data/2023_carryforward.json")
        >>> df = build_fa(..., section_179_carryforward_from_prior_year=carryforward)
    """
    import json
    import os

    if not os.path.exists(file_path):
        print(f"ℹ️  No prior year carryforward file found at {file_path}")
        print(f"   Using $0 carryforward (first year or no carryforward)")
        return 0.0

    try:
        with open(file_path, 'r') as f:
            data = json.load(f)

        carryforward = float(data.get("section_179_carryforward", 0.0))

        if carryforward > 0:
            tax_year = data.get("tax_year", "unknown")
            print(f"✓ Loaded Section 179 carryforward: ${carryforward:,.2f} from {tax_year}")

        return carryforward

    except Exception as e:
        print(f"⚠️  Error loading prior year carryforward from {file_path}: {e}")
        print(f"   Using $0 carryforward")
        return 0.0


def save_section_179_carryforward_to_file(
    tax_year: int,
    carryforward_amount: float,
    file_path: str = "prior_year_data.json"
):
    """
    Save Section 179 carryforward data for next year's processing.

    Call this after generating exports to save carryforward data for next year.

    Args:
        tax_year: Current tax year
        carryforward_amount: Section 179 carryforward to next year
        file_path: Path to save data

    Example:
        >>> # After processing 2024 returns
        >>> save_section_179_carryforward_to_file(2024, 15000.00, "data/2024_carryforward.json")
        >>> # Next year, load it:
        >>> cf = load_section_179_carryforward_from_file("data/2024_carryforward.json")
    """
    import json
    import os

    data = {
        "tax_year": tax_year,
        "section_179_carryforward": carryforward_amount,
        "generated_date": datetime.now().isoformat(),
        "note": f"Section 179 carryforward from {tax_year} to {tax_year + 1} (Form 4562 Part I Line 13)"
    }

    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else ".", exist_ok=True)

    with open(file_path, 'w') as f:
        json.dump(data, f, indent=2)

    print(f"✓ Saved Section 179 carryforward data to {file_path}")
    print(f"  Tax Year: {tax_year}")
    print(f"  Carryforward to {tax_year + 1}: ${carryforward_amount:,.2f}")
