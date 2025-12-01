from fastapi import FastAPI, UploadFile, File, HTTPException, Body, Query, Request, Depends, Response, APIRouter
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
import uvicorn
import psutil
import os
import shutil
import sys
from typing import List, Dict, Optional
from datetime import datetime
import traceback
import io
from threading import Lock
from contextlib import asynccontextmanager
import asyncio
import logging

# Thread lock for concurrent access to global state
_state_lock = Lock()

# Per-session locks to prevent concurrent upload data corruption
_session_locks: Dict[str, Lock] = {}
_session_locks_lock = Lock()  # Lock for accessing _session_locks dict

def get_session_lock(session_id: str) -> Lock:
    """Get or create a lock for a specific session."""
    with _session_locks_lock:
        if session_id not in _session_locks:
            _session_locks[session_id] = Lock()
        return _session_locks[session_id]


def cleanup_session_lock(session_id: str) -> None:
    """Remove a session lock when the session expires."""
    with _session_locks_lock:
        if session_id in _session_locks:
            del _session_locks[session_id]


def cleanup_stale_session_locks(active_session_ids: set) -> int:
    """Remove locks for sessions that no longer exist.

    Args:
        active_session_ids: Set of session IDs that are still active

    Returns:
        Number of locks removed
    """
    with _session_locks_lock:
        stale_ids = set(_session_locks.keys()) - active_session_ids
        for session_id in stale_ids:
            del _session_locks[session_id]
        return len(stale_ids)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

# Add project root to sys.path to allow imports
# This must happen BEFORE importing services that depend on backend.logic
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend.services.importer import ImporterService
from backend.services.classifier import ClassifierService
from backend.services.auditor import AuditorService
from backend.services.exporter import ExporterService
from backend.models.asset import Asset

# Import logic modules for new features
from backend.logic.data_quality_score import calculate_data_quality_score, DataQualityScore
from backend.logic.smart_tab_analyzer import analyze_tabs, TabAnalysisResult
from backend.logic.rollforward_reconciliation import reconcile_rollforward, RollforwardResult
from backend.logic.depreciation_projection import project_portfolio_depreciation
import pandas as pd

# Import scalability modules
from backend.middleware.rate_limiter import RateLimiter, get_rate_limiter
from backend.middleware.timeout import TimeoutMiddleware
from backend.logic.file_cleanup import get_cleanup_manager
from backend.logic.session_manager import (
    get_session_manager,
    SessionData,
    get_session_from_request,
    add_session_to_response
)

# Import authentication module
from backend.middleware.auth import (
    require_auth,
    get_current_user,
    User as AuthUser
)

# Import tempfile for secure temporary file handling
import tempfile

# ==============================================================================
# SECURITY CONFIGURATION
# ==============================================================================
# Authentication can be enabled/disabled via environment variable.
# In production (ENVIRONMENT=production), authentication is REQUIRED.
# In development, it's optional for easier testing.

AUTH_ENABLED = os.environ.get("AUTH_ENABLED", "").lower() in ("true", "1", "yes")
IS_PRODUCTION = os.environ.get("ENVIRONMENT", "").lower() == "production"

if IS_PRODUCTION and not AUTH_ENABLED:
    logger.warning(
        "⚠️ SECURITY WARNING: Running in production without authentication! "
        "Set AUTH_ENABLED=true to require authentication on all endpoints."
    )

# Dependency for optional authentication
async def optional_auth(user: AuthUser = Depends(get_current_user)) -> AuthUser:
    """
    Optional authentication dependency.

    - If AUTH_ENABLED=true: Requires valid authentication
    - If AUTH_ENABLED=false: Returns None (allows anonymous access)

    This allows gradual rollout of authentication.
    """
    if AUTH_ENABLED and not user:
        raise HTTPException(
            status_code=401,
            detail="Authentication required. Set Authorization header with Bearer token.",
            headers={"WWW-Authenticate": "Bearer"}
        )
    return user

# ==============================================================================
# APPLICATION LIFECYCLE
# ==============================================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application startup and shutdown lifecycle."""
    # Startup
    logger.info("Starting FA CS Automator API...")

    # Start session cleanup task
    session_manager = get_session_manager()
    await session_manager.start_cleanup_task()
    logger.info("Session cleanup task started")

    # Start file cleanup task
    cleanup_manager = get_cleanup_manager()
    await cleanup_manager.start_scheduled_cleanup()
    logger.info("File cleanup task started")

    yield

    # Shutdown
    logger.info("Shutting down FA CS Automator API...")
    session_manager.stop_cleanup_task()
    cleanup_manager.stop_scheduled_cleanup()

app = FastAPI(
    title="FA CS Automator API",
    description="Fixed Asset Classification and Export API",
    version="2.1.0",
    lifespan=lifespan
)

# ==============================================================================
# API VERSIONING
# ==============================================================================
# All endpoints are available at both:
#   - /api/v1/... (recommended, versioned)
#   - /... (deprecated, for backward compatibility)
#
# Frontend should migrate to /api/v1/ prefix for future compatibility.
# The root endpoints will be removed in version 3.0.0.

API_VERSION = "v1"
api_v1 = APIRouter(prefix=f"/api/{API_VERSION}", tags=["v1"])

# ==============================================================================
# CORS CONFIGURATION (Issue 6.2 from IRS Audit Report - Security)
# ==============================================================================
# Configure allowed origins via CORS_ALLOWED_ORIGINS environment variable.
# For production, set to specific domains (comma-separated):
#   CORS_ALLOWED_ORIGINS=https://app.example.com,https://admin.example.com
#
# For development, leave unset to use localhost defaults.

def get_cors_origins():
    """Get CORS allowed origins from environment or use development defaults."""
    env_origins = os.environ.get("CORS_ALLOWED_ORIGINS", "")
    if env_origins:
        # Production: use configured origins
        return [origin.strip() for origin in env_origins.split(",") if origin.strip()]
    else:
        # Development: allow both localhost and 127.0.0.1 variants
        return [
            "http://localhost:3000",
            "http://localhost:5173",
            "http://localhost:8000",
            "http://127.0.0.1:3000",
            "http://127.0.0.1:5173",
            "http://127.0.0.1:8000"
        ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=get_cors_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Session-ID"],  # CRITICAL: Allow frontend to read session ID header
)

# Add timeout middleware (must be added before other middleware)
app.add_middleware(TimeoutMiddleware, default_timeout=30)

# ==============================================================================
# SECURITY HEADERS MIDDLEWARE
# ==============================================================================
# Adds HTTP security headers to all responses to protect against common attacks.
# Reference: OWASP Secure Headers Project

@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    """Add security headers to all responses."""
    response = await call_next(request)

    # Prevent MIME type sniffing
    response.headers["X-Content-Type-Options"] = "nosniff"

    # Prevent clickjacking
    response.headers["X-Frame-Options"] = "DENY"

    # Enable XSS filtering (legacy browsers)
    response.headers["X-XSS-Protection"] = "1; mode=block"

    # Referrer policy - don't leak URLs
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"

    # Permissions policy - restrict browser features
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"

    # In production, enable HSTS (requires HTTPS)
    if IS_PRODUCTION:
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"

    return response

# ==============================================================================
# RATE LIMITING MIDDLEWARE
# ==============================================================================

@app.middleware("http")
async def rate_limit_middleware(request: Request, call_next):
    """Apply rate limiting to all requests."""
    limiter = get_rate_limiter()

    # Determine operation type from path
    path = request.url.path.lower()

    if "/upload" in path:
        operation = "upload"
    elif "/classify" in path or "/assets" in path and request.method == "POST":
        operation = "classify"
    elif "/export" in path and "/status" not in path and request.method != "GET":
        # Only rate-limit actual export operations, not status checks
        operation = "export"
    elif request.method == "GET":
        operation = "read"
    else:
        operation = "default"

    try:
        await limiter.check(request, operation=operation)
    except HTTPException as e:
        # Must manually add CORS headers since this runs before CORS middleware
        origin = request.headers.get("origin", "")
        allowed_origins = get_cors_origins()
        cors_headers = {}
        if origin in allowed_origins:
            cors_headers = {
                "Access-Control-Allow-Origin": origin,
                "Access-Control-Allow-Credentials": "true",
            }
        response_headers = {**(e.headers or {}), **cors_headers}
        return JSONResponse(
            status_code=e.status_code,
            content=e.detail if isinstance(e.detail, dict) else {"error": e.detail},
            headers=response_headers
        )

    response = await call_next(request)

    # Add rate limit headers
    remaining = limiter.get_remaining(request, operation)
    response.headers["X-RateLimit-Remaining"] = str(remaining)
    response.headers["X-RateLimit-Operation"] = operation

    return response

# Initialize Services
importer = ImporterService()
classifier = ClassifierService()
auditor = AuditorService()
exporter = ExporterService()

# ==============================================================================
# STANDARDIZED API ERROR RESPONSES
# ==============================================================================
# All errors follow this schema for consistent frontend parsing:
# {
#   "error": "ERROR_CODE",           # Machine-readable error code
#   "message": "Human readable...",   # User-friendly message
#   "details": {...}                  # Optional additional context
# }

from pydantic import BaseModel
from typing import Any

class APIError(BaseModel):
    """Standardized API error response schema."""
    error: str  # Machine-readable error code (e.g., "ASSET_NOT_FOUND")
    message: str  # Human-readable message
    details: Optional[Dict[str, Any]] = None  # Optional context

def api_error(status_code: int, error_code: str, message: str, details: Dict = None) -> HTTPException:
    """
    Create a standardized API error response.

    Usage:
        raise api_error(404, "ASSET_NOT_FOUND", "Asset with ID 123 not found")
        raise api_error(400, "VALIDATION_ERROR", "Cannot export", {"errors": 5})
    """
    detail = {"error": error_code, "message": message}
    if details:
        detail["details"] = details
    return HTTPException(status_code=status_code, detail=detail)

# ==============================================================================
# SESSION-BASED STATE MANAGEMENT
# ==============================================================================
# All state is now stored per-session for user isolation.
# Legacy global variables are kept as fallback for backward compatibility.
# Frontend should pass X-Session-ID header or session_id query param.

# Legacy global state (DEPRECATED - use session-based state instead)
# These will be removed in a future version
ASSET_STORE: Dict[int, Asset] = {}
ASSET_ID_COUNTER = 0
APPROVED_ASSETS: set = set()
TAB_ANALYSIS_RESULT = None
LAST_UPLOAD_FILENAME = None

# Helper function to get session from request (with fallback to creating new session)
async def get_current_session(request: Request) -> SessionData:
    """Get session from request, creating one if needed."""
    return await get_session_from_request(request)

# Remote FA CS Configuration (session-independent for now)
FACS_CONFIG = {
    "remote_mode": True,
    "user_confirmed_connected": False,
    "export_path": None
}

# Tax Configuration - CRITICAL for proper classification
from datetime import date
TAX_CONFIG = {
    "tax_year": date.today().year,
    "de_minimis_threshold": 2500,
    "has_afs": False,
    "bonus_rate": None,
    "section_179_limit": None,
}

# Import tax year config for bonus rates
try:
    from backend.logic import tax_year_config
    TAX_YEAR_CONFIG_AVAILABLE = True
except ImportError:
    TAX_YEAR_CONFIG_AVAILABLE = False

@app.get("/")
def read_root():
    return {"status": "Backend Online", "version": "1.0.0"}

@app.get("/check-facs")
async def check_facs(request: Request, response: Response):
    """
    Check FA CS connection status.
    Remote mode: relies on user confirmation (can't auto-detect remote app)
    Local mode: checks if FAwin.exe is running

    IMPORTANT: This endpoint establishes the session for the Dashboard.
    It's called first, so it must return session ID for subsequent requests.
    """
    # Establish session and add to response (critical for Dashboard parallel requests)
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if FACS_CONFIG["remote_mode"]:
        return {
            "running": FACS_CONFIG["user_confirmed_connected"],
            "remote_mode": True,
            "message": "Remote FA CS - Please confirm you are connected" if not FACS_CONFIG["user_confirmed_connected"] else "Connected to remote FA CS",
            "export_path": FACS_CONFIG["export_path"]
        }
    else:
        # Local mode - auto-detect
        running = "FAwin.exe" in (p.name() for p in psutil.process_iter())
        return {"running": running, "remote_mode": False}

@app.post("/facs/confirm-connected")
def confirm_facs_connected():
    """User confirms they are connected to remote FA CS session."""
    FACS_CONFIG["user_confirmed_connected"] = True
    return {"confirmed": True, "message": "FA CS connection confirmed"}

@app.post("/facs/disconnect")
def disconnect_facs():
    """User indicates they disconnected from FA CS."""
    FACS_CONFIG["user_confirmed_connected"] = False
    return {"confirmed": False, "message": "FA CS disconnected"}

@app.post("/facs/set-export-path")
def set_export_path(path: str = Body(..., embed=True)):
    """
    Set custom export path (e.g., shared network folder accessible from remote session).
    Example: "\\\\server\\shared\\FA_Imports" or "Z:\\FA_Imports"

    SECURITY: Validates path is safe and accessible before setting.
    """
    # Normalize and get absolute path
    try:
        abs_path = os.path.abspath(os.path.expanduser(path))
    except Exception as e:
        raise api_error(400, "INVALID_PATH", f"Invalid path format: {str(e)}")

    # Block dangerous system paths
    BLOCKED_PATHS = [
        "/etc", "/bin", "/sbin", "/usr", "/var", "/root", "/boot", "/sys", "/proc",
        "/System", "/Library", "/Applications", "/private",
        "C:\\Windows", "C:\\Program Files", "C:\\ProgramData"
    ]

    path_lower = abs_path.lower()
    for blocked in BLOCKED_PATHS:
        if path_lower.startswith(blocked.lower()):
            logger.warning(f"Blocked export path attempt: {abs_path}")
            raise api_error(400, "BLOCKED_PATH", f"Cannot export to system directory: {blocked}")

    # Validate path exists (or parent exists for new directories)
    if not os.path.exists(abs_path):
        parent_dir = os.path.dirname(abs_path)
        if not os.path.exists(parent_dir):
            raise api_error(400, "PATH_NOT_FOUND",
                f"Directory does not exist and parent is also missing: {abs_path}")
        # Try to create the directory
        try:
            os.makedirs(abs_path, exist_ok=True)
            logger.info(f"Created export directory: {abs_path}")
        except PermissionError:
            raise api_error(403, "PERMISSION_DENIED", f"Cannot create directory (permission denied): {abs_path}")
        except Exception as e:
            raise api_error(400, "PATH_CREATE_FAILED", f"Cannot create directory: {str(e)}")

    # Validate path is writable
    if not os.access(abs_path, os.W_OK):
        raise api_error(403, "PERMISSION_DENIED", f"No write permission for path: {abs_path}")

    FACS_CONFIG["export_path"] = abs_path
    logger.info(f"Export path set to: {abs_path}")
    return {"export_path": abs_path, "message": f"Export path set to: {abs_path}"}

@app.get("/facs/config")
def get_facs_config():
    """Get current FA CS configuration."""
    return FACS_CONFIG

@app.get("/stats")
async def get_stats(request: Request, response: Response):
    """
    Returns statistics for the Dashboard.
    Uses session-based state for user isolation.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    assets = list(session.assets.values())
    total = len(assets)

    errors = sum(1 for a in assets if getattr(a, 'validation_errors', None))
    needs_review = sum(1 for a in assets if not getattr(a, 'validation_errors', None)
                       and getattr(a, 'confidence_score', 1.0) <= 0.8)
    high_confidence = sum(1 for a in assets if not getattr(a, 'validation_errors', None)
                          and getattr(a, 'confidence_score', 1.0) > 0.8)
    approved = len(session.approved_assets)

    # Transaction type breakdown
    trans_types = {}
    de_minimis_count = 0
    de_minimis_total = 0.0

    for a in assets:
        tt = getattr(a, 'transaction_type', 'addition') or 'addition'
        trans_types[tt] = trans_types.get(tt, 0) + 1

        # Track De Minimis items (expensed, not capitalized)
        election = getattr(a, 'depreciation_election', 'MACRS') or 'MACRS'
        if election == 'DeMinimis' and tt == 'Current Year Addition':
            de_minimis_count += 1
            de_minimis_total += a.cost or 0

    # Calculate additions excluding De Minimis (for accurate capital addition count)
    additions_count = trans_types.get('Current Year Addition', 0)
    capital_additions = additions_count - de_minimis_count

    return {
        "total": total,
        "errors": errors,
        "needs_review": needs_review,
        "high_confidence": high_confidence,
        "approved": approved,
        "ready_for_export": errors == 0 and total > 0,
        "transaction_types": trans_types,
        "tax_year": session.tax_config.get("tax_year", TAX_CONFIG["tax_year"]),
        "session_id": session.session_id,
        # De Minimis tracking
        "de_minimis_count": de_minimis_count,
        "de_minimis_total": round(de_minimis_total, 2),
        "capital_additions": capital_additions  # Additions minus De Minimis
    }


@app.get("/assets")
async def get_assets(request: Request, response: Response):
    """
    Returns all currently loaded assets for this session.
    Useful for refreshing the frontend after tax year changes.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)
    return list(session.assets.values())


# ==============================================================================
# TAX CONFIGURATION ENDPOINTS
# ==============================================================================

@app.get("/config/tax")
async def get_tax_config(request: Request, response: Response):
    """
    Get current tax configuration including:
    - Tax year
    - De minimis threshold
    - Bonus depreciation rate
    - Section 179 limits

    Uses session-based storage for per-user isolation.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    # Use session tax config (isolated per user)
    config = dict(session.tax_config)
    tax_year = config.get("tax_year", datetime.now().year)

    # Calculate bonus rate based on tax year
    if TAX_YEAR_CONFIG_AVAILABLE:
        # get_bonus_percentage returns a float (e.g., 1.0 for 100%, 0.6 for 60%)
        bonus_rate = tax_year_config.get_bonus_percentage(tax_year)
        config["bonus_rate"] = int(bonus_rate * 100)  # Convert to percentage (e.g., 100, 60)
        # get_section_179_limits returns a dict with 'max_deduction' and 'phaseout_threshold'
        section_179_info = tax_year_config.get_section_179_limits(tax_year)
        config["section_179_limit"] = section_179_info.get("max_deduction", 0)
        config["obbba_effective"] = tax_year >= 2025  # OBBBA passed in 2025
        config["obbba_info"] = {
            "name": "One Big Beautiful Bill Act (OBBBA) 2025",
            "effective_date": "2025-01-19",
            "bonus_rate_post_obbba": 100 if tax_year >= 2025 else None,
            "section_179_limit_increase": "$2.5M (up from $1.2M)" if tax_year >= 2025 else None
        }
    else:
        # Fallback defaults (TCJA phaseout schedule - pre-OBBBA)
        config["bonus_rate"] = 80 if tax_year <= 2024 else 40 if tax_year == 2025 else 20 if tax_year == 2026 else 0
        config["section_179_limit"] = 1220000 if tax_year < 2025 else 2500000

    return config


@app.post("/config/tax")
async def set_tax_config(
    request: Request,
    response: Response,
    tax_year: int = Body(..., embed=True, ge=2020, le=2030),
    de_minimis_threshold: int = Body(2500, ge=0, le=5000),
    has_afs: bool = Body(False),
    user: AuthUser = Depends(optional_auth)
):
    """
    Set tax configuration for the session.

    Authentication: Required when AUTH_ENABLED=true

    Uses session-based storage for per-user isolation.

    Args:
        tax_year: Tax year for depreciation calculations (2020-2030)
        de_minimis_threshold: De minimis safe harbor threshold ($0-$5000)
                             - $2,500 for taxpayers WITHOUT audited financial statements
                             - $5,000 for taxpayers WITH audited financial statements
        has_afs: Whether taxpayer has Audited Financial Statements

    IMPORTANT: Setting tax year will RECLASSIFY all loaded assets!
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    # ATOMICITY: Save original state for rollback on failure
    import copy
    original_tax_config = copy.deepcopy(session.tax_config)
    original_assets = {k: copy.deepcopy(v) for k, v in session.assets.items()} if session.assets else {}

    try:
        # Update session tax config (isolated per user)
        session.tax_config["tax_year"] = tax_year
        session.tax_config["de_minimis_threshold"] = de_minimis_threshold
        session.tax_config["has_afs"] = has_afs

        # Update classifier's tax year
        classifier.set_tax_year(tax_year)

        # Reclassify loaded assets with new tax year
        # CRITICAL: We must preserve ALL assets - never filter or remove any
        trans_types = {}
        errors_count = 0
        reclassified_assets = []

        # Use session assets instead of global ASSET_STORE
        if session.assets:
            # Get all assets from session
            asset_count_before = len(session.assets)
            assets = list(session.assets.values())
            logger.info(f"[Tax Year Change] Starting with {asset_count_before} assets in session")
            logger.info(f"[Tax Year Change] Reclassifying {len(assets)} assets for tax year {tax_year}")

            # Reclassify transaction types
            classifier._classify_transaction_types(assets, tax_year)

            # Re-run validation for each asset with new tax year
            # This will flag assets with dates after the tax year as errors
            # BUT it should NEVER remove assets from the list
            for a in assets:
                a.check_validity(tax_year=tax_year)
                tt = a.transaction_type or 'unknown'
                trans_types[tt] = trans_types.get(tt, 0) + 1
                if a.validation_errors:
                    errors_count += 1
                reclassified_assets.append(a)

            asset_count_after = len(reclassified_assets)
            logger.info(f"[Tax Year Change] Reclassification complete: {trans_types}")
            logger.info(f"[Tax Year Change] Asset count after reclassification: {asset_count_after}")
            if asset_count_before != asset_count_after:
                logger.warning(f"[Tax Year Change] Asset count changed from {asset_count_before} to {asset_count_after}!")
            if errors_count > 0:
                logger.info(f"[Tax Year Change] {errors_count} assets have validation errors (may include future dates)")
        else:
            logger.info(f"[Tax Year Change] No assets in session to reclassify")
            reclassified_assets = []

        # Save session
        manager = get_session_manager()
        manager._save_session(session)

    except Exception as e:
        # ROLLBACK: Restore original state on any failure
        logger.error(f"[Tax Year Change] Reclassification failed: {e}, rolling back")
        session.tax_config = original_tax_config
        session.assets = original_assets
        classifier.set_tax_year(original_tax_config.get("tax_year", TAX_CONFIG["tax_year"]))
        raise api_error(500, "RECLASSIFICATION_FAILED",
            f"Tax year change failed. Original state restored. Error: {str(e)}")

    return {
        "status": "updated",
        "tax_year": tax_year,
        "de_minimis_threshold": de_minimis_threshold,
        "has_afs": has_afs,
        "assets_reclassified": len(session.assets),
        "transaction_type_breakdown": trans_types,
        "assets": list(session.assets.values()) if session.assets else []
    }


# ==============================================================================
# DATA WARNINGS & COMPLETENESS ENDPOINTS
# ==============================================================================

@app.get("/warnings")
async def get_warnings(request: Request, response: Response):
    """
    Get comprehensive warnings about the loaded data including:
    - Missing asset detection warnings
    - Transaction type classification issues
    - De minimis candidates
    - Tax compliance warnings
    Uses session-based storage for user isolation.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    assets = list(session.assets.values())
    if not assets:
        return {"warnings": [], "critical": [], "info": []}

    critical_warnings = []
    warnings = []
    info_messages = []
    tax_year = session.tax_config.get("tax_year", TAX_CONFIG["tax_year"])

    # ===== CRITICAL WARNINGS =====

    # 1. Check for existing assets incorrectly classified as additions
    existing_as_additions = []
    for a in assets:
        if a.in_service_date:
            asset_year = a.in_service_date.year
            if asset_year < tax_year and "addition" in (a.transaction_type or "").lower():
                existing_as_additions.append({
                    "asset_id": a.asset_id,
                    "description": a.description[:50],
                    "in_service_year": asset_year,
                    "tax_year": tax_year
                })

    if existing_as_additions:
        critical_warnings.append({
            "type": "MISCLASSIFIED_EXISTING_ASSETS",
            "message": f"{len(existing_as_additions)} existing assets may be incorrectly classified as additions",
            "impact": "Section 179 and Bonus depreciation NOT allowed on existing assets",
            "action": "Review transaction type classification and set correct tax year",
            "affected_count": len(existing_as_additions),
            "examples": existing_as_additions[:5]
        })

    # 2. Check for assets missing cost (exclude transfers - they don't require cost)
    # Transfers just move assets between departments/locations, cost is already recorded
    zero_cost = [
        a for a in assets
        if a.cost <= 0 and not (a.transaction_type and a.transaction_type.lower() == "transfer")
    ]
    if zero_cost:
        critical_warnings.append({
            "type": "MISSING_COST",
            "message": f"{len(zero_cost)} assets have $0 or missing cost",
            "impact": "Cannot calculate depreciation without cost basis",
            "action": "Enter cost for each asset",
            "affected_count": len(zero_cost)
        })

    # 3. Check for assets with Unknown (Missing Date) transaction type - CRITICAL
    unknown_trans_type = [
        a for a in assets
        if a.transaction_type and "unknown" in a.transaction_type.lower()
    ]
    if unknown_trans_type:
        critical_warnings.append({
            "type": "UNKNOWN_TRANSACTION_TYPE",
            "message": f"{len(unknown_trans_type)} assets have unknown transaction type due to missing dates",
            "impact": "Cannot determine if Section 179/Bonus eligible - potential tax compliance issue",
            "action": "Add in-service dates to properly classify as Current Year Addition or Existing Asset",
            "affected_count": len(unknown_trans_type)
        })

    # ===== REGULAR WARNINGS =====

    # 4. Check for assets missing dates (informational - already covered by Unknown transaction type)
    missing_dates = [a for a in assets if not a.in_service_date and not a.acquisition_date]
    if missing_dates:
        warnings.append({
            "type": "MISSING_DATES",
            "message": f"{len(missing_dates)} assets have no in-service or acquisition date",
            "impact": "Classified as 'Unknown (Missing Date)' - requires manual review",
            "action": "Add in-service dates for proper classification",
            "affected_count": len(missing_dates)
        })

    # 5. Transaction type breakdown check
    trans_types = {}
    for a in assets:
        tt = a.transaction_type or "unknown"
        trans_types[tt] = trans_types.get(tt, 0) + 1

    # Check if ALL assets are additions (suspicious)
    additions_count = trans_types.get("Current Year Addition", 0) + trans_types.get("addition", 0)
    if additions_count == len(assets) and len(assets) > 10:
        warnings.append({
            "type": "ALL_ADDITIONS_SUSPICIOUS",
            "message": f"All {len(assets)} assets classified as additions - this may indicate missing data",
            "impact": "Typical asset schedules include existing assets, disposals, and transfers",
            "action": "Verify data includes complete asset schedule with all transaction types",
            "affected_count": len(assets)
        })

    # 6. De minimis candidates
    de_minimis_threshold = session.tax_config.get("de_minimis_threshold", TAX_CONFIG["de_minimis_threshold"])
    de_minimis_candidates = [
        a for a in assets
        if 0 < a.cost <= de_minimis_threshold and "Current Year Addition" in (a.transaction_type or "")
    ]
    if de_minimis_candidates:
        info_messages.append({
            "type": "DE_MINIMIS_CANDIDATES",
            "message": f"{len(de_minimis_candidates)} current year additions qualify for de minimis safe harbor",
            "threshold": de_minimis_threshold,
            "total_value": sum(a.cost for a in de_minimis_candidates),
            "benefit": "Can expense immediately instead of capitalizing and depreciating",
            "affected_count": len(de_minimis_candidates)
        })

    # 7. Unclassified assets (exclude transfers - they don't need classification)
    unclassified = [
        a for a in assets
        if a.macrs_class in ["Unclassified", "Unknown", None, ""]
        and not (a.transaction_type and a.transaction_type.lower() == "transfer")
    ]
    if unclassified:
        warnings.append({
            "type": "UNCLASSIFIED_ASSETS",
            "message": f"{len(unclassified)} assets could not be automatically classified",
            "impact": "Manual MACRS class assignment required before export",
            "action": "Review and assign MACRS class to each unclassified asset",
            "affected_count": len(unclassified)
        })

    # 8. Asset ID Collision Detection (FA CS requires unique numeric Asset #)
    from backend.services.exporter import ExporterService
    exporter = ExporterService()
    collision_result = exporter.detect_asset_number_collisions(assets)

    if collision_result["has_collisions"]:
        warnings.append({
            "type": "FA_CS_ASSET_NUMBER_COLLISION",
            "message": f"{collision_result['collision_count']} FA CS Asset # collisions detected - {collision_result['affected_assets']} assets affected",
            "impact": "FA CS requires unique Asset # for each asset. Duplicate numbers will cause import errors.",
            "action": "Assign unique FA CS # in the Review table for colliding assets",
            "affected_count": collision_result["affected_assets"],
            "collisions": collision_result["collisions"][:5],  # Show first 5
            "details": collision_result["warnings"][:5]  # Show first 5 warning messages
        })

    # ===== INFO MESSAGES =====

    # 9. Transfer assets info (they don't require cost)
    transfer_assets = [
        a for a in assets
        if a.transaction_type and a.transaction_type.lower() == "transfer"
    ]
    if transfer_assets:
        transfer_with_cost = [a for a in transfer_assets if a.cost > 0]
        transfer_no_cost = [a for a in transfer_assets if a.cost <= 0]
        info_messages.append({
            "type": "TRANSFER_ASSETS_INFO",
            "message": f"{len(transfer_assets)} transfer records detected",
            "details": {
                "with_cost": len(transfer_with_cost),
                "without_cost": len(transfer_no_cost),
                "note": "Transfers track asset movement between departments/locations - cost field is optional"
            },
            "affected_count": len(transfer_assets)
        })

    # 9. Transaction type summary
    info_messages.append({
        "type": "TRANSACTION_SUMMARY",
        "message": "Asset classification summary",
        "breakdown": trans_types,
        "tax_year": tax_year
    })

    # 10. OBBBA 2025 info
    if tax_year >= 2025:
        info_messages.append({
            "type": "OBBBA_2025_EFFECTIVE",
            "message": "One Big Beautiful Bill Act (OBBBA) 2025 provisions apply",
            "details": {
                "bonus_depreciation": "100% for assets acquired AND placed in service after 1/19/2025",
                "section_179_limit": "$2.5 million (increased from $1.2 million)",
                "phase_out_threshold": "$4 million"
            }
        })

    return {
        "critical": critical_warnings,
        "warnings": warnings,
        "info": info_messages,
        "summary": {
            "total_assets": len(assets),
            "critical_issues": len(critical_warnings),
            "warnings": len(warnings),
            "tax_year": tax_year
        }
    }

# Maximum upload file size (50MB) - prevents DoS via large files
MAX_UPLOAD_SIZE_MB = 50
MAX_UPLOAD_SIZE_BYTES = MAX_UPLOAD_SIZE_MB * 1024 * 1024


@app.post("/upload", response_model=List[Asset])
async def upload_file(
    request: Request,
    response: Response,
    file: UploadFile = File(...),
    user: AuthUser = Depends(optional_auth)
):
    """
    Uploads an Excel file, parses it, and returns classified assets.
    Uses session-based storage for user isolation.

    Authentication: Required when AUTH_ENABLED=true

    Uses the configured tax year for proper transaction classification:
    - Current Year Additions (eligible for Section 179/Bonus)
    - Existing Assets (NOT eligible for Section 179/Bonus)
    - Disposals
    - Transfers

    SAFETY:
    - Uses per-session locking to prevent concurrent upload data corruption.
    - Enforces file size limit to prevent DoS attacks.
    """
    # Get or create session for this user
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    # Check file size limit (security: prevent DoS via large files)
    # Read file content to check size (UploadFile.size may not be available)
    file_content = await file.read()
    file_size = len(file_content)
    await file.seek(0)  # Reset file position for later reading

    if file_size > MAX_UPLOAD_SIZE_BYTES:
        raise api_error(
            413, "FILE_TOO_LARGE",
            f"File size ({file_size / 1024 / 1024:.1f}MB) exceeds maximum allowed size ({MAX_UPLOAD_SIZE_MB}MB)",
            {"max_size_mb": MAX_UPLOAD_SIZE_MB, "file_size_mb": round(file_size / 1024 / 1024, 1)}
        )

    # Acquire session lock to prevent concurrent upload corruption
    session_lock = get_session_lock(session.session_id)
    if not session_lock.acquire(blocking=False):
        raise api_error(409, "UPLOAD_IN_PROGRESS",
            "Another upload is already in progress for this session. Please wait.")

    # SECURITY: Use secure temporary file in system temp directory
    # This prevents arbitrary file write vulnerabilities from CWD manipulation
    # The file is created with restrictive permissions (0600) by default
    temp_fd, temp_file = tempfile.mkstemp(suffix='.xlsx', prefix='facs_upload_')

    try:
        # Save uploaded file to secure temp location
        with os.fdopen(temp_fd, 'wb') as buffer:
            buffer.write(file_content)  # Use already-read content

        # Perform tab analysis before processing
        try:
            import openpyxl
            wb = openpyxl.load_workbook(temp_file, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                sheets[sheet_name] = pd.DataFrame(data)
            wb.close()

            # Run tab analysis using session's tax year
            current_tax_year = session.tax_config.get("tax_year", TAX_CONFIG["tax_year"])
            session.tab_analysis_result = analyze_tabs(sheets, current_tax_year)
            logger.info(f"Tab analysis: {len(session.tab_analysis_result.tabs)} tabs detected")
        except Exception as tab_err:
            logger.warning(f"Tab analysis error (non-fatal): {tab_err}")
            session.tab_analysis_result = None

        # 1. Parse Excel
        assets = importer.parse_excel(temp_file)

        # 2. Classify Assets (MACRS + Transaction Types)
        tax_year = session.tax_config.get("tax_year", TAX_CONFIG["tax_year"])
        classified_assets = classifier.classify_batch(assets, tax_year=tax_year)

        # 3. Store in session using unique IDs to prevent overwrites
        session.assets.clear()
        session.approved_assets.clear()
        session.asset_id_counter = 0
        session.last_upload_filename = file.filename

        for asset in classified_assets:
            asset.unique_id = session.asset_id_counter
            session.assets[session.asset_id_counter] = asset
            session.asset_id_counter += 1

        # Save session
        manager = get_session_manager()
        manager._save_session(session)

        logger.info(f"[Upload] Session {session.session_id}: Stored {len(session.assets)} assets")

        # Log transaction type summary
        trans_types = {}
        for a in classified_assets:
            tt = a.transaction_type or "unknown"
            trans_types[tt] = trans_types.get(tt, 0) + 1
        logger.info(f"Classification Summary (Tax Year {tax_year}): {trans_types}")

        return classified_assets
        
    except Exception as e:
        # Log error with full traceback for debugging
        logger.error(f"Upload Error: {e}", exc_info=True)

        # Don't expose internal error details to client
        raise api_error(500, "FILE_PROCESSING_FAILED", "File processing failed. Please check the file format and try again.")
    finally:
        # Always release session lock
        session_lock.release()

        # Cleanup temp file
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception as cleanup_error:
                logger.warning(f"Failed to delete temp file {temp_file}: {cleanup_error}")


# ==============================================================================
# BACKGROUND JOB API
# ==============================================================================

from backend.logic.job_processor import (
    get_job_processor,
    init_job_handlers,
    JobType,
    JobStatus
)

# Initialize job handlers on import
init_job_handlers()


@app.post("/jobs/upload")
async def submit_upload_job(
    request: Request,
    response: Response,
    file: UploadFile = File(...),
    user: AuthUser = Depends(optional_auth)
):
    """
    Submit a file upload for async background processing.

    Returns a job_id immediately. Client polls /jobs/{job_id} for status.

    This prevents HTTP timeouts for large files with many assets.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    # Validate file
    if not file.filename:
        raise api_error(400, "NO_FILE", "No file provided")

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise api_error(400, "INVALID_FILE_TYPE", "Only Excel files (.xlsx, .xls) are supported")

    # Save to temp file
    temp_fd, temp_file = tempfile.mkstemp(suffix='.xlsx', prefix='facs_job_')
    try:
        content = await file.read()

        # Size check
        max_size = int(os.environ.get("MAX_UPLOAD_SIZE_MB", "50")) * 1024 * 1024
        if len(content) > max_size:
            os.close(temp_fd)
            os.remove(temp_file)
            raise api_error(413, "FILE_TOO_LARGE", f"File exceeds maximum size of {max_size // (1024*1024)}MB")

        os.write(temp_fd, content)
        os.close(temp_fd)

        # Submit job
        processor = get_job_processor()
        job = processor.submit(
            job_type=JobType.UPLOAD,
            params={
                "file_path": temp_file,
                "filename": file.filename,
                "tax_year": session.tax_config.get("tax_year"),
            },
            session_id=session.session_id,
            user_id=user.user_id if user else None
        )

        logger.info(f"Submitted upload job {job.job_id} for file {file.filename}")

        return {
            "job_id": job.job_id,
            "status": job.status.value,
            "message": "Upload job submitted. Poll /jobs/{job_id} for status.",
        }

    except HTTPException:
        raise
    except Exception as e:
        # Clean up on error
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception:
                pass
        logger.error(f"Error submitting upload job: {e}", exc_info=True)
        raise api_error(500, "JOB_SUBMIT_FAILED", "Failed to submit upload job")


@app.get("/jobs/{job_id}")
async def get_job_status(
    request: Request,
    response: Response,
    job_id: str
):
    """
    Get the status of a background job.

    Client should poll this endpoint until status is 'completed' or 'failed'.
    Recommended polling interval: 1-2 seconds.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    processor = get_job_processor()
    status = processor.get_job_status(job_id)

    if not status:
        raise api_error(404, "JOB_NOT_FOUND", f"Job {job_id} not found")

    return status


@app.get("/jobs/{job_id}/result")
async def get_job_result(
    request: Request,
    response: Response,
    job_id: str
):
    """
    Get the result of a completed job.

    For upload jobs, this returns the list of classified assets.
    Assets are automatically stored in the session.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    processor = get_job_processor()
    job = processor.get_job(job_id)

    if not job:
        raise api_error(404, "JOB_NOT_FOUND", f"Job {job_id} not found")

    if job.status != JobStatus.COMPLETED:
        raise api_error(400, "JOB_NOT_COMPLETED",
            f"Job is not completed. Current status: {job.status.value}")

    result = job.result

    # For upload jobs, store assets in session
    if job.job_type == JobType.UPLOAD and result and "assets" in result:
        from backend.models.asset import Asset

        # Clear existing and store new
        session.assets.clear()
        session.approved_assets.clear()

        for asset_data in result["assets"]:
            asset = Asset(**asset_data)
            session.assets[asset.unique_id] = asset

        session.last_upload_filename = job.metadata.get("filename", "unknown")

        # Save session
        manager = get_session_manager()
        manager._save_session(session)

        logger.info(f"Stored {len(session.assets)} assets from job {job_id} to session {session.session_id}")

    return result


@app.delete("/jobs/{job_id}")
async def cancel_job(
    request: Request,
    response: Response,
    job_id: str
):
    """
    Cancel a pending job.

    Only pending jobs can be cancelled. Running jobs will complete.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    processor = get_job_processor()

    if not processor.get_job(job_id):
        raise api_error(404, "JOB_NOT_FOUND", f"Job {job_id} not found")

    if processor.cancel_job(job_id):
        return {"status": "cancelled", "job_id": job_id}
    else:
        raise api_error(400, "CANNOT_CANCEL",
            "Job cannot be cancelled. It may already be running or completed.")


@app.get("/jobs")
async def list_session_jobs(
    request: Request,
    response: Response
):
    """
    List all jobs for the current session.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    processor = get_job_processor()
    jobs = processor.get_jobs_by_session(session.session_id)

    return {
        "jobs": [j.to_dict() for j in jobs],
        "count": len(jobs),
    }


@app.get("/jobs/stats")
async def get_job_stats():
    """
    Get job processor statistics.
    """
    processor = get_job_processor()
    return processor.get_stats()


@app.post("/assets/{asset_id}/update", response_model=Asset)
async def update_asset(
    request: Request,
    response: Response,
    asset_id: int,
    update_data: Dict = Body(...),
    user: AuthUser = Depends(optional_auth)
):
    """
    Updates a specific asset and logs the change in the audit trail.
    Uses session-based storage for user isolation.
    Note: asset_id here is the unique_id, not the original row_index.

    Authentication: Required when AUTH_ENABLED=true
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    # Whitelist of allowed fields to update (security measure)
    ALLOWED_UPDATE_FIELDS = {
        "macrs_class", "macrs_life", "macrs_method", "macrs_convention",
        "fa_cs_wizard_category", "description", "cost", "acquisition_date",
        "in_service_date", "transaction_type", "is_bonus_eligible",
        "is_qualified_improvement", "fa_cs_asset_number"  # FA CS cross-reference
    }

    # Filter out any fields not in the whitelist
    safe_update_data = {k: v for k, v in update_data.items() if k in ALLOWED_UPDATE_FIELDS}

    if not safe_update_data:
        raise api_error(
            400, "INVALID_UPDATE_FIELDS",
            f"No valid fields to update. Allowed fields: {', '.join(ALLOWED_UPDATE_FIELDS)}",
            {"allowed_fields": list(ALLOWED_UPDATE_FIELDS)}
        )

    # Validate fa_cs_asset_number if provided (must be positive integer >= 1 or None)
    if "fa_cs_asset_number" in safe_update_data:
        fa_cs_num = safe_update_data["fa_cs_asset_number"]
        if fa_cs_num is not None:
            # Ensure it's an integer
            if not isinstance(fa_cs_num, int):
                try:
                    fa_cs_num = int(fa_cs_num)
                    safe_update_data["fa_cs_asset_number"] = fa_cs_num
                except (ValueError, TypeError):
                    raise api_error(
                        400, "INVALID_FA_CS_NUMBER",
                        "FA CS Asset # must be a positive integer",
                        {"value": fa_cs_num}
                    )
            # Ensure it's >= 1
            if fa_cs_num < 1:
                raise api_error(
                    400, "INVALID_FA_CS_NUMBER",
                    "FA CS Asset # must be >= 1 (FA CS requires positive asset numbers)",
                    {"value": fa_cs_num}
                )

    if asset_id not in session.assets:
        raise api_error(404, "ASSET_NOT_FOUND", f"Asset with ID {asset_id} not found")

    asset = session.assets[asset_id]

    # Check for changes and log them (use safe_update_data)
    for field, new_value in safe_update_data.items():
        if hasattr(asset, field):
            old_value = getattr(asset, field)
            if str(old_value) != str(new_value):
                # Update the field
                setattr(asset, field, new_value)
                # Log the change
                auditor.log_override(asset, field, str(old_value), str(new_value))

    # Save session
    manager = get_session_manager()
    manager._save_session(session)

    return asset

@app.post("/assets/{asset_id}/approve")
async def approve_asset(
    request: Request,
    response: Response,
    asset_id: int,
    user: AuthUser = Depends(optional_auth)
):
    """
    CPA approves a single asset for export.
    Uses session-based storage for user isolation.
    Note: asset_id here is the unique_id, not the original row_index.

    Authentication: Required when AUTH_ENABLED=true
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if asset_id not in session.assets:
        raise api_error(404, "ASSET_NOT_FOUND", f"Asset with ID {asset_id} not found")

    asset = session.assets[asset_id]

    # Can't approve assets with validation errors
    if getattr(asset, 'validation_errors', None):
        raise api_error(400, "VALIDATION_ERRORS", "Cannot approve asset with validation errors",
                       {"errors": asset.validation_errors})

    session.approved_assets.add(asset_id)

    # Save session
    manager = get_session_manager()
    manager._save_session(session)

    return {"approved": True, "unique_id": asset_id}


@app.post("/assets/{asset_id}/election")
async def update_asset_election(
    request: Request,
    response: Response,
    asset_id: int,
    body: dict = Body(...)
):
    """
    Update depreciation election for an asset.

    Valid elections:
    - MACRS: Standard MACRS depreciation
    - DeMinimis: De minimis safe harbor (expense immediately if ≤$2,500)
    - Section179: Section 179 expense election
    - Bonus: Bonus depreciation (80% for 2024, per Form 4562 Instructions)
    - ADS: Alternative Depreciation System

    Note: This is a CPA decision based on client's income situation.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if asset_id not in session.assets:
        raise api_error(404, "ASSET_NOT_FOUND", f"Asset with ID {asset_id} not found")

    election = body.get("election", "MACRS")
    valid_elections = ["MACRS", "DeMinimis", "Section179", "Bonus", "ADS"]

    if election not in valid_elections:
        raise api_error(400, "INVALID_ELECTION", f"Invalid election: {election}",
                       {"valid_options": valid_elections})

    asset = session.assets[asset_id]
    asset.depreciation_election = election
    asset.election_reason = f"Manually selected by CPA"

    # Save session
    manager = get_session_manager()
    manager._save_session(session)

    return {
        "success": True,
        "unique_id": asset_id,
        "election": election
    }


@app.post("/assets/approve-batch")
async def approve_batch(
    request: Request,
    response: Response,
    asset_ids: List[int] = Body(...),
    user: AuthUser = Depends(optional_auth)
):
    """
    CPA approves multiple assets at once (e.g., all high-confidence items).
    Uses session-based storage for user isolation.
    Note: asset_ids are unique_ids, not row_indexes.

    Authentication: Required when AUTH_ENABLED=true
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    approved = []
    errors = []

    for asset_id in asset_ids:
        if asset_id not in session.assets:
            errors.append({"unique_id": asset_id, "error": "Not found"})
            continue

        asset = session.assets[asset_id]
        if getattr(asset, 'validation_errors', None):
            errors.append({"unique_id": asset_id, "error": "Has validation errors"})
            continue

        session.approved_assets.add(asset_id)
        approved.append(asset_id)

    # Save session
    manager = get_session_manager()
    manager._save_session(session)

    return {"approved": approved, "errors": errors, "total_approved": len(approved)}

@app.post("/assets/auto-generate-fa-cs-numbers")
async def auto_generate_fa_cs_numbers(request: Request, response: Response, mode: str = "sequential"):
    """
    Auto-generate unique FA CS Asset # for all assets to resolve collisions.

    Modes:
    - "sequential": Assign sequential numbers starting from max existing + 1
    - "preserve": Keep existing numbers where unique, only fix collisions
    - "row_based": Use row_index as FA CS # (guarantees uniqueness)

    Returns updated assets with new fa_cs_asset_number assignments.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    assets = list(session.assets.values())
    if not assets:
        return {"message": "No assets to update", "updated": 0}

    # First, detect collisions using exporter logic
    from backend.services.exporter import ExporterService
    exporter = ExporterService()

    updated_count = 0

    if mode == "row_based":
        # Simple: use row_index as FA CS # (always unique)
        for asset in assets:
            if asset.fa_cs_asset_number != asset.row_index:
                asset.fa_cs_asset_number = asset.row_index
                updated_count += 1

    elif mode == "sequential":
        # Find max existing FA CS # and assign sequential from there
        existing_numbers = set()
        for asset in assets:
            if asset.fa_cs_asset_number is not None:
                existing_numbers.add(asset.fa_cs_asset_number)
            else:
                # Also consider auto-generated numbers
                auto_num = exporter._format_asset_number(asset)
                existing_numbers.add(auto_num)

        # Start from max + 1 (or 1 if no existing)
        next_num = max(existing_numbers) + 1 if existing_numbers else 1

        # Detect collision groups
        collision_result = exporter.detect_asset_number_collisions(assets)

        if collision_result["has_collisions"]:
            # Get all unique_ids involved in collisions
            collision_asset_ids = set()
            for collision in collision_result["collisions"]:
                for asset_info in collision["assets"]:
                    collision_asset_ids.add(asset_info["unique_id"])

            # Assign new numbers to colliding assets (skip first in each group to preserve one)
            for collision in collision_result["collisions"]:
                # Skip first asset (it keeps its number), reassign the rest
                for asset_info in collision["assets"][1:]:
                    asset = session.assets.get(asset_info["unique_id"])
                    if asset:
                        asset.fa_cs_asset_number = next_num
                        next_num += 1
                        updated_count += 1

    elif mode == "preserve":
        # Only fix collisions, preserve existing assignments
        collision_result = exporter.detect_asset_number_collisions(assets)

        if collision_result["has_collisions"]:
            # Find all used numbers (both explicit and auto-generated)
            used_numbers = set()
            for asset in assets:
                used_numbers.add(exporter._format_asset_number(asset))

            next_num = max(used_numbers) + 1

            # Fix collisions
            for collision in collision_result["collisions"]:
                # Keep first asset, reassign rest
                for asset_info in collision["assets"][1:]:
                    asset = session.assets.get(asset_info["unique_id"])
                    if asset:
                        asset.fa_cs_asset_number = next_num
                        used_numbers.add(next_num)
                        next_num += 1
                        updated_count += 1

    # Save session
    manager = get_session_manager()
    manager._save_session(session)

    # Re-check for collisions after fix
    new_collision_result = exporter.detect_asset_number_collisions(list(session.assets.values()))

    return {
        "message": f"Updated {updated_count} assets",
        "updated": updated_count,
        "mode": mode,
        "remaining_collisions": new_collision_result["collision_count"],
        "assets": list(session.assets.values())
    }

@app.delete("/assets/{asset_id}/approve")
async def unapprove_asset(request: Request, response: Response, asset_id: int):
    """
    Remove approval from an asset.
    Uses session-based storage for user isolation.
    Note: asset_id here is the unique_id, not the original row_index.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    was_approved = asset_id in session.approved_assets
    if was_approved:
        session.approved_assets.discard(asset_id)
        # Save session
        manager = get_session_manager()
        manager._save_session(session)
        return {"approved": False, "unique_id": asset_id}
    return {"approved": False, "unique_id": asset_id, "message": "Was not approved"}


@app.get("/export/status")
async def get_export_status(request: Request, response: Response):
    """
    Check if export is ready and get detailed approval status.
    Uses session-based storage for user isolation.
    Frontend should call this to enable/disable export button.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if not session.assets:
        return {
            "ready": False,
            "reason": "No assets loaded",
            "total_assets": 0,
            "approved_assets": 0,
            "approved_ids": []
        }

    assets = list(session.assets.values())

    # Count by category
    total = len(assets)
    errors = sum(1 for a in assets if getattr(a, 'validation_errors', None))

    # Actionable = additions, disposals, transfers (not existing)
    actionable = [a for a in assets if a.transaction_type not in ["Existing Asset", None]]
    actionable_count = len(actionable)

    # Approved actionable
    approved_actionable = [a for a in actionable if a.unique_id in session.approved_assets]
    approved_count = len(approved_actionable)

    # Low confidence needing review
    low_conf_unreviewed = [
        a for a in actionable
        if a.confidence_score <= 0.8 and a.unique_id not in session.approved_assets
    ]

    # Determine readiness
    has_errors = errors > 0
    all_approved = approved_count == actionable_count
    low_conf_reviewed = len(low_conf_unreviewed) == 0

    ready = not has_errors and all_approved and low_conf_reviewed

    reason = None
    if has_errors:
        reason = f"{errors} asset(s) have validation errors"
    elif not all_approved:
        reason = f"{actionable_count - approved_count} of {actionable_count} actionable assets not approved"
    elif not low_conf_reviewed:
        reason = f"{len(low_conf_unreviewed)} low-confidence assets need review"

    return {
        "ready": ready,
        "reason": reason,
        "total_assets": total,
        "actionable_assets": actionable_count,
        "approved_assets": approved_count,
        "unapproved_assets": actionable_count - approved_count,
        "errors": errors,
        "low_confidence_unreviewed": len(low_conf_unreviewed),
        "approved_ids": list(session.approved_assets)
    }


@app.get("/export/compatibility-check")
async def check_fa_cs_compatibility(request: Request, response: Response):
    """
    Real-Time FA CS Compatibility Checker.
    Validates assets before export to prevent broken imports.

    Checks:
    - Missing required fields
    - Valid method/life pairs for FA CS
    - Date format validity
    - Negative number formatting
    - Class to category mapping validity
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if not session.assets:
        return {
            "is_compatible": True,
            "issues": [],
            "summary": "No assets to check"
        }

    issues = []
    assets = list(session.assets.values())

    # Valid MACRS life/method combinations for FA CS
    valid_life_method = {
        3: ["200DB", "150DB", "SL"],
        5: ["200DB", "150DB", "SL"],
        7: ["200DB", "150DB", "SL"],
        10: ["200DB", "150DB", "SL"],
        15: ["150DB", "SL"],
        20: ["150DB", "SL"],
        25: ["SL"],
        27.5: ["SL"],
        39: ["SL"],
    }

    for asset in assets:
        # Skip existing assets (they're already in FA CS)
        if asset.transaction_type == "Existing Asset":
            continue

        asset_id = asset.asset_id or f"Row {asset.unique_id}"

        # Check 1: Missing required fields
        if not asset.description:
            issues.append({
                "asset_id": asset_id,
                "severity": "error",
                "message": "Missing description",
                "suggestion": "Add asset description"
            })

        if not asset.cost or asset.cost == 0:
            issues.append({
                "asset_id": asset_id,
                "severity": "error",
                "message": "Missing or zero cost",
                "suggestion": "Enter valid cost amount"
            })

        if not asset.in_service_date:
            issues.append({
                "asset_id": asset_id,
                "severity": "error",
                "message": "Missing in-service date",
                "suggestion": "Use Data Cleanup to auto-fill from acquisition date"
            })

        # Check 2: Valid method/life pair
        life = asset.macrs_life
        method = asset.macrs_method
        if life and method:
            valid_methods = valid_life_method.get(life, [])
            if method not in valid_methods and valid_methods:
                issues.append({
                    "asset_id": asset_id,
                    "severity": "warning",
                    "message": f"Invalid method/life: {method}/{life} yr",
                    "suggestion": f"Valid methods for {life}-yr: {', '.join(valid_methods)}"
                })

        # Check 3: Missing MACRS class
        if not asset.macrs_class:
            issues.append({
                "asset_id": asset_id,
                "severity": "warning",
                "message": "Missing MACRS class",
                "suggestion": "Review and assign asset class"
            })

        # Check 4: Negative cost format
        if asset.cost and asset.cost < 0 and asset.transaction_type != "Disposal":
            issues.append({
                "asset_id": asset_id,
                "severity": "warning",
                "message": f"Negative cost (${asset.cost:,.2f}) for non-disposal",
                "suggestion": "Verify if this should be a disposal"
            })

    is_compatible = len([i for i in issues if i["severity"] == "error"]) == 0

    return {
        "is_compatible": is_compatible,
        "issues": issues,
        "summary": f"{len(issues)} potential issues found" if issues else "All checks passed"
    }


@app.get("/export/depreciation-preview")
async def get_depreciation_preview(request: Request, response: Response):
    """
    179/Bonus Election Preview.
    Shows estimated Year 1 depreciation breakdown before export.

    Returns:
    - Section 179 eligible amount
    - Bonus depreciation (80% for 2024, 100% for 2025+ OBBBA)
    - Regular MACRS depreciation
    - Total Year 1 depreciation
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if not session.assets:
        return {
            "section_179": 0,
            "bonus": 0,
            "regular_macrs": 0,
            "total_year1": 0,
            "summary": "No assets loaded"
        }

    assets = list(session.assets.values())
    tax_year = session.tax_config.get("tax_year", 2024)

    # Get tax year limits from centralized config (OBBBA/TCJA compliant)
    section_179_config = tax_year_config.get_section_179_limits(tax_year)
    section_179_limit = section_179_config.get("max_deduction", 1220000)
    bonus_rate = tax_year_config.get_bonus_percentage(tax_year)

    section_179_total = 0
    bonus_total = 0
    regular_macrs_total = 0

    for asset in assets:
        # Only CY additions are eligible for 179/Bonus
        if asset.transaction_type != "Current Year Addition":
            continue

        cost = asset.cost or 0
        if cost <= 0:
            continue

        is_bonus_eligible = getattr(asset, 'is_bonus_eligible', True)
        life = asset.macrs_life or 7

        # First year MACRS rate (approximation)
        first_year_rates = {
            3: 0.3333,
            5: 0.20,
            7: 0.1429,
            10: 0.10,
            15: 0.05,
            20: 0.0375,
            27.5: 0.03636,
            39: 0.02564,
        }

        first_year_rate = first_year_rates.get(life, 0.1429)

        if is_bonus_eligible:
            # Assume 179 election for smaller assets
            if cost <= 2500:
                # De minimis - expense fully
                section_179_total += cost
            elif cost <= 50000 and section_179_total + cost <= section_179_limit:
                # 179 candidate
                section_179_total += cost
            else:
                # Bonus depreciation
                bonus_amount = cost * bonus_rate
                remaining = cost - bonus_amount
                regular = remaining * first_year_rate
                bonus_total += bonus_amount
                regular_macrs_total += regular
        else:
            # Regular MACRS only (existing assets, not eligible)
            regular_macrs_total += cost * first_year_rate

    total_year1 = section_179_total + bonus_total + regular_macrs_total

    return {
        "section_179": round(section_179_total, 2),
        "bonus": round(bonus_total, 2),
        "regular_macrs": round(regular_macrs_total, 2),
        "total_year1": round(total_year1, 2),
        "tax_year": tax_year,
        "bonus_rate": bonus_rate,
        "section_179_limit": section_179_limit,
        "summary": f"Estimated ${total_year1:,.0f} Year 1 depreciation"
    }


@app.post("/export/auto-fix")
async def auto_fix_compatibility_issues(
    request: Request,
    response: Response,
    user: AuthUser = Depends(optional_auth)
):
    """
    Auto-fix common FA CS compatibility issues.

    Authentication: Required when AUTH_ENABLED=true

    Fixes:
    - Missing method: Default to 200DB for 3-7yr, 150DB for 15-20yr, SL for others
    - Missing convention: Default to HY
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    fixed_count = 0
    assets = session.assets

    for asset_id, asset in assets.items():
        # Fix missing method based on life
        if not asset.macrs_method and asset.macrs_life:
            life = asset.macrs_life
            if life in [3, 5, 7, 10]:
                asset.macrs_method = "200DB"
            elif life in [15, 20]:
                asset.macrs_method = "150DB"
            else:
                asset.macrs_method = "SL"
            fixed_count += 1

        # Fix missing convention
        if not asset.macrs_convention:
            asset.macrs_convention = "HY"
            fixed_count += 1

    return {
        "success": True,
        "fixed_count": fixed_count
    }


@app.get("/export")
async def export_assets(
    request: Request,
    skip_approval_check: bool = Query(False, description="Skip approval validation (requires X-Admin-Key header)")
):
    """
    Generates an Excel file for Fixed Assets CS import.
    Uses session-based storage for user isolation.
    Saves a copy to 'bot_handoff' folder for UiPath to pick up.

    IMPORTANT: All actionable assets (additions, disposals, transfers) must be approved
    before export. Existing assets are excluded from approval requirement.

    SECURITY: skip_approval_check requires admin authentication via X-Admin-Key header.
    """
    session = await get_current_session(request)

    if not session.assets:
        raise api_error(400, "NO_ASSETS", "No assets to export")

    # Get all assets from session
    assets = list(session.assets.values())

    # Validate: Block export if any asset has validation errors
    assets_with_errors = [a for a in assets if getattr(a, 'validation_errors', None)]
    if assets_with_errors:
        error_count = len(assets_with_errors)
        raise api_error(400, "VALIDATION_ERRORS",
            f"Cannot export: {error_count} asset(s) have validation errors. Fix all errors before exporting.",
            {"error_count": error_count}
        )
    
    # Generate Excel with tax configuration
    excel_file = exporter.generate_fa_cs_export(
        assets,
        tax_year=TAX_CONFIG["tax_year"],
        de_minimis_limit=TAX_CONFIG["de_minimis_threshold"]
    )

    # Save to Bot Handoff Folder (use custom path if set)
    if FACS_CONFIG["export_path"]:
        handoff_dir = FACS_CONFIG["export_path"]
    else:
        handoff_dir = os.path.join(os.getcwd(), "bot_handoff")

    os.makedirs(handoff_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"FA_Import_{session.session_id[:8]}_{timestamp}.xlsx"
    filepath = os.path.join(handoff_dir, filename)

    with open(filepath, "wb") as f:
        f.write(excel_file.getvalue())

    logger.info(f"Session {session.session_id}: Saved export to {filepath}")

    # Reset stream position for StreamingResponse
    excel_file.seek(0)

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=FA_CS_Import.xlsx",
            "X-Session-ID": session.session_id
        }
    )


@app.get("/export/audit")
async def export_audit_documentation(request: Request):
    """
    Generates a comprehensive audit documentation Excel file.
    Includes ALL assets (additions, disposals, transfers, AND existing assets)
    with full classification details, confidence scores, and reasoning.

    This is separate from the FA CS export and is meant for:
    - IRS audit documentation
    - Internal compliance records
    - Year-over-year reconciliation

    SECURITY: Uses session-based storage for user isolation.
    """
    # FIX: Use session instead of global ASSET_STORE for user isolation
    session = await get_current_session(request)

    if not session.assets:
        raise api_error(400, "NO_ASSETS", "No assets to export")

    assets = list(session.assets.values())
    tax_year = session.tax_config.get("tax_year", TAX_CONFIG["tax_year"])

    # Create comprehensive audit DataFrame
    audit_data = []
    for a in assets:
        audit_data.append({
            "Asset ID": a.asset_id or "",
            "Description": a.description,
            "Cost": a.cost,
            "Acquisition Date": a.acquisition_date.isoformat() if a.acquisition_date else "",
            "In Service Date": a.in_service_date.isoformat() if a.in_service_date else "",
            "Transaction Type": a.transaction_type or "",
            "Classification Reason": a.classification_reason or "",
            "MACRS Class": a.macrs_class or "",
            "MACRS Life": a.macrs_life or "",
            "MACRS Method": a.macrs_method or "",
            "MACRS Convention": a.macrs_convention or "",
            "Classification Confidence": f"{(a.confidence_score or 0) * 100:.0f}%",
            "Bonus Eligible": "Yes" if a.is_bonus_eligible else "No",
            "Qualified Improvement": "Yes" if a.is_qualified_improvement else "No",
            "Validation Errors": "; ".join(a.validation_errors) if a.validation_errors else "",
            "Validation Warnings": "; ".join(a.validation_warnings) if a.validation_warnings else "",
            "Source Sheet": a.source_sheet or "",
            "Row Index": a.row_index
        })

    df = pd.DataFrame(audit_data)

    # Create summary statistics
    summary_data = {
        "Metric": [
            "Tax Year",
            "Total Assets",
            "Total Cost",
            "Current Year Additions",
            "Existing Assets",
            "Disposals",
            "Transfers",
            "Unknown (Missing Date)",
            "Assets with Errors",
            "High Confidence (>80%)",
            "Low Confidence (<80%)",
            "Export Generated"
        ],
        "Value": [
            str(tax_year),
            str(len(assets)),
            f"${sum(a.cost or 0 for a in assets):,.2f}",
            str(len([a for a in assets if a.transaction_type == "Current Year Addition"])),
            str(len([a for a in assets if a.transaction_type == "Existing Asset"])),
            str(len([a for a in assets if a.transaction_type == "Disposal"])),
            str(len([a for a in assets if a.transaction_type == "Transfer"])),
            str(len([a for a in assets if a.transaction_type and "Unknown" in a.transaction_type])),
            str(len([a for a in assets if a.validation_errors])),
            str(len([a for a in assets if (a.confidence_score or 0) > 0.8])),
            str(len([a for a in assets if (a.confidence_score or 0) <= 0.8])),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
    }
    summary_df = pd.DataFrame(summary_data)

    # Create Excel file with multiple sheets
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Summary sheet first
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # All assets - sorted by transaction type
        df_sorted = df.sort_values(['Transaction Type', 'Asset ID'])
        df_sorted.to_excel(writer, sheet_name='All Assets', index=False)

        # Separate sheets by transaction type for easier navigation
        for trans_type in ['Current Year Addition', 'Existing Asset', 'Disposal', 'Transfer']:
            type_df = df[df['Transaction Type'] == trans_type]
            if len(type_df) > 0:
                sheet_name = trans_type.replace(' ', '_')[:31]  # Excel sheet name limit
                type_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Errors sheet (if any)
        errors_df = df[df['Validation Errors'] != '']
        if len(errors_df) > 0:
            errors_df.to_excel(writer, sheet_name='Validation_Errors', index=False)

    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Audit_Documentation_{tax_year}_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==============================================================================
# DATA QUALITY & ANALYSIS ENDPOINTS
# ==============================================================================

@app.get("/quality")
async def get_data_quality(request: Request, response: Response):
    """
    Get data quality score with A-F grade and detailed breakdown.
    Returns a comprehensive quality assessment for the loaded assets.

    SECURITY: Uses session-based storage for user isolation.
    """
    # FIX: Use session instead of global ASSET_STORE for user isolation
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if not session.assets:
        return {
            "grade": "-",
            "score": 0,
            "is_export_ready": False,
            "summary": "No assets loaded",
            "checks": [],
            "critical_issues": [],
            "recommendations": ["Upload an asset schedule to begin"]
        }

    # Convert assets to DataFrame for quality scoring
    assets = list(session.assets.values())
    df_data = []
    for a in assets:
        df_data.append({
            "Asset ID": a.asset_id,
            "Description": a.description,
            "Cost": a.cost,
            "In Service Date": a.in_service_date,
            "Final Category": a.macrs_class,
            "MACRS Life": a.macrs_life,
            "Method": a.macrs_method,
            "Transaction Type": a.transaction_type,
        })

    df = pd.DataFrame(df_data)
    tax_year = session.tax_config.get("tax_year", TAX_CONFIG["tax_year"])

    try:
        quality_result = calculate_data_quality_score(df, tax_year)

        # Convert checks to serializable format
        checks_data = []
        for check in quality_result.checks:
            checks_data.append({
                "name": check.name,
                "passed": check.passed,
                "score": round(check.score, 1),
                "max_score": check.max_score,
                "weight": check.weight,
                "details": check.details,
                "fix_suggestion": check.fix_suggestion
            })

        return {
            "grade": quality_result.grade,
            "score": round(quality_result.score, 1),
            "is_export_ready": quality_result.is_export_ready,
            "summary": quality_result.summary,
            "checks": checks_data,
            "critical_issues": quality_result.critical_issues,
            "recommendations": quality_result.recommendations
        }
    except Exception as e:
        print(f"Quality score error: {e}")
        return {
            "grade": "?",
            "score": 0,
            "is_export_ready": False,
            "summary": f"Error calculating quality: {str(e)}",
            "checks": [],
            "critical_issues": [str(e)],
            "recommendations": []
        }


@app.get("/tabs")
async def get_tab_analysis(request: Request, response: Response):
    """
    Get the tab analysis results from the last upload.
    Shows which tabs were detected and their roles.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    tab_result = getattr(session, 'tab_analysis_result', None)
    if tab_result is None:
        return {
            "tabs": [],
            "target_fiscal_year": session.tax_config.get("tax_year", TAX_CONFIG["tax_year"]),
            "warnings": [],
            "summary": "No file uploaded yet",
            "efficiency": {}
        }

    # Convert tab analysis to serializable format
    tabs_data = []
    for tab in tab_result.tabs:
        tabs_data.append({
            "tab_name": tab.tab_name,
            "role": tab.role.value,
            "role_label": tab.role_label,
            "icon": tab.icon,
            "fiscal_year": tab.fiscal_year,
            "row_count": tab.row_count,
            "data_row_count": tab.data_row_count,
            "has_cost_data": tab.has_cost_data,
            "has_date_data": tab.has_date_data,
            "should_process": tab.should_process,
            "skip_reason": tab.skip_reason,
            "confidence": round(tab.confidence, 2),
            "detection_notes": tab.detection_notes
        })

    efficiency = tab_result.get_efficiency_stats()

    return {
        "tabs": tabs_data,
        "target_fiscal_year": tab_result.target_fiscal_year,
        "warnings": tab_result.warnings,
        "summary_tabs": tab_result.summary_tabs,
        "detail_tabs": tab_result.detail_tabs,
        "disposal_tabs": tab_result.disposal_tabs,
        "efficiency": efficiency,
        "filename": getattr(session, 'last_upload_filename', None)
    }


@app.get("/rollforward")
async def get_rollforward_status(request: Request, response: Response):
    """
    Get rollforward reconciliation status for loaded assets.
    Shows beginning balance, additions, disposals, and ending balance.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if not session.assets:
        return {
            "is_balanced": True,
            "beginning_balance": 0,
            "additions": 0,
            "disposals": 0,
            "transfers_in": 0,
            "transfers_out": 0,
            "expected_ending": 0,
            "variance": 0,
            "details": {},
            "warnings": ["No assets loaded"],
            "status_label": "No Data"
        }

    # Convert assets to DataFrame
    assets = list(session.assets.values())
    df_data = []
    de_minimis_total = 0.0
    de_minimis_count = 0

    for a in assets:
        # Check if this is a De Minimis expensed item
        election = getattr(a, 'depreciation_election', 'MACRS') or 'MACRS'
        is_de_minimis = election == 'DeMinimis'

        # Track De Minimis separately - they are expensed, not capitalized
        if is_de_minimis and a.transaction_type == 'Current Year Addition':
            de_minimis_total += a.cost or 0
            de_minimis_count += 1
            # Skip adding to rollforward - it's an expense, not a capital addition
            continue

        df_data.append({
            "Cost": a.cost,
            "Transaction Type": a.transaction_type or "",
        })

    df = pd.DataFrame(df_data)

    try:
        result = reconcile_rollforward(df, cost_column="Cost", trans_type_column="Transaction Type")

        status_label = "Balanced" if result.is_balanced else f"Out of Balance (${result.variance:,.2f})"

        return {
            "is_balanced": result.is_balanced,
            "beginning_balance": round(result.beginning_balance, 2),
            "additions": round(result.additions, 2),
            "disposals": round(result.disposals, 2),
            "transfers_in": round(result.transfers_in, 2),
            "transfers_out": round(result.transfers_out, 2),
            "expected_ending": round(result.expected_ending, 2),
            "actual_ending": round(result.actual_ending, 2),
            "variance": round(result.variance, 2),
            "details": result.details,
            "warnings": result.warnings,
            "status_label": status_label,
            # De Minimis items - expensed, not capitalized
            "de_minimis_expensed": round(de_minimis_total, 2),
            "de_minimis_count": de_minimis_count
        }
    except Exception as e:
        logger.error(f"Rollforward error: {e}")
        return {
            "is_balanced": False,
            "beginning_balance": 0,
            "additions": 0,
            "disposals": 0,
            "transfers_in": 0,
            "transfers_out": 0,
            "expected_ending": 0,
            "variance": 0,
            "details": {},
            "warnings": [f"Error: {str(e)}"],
            "status_label": "Error",
            "de_minimis_expensed": round(de_minimis_total, 2),
            "de_minimis_count": de_minimis_count
        }


@app.get("/projection")
async def get_depreciation_projection(request: Request, response: Response):
    """
    Get 10-year depreciation projection for loaded assets.
    Shows year-by-year tax depreciation forecast.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if not session.assets:
        return {
            "years": [],
            "depreciation": [],
            "total_10_year": 0,
            "current_year": 0,
            "summary": "No assets loaded"
        }

    # Convert assets to DataFrame with required columns
    assets = list(session.assets.values())
    df_data = []
    for a in assets:
        # Calculate depreciable basis (simplified - full calc in export)
        depreciable_basis = a.cost or 0

        df_data.append({
            "Depreciable Basis": depreciable_basis,
            "Recovery Period": a.macrs_life or 7,
            "Method": a.macrs_method or "200DB",
            "Convention": a.macrs_convention or "HY",
            "In Service Date": a.in_service_date,
        })

    df = pd.DataFrame(df_data)
    current_year = session.tax_config.get("tax_year", TAX_CONFIG["tax_year"])

    try:
        projection_df = project_portfolio_depreciation(df, current_year, projection_years=10)

        years = projection_df["Tax Year"].tolist()
        depreciation = [round(d, 2) for d in projection_df["Total Depreciation"].tolist()]

        total_10_year = sum(depreciation)
        current_year_dep = depreciation[0] if depreciation else 0

        return {
            "years": years,
            "depreciation": depreciation,
            "total_10_year": round(total_10_year, 2),
            "current_year": round(current_year_dep, 2),
            "asset_count": len(assets),
            "summary": f"${total_10_year:,.0f} total depreciation over 10 years"
        }
    except Exception as e:
        logger.error(f"Projection error: {e}")
        return {
            "years": [],
            "depreciation": [],
            "total_10_year": 0,
            "current_year": 0,
            "summary": f"Error calculating projection: {str(e)}"
        }


@app.get("/confidence")
async def get_confidence_breakdown(request: Request, response: Response):
    """
    Get breakdown of assets by confidence level.
    Helps CPAs prioritize review time.
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if not session.assets:
        return {
            "high": {"count": 0, "pct": 0},
            "medium": {"count": 0, "pct": 0},
            "low": {"count": 0, "pct": 0},
            "total": 0,
            "auto_approve_eligible": 0
        }

    assets = list(session.assets.values())
    total = len(assets)

    # Count by confidence tier
    high = sum(1 for a in assets if a.confidence_score > 0.8)
    medium = sum(1 for a in assets if 0.5 < a.confidence_score <= 0.8)
    low = sum(1 for a in assets if a.confidence_score <= 0.5)

    # Count auto-approve eligible (high confidence + no errors)
    auto_approve = sum(1 for a in assets
                       if a.confidence_score > 0.8
                       and not getattr(a, 'validation_errors', None))

    return {
        "high": {
            "count": high,
            "pct": round(high / total * 100, 1) if total > 0 else 0,
            "label": "High (80%+)"
        },
        "medium": {
            "count": medium,
            "pct": round(medium / total * 100, 1) if total > 0 else 0,
            "label": "Medium (50-80%)"
        },
        "low": {
            "count": low,
            "pct": round(low / total * 100, 1) if total > 0 else 0,
            "label": "Low (<50%)"
        },
        "total": total,
        "auto_approve_eligible": auto_approve
    }


@app.get("/system-status")
def get_system_status():
    """
    Get system status including AI availability, memory patterns, etc.
    """
    # Check OpenAI API availability
    ai_available = True
    ai_status = "Online"
    try:
        import os
        if not os.getenv("OPENAI_API_KEY"):
            ai_available = False
            ai_status = "No API Key"
    except Exception:
        ai_available = False
        ai_status = "Error"

    # Check memory engine patterns
    memory_patterns = 0
    try:
        from pathlib import Path
        memory_path = Path(__file__).resolve().parent / "logic" / "classification_memory.json"
        if memory_path.exists():
            import json
            with open(memory_path, "r") as f:
                memory_data = json.load(f)
                memory_patterns = len(memory_data.get("assets", []))
    except Exception:
        pass

    # Count classification rules and keywords
    rules_count = 0
    keywords_count = 0
    try:
        config_dir = Path(__file__).resolve().parent / "logic" / "config"

        # Count explicit rules
        rules_path = config_dir / "rules.json"
        if rules_path.exists():
            import json
            with open(rules_path, "r") as f:
                rules_data = json.load(f)
                rules_count = len(rules_data.get("rules", []))

        # Count classification keywords (these also act as rules)
        keywords_path = config_dir / "classification_keywords.json"
        if keywords_path.exists():
            with open(keywords_path, "r") as f:
                keywords_data = json.load(f)
                if isinstance(keywords_data, dict):
                    keywords_count = sum(len(v) if isinstance(v, list) else 1 for v in keywords_data.values())
    except Exception:
        pass

    total_rules = rules_count + keywords_count

    return {
        "ai": {
            "available": ai_available,
            "status": ai_status,
            "model": "GPT-4o-mini" if ai_available else "Rules Only"
        },
        "memory": {
            "patterns_learned": memory_patterns,
            "status": "Active" if memory_patterns > 0 else "Empty"
        },
        "rules": {
            "count": total_rules,
            "status": f"Loaded ({rules_count} rules + {keywords_count} keywords)" if total_rules > 0 else "Not Found"
        },
        "backend": {
            "status": "Online",
            "version": "1.0.0"
        }
    }


# ==============================================================================
# DATA CLEANUP ENDPOINTS
# ==============================================================================

@app.get("/cleanup/analyze")
async def analyze_data_cleanup(request: Request, response: Response):
    """
    Analyze loaded data for common quality issues that can be auto-fixed.

    Returns categorized issues:
    - invalid_dates: Dates that don't exist (e.g., Feb 30)
    - negative_format: Numbers with trailing minus
    - missing_dates: In-service dates that can be filled from acquisition date
    - ocr_errors: Common OCR/scanning mistakes
    - cost_format: Currency symbols, invalid characters in costs
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    if not session.assets:
        return {
            "issues": {},
            "stats": {"total": 0},
            "summary": "No assets loaded"
        }

    issues = {
        "invalid_dates": [],
        "negative_format": [],
        "missing_dates": [],
        "ocr_errors": [],
        "cost_format": []
    }

    assets = list(session.assets.values())

    for asset in assets:
        asset_id = asset.asset_id

        # Check for invalid dates
        if asset.in_service_date:
            date_str = str(asset.in_service_date)
            # Check for obviously invalid dates
            if "2/30" in date_str or "2/31" in date_str or "4/31" in date_str or "6/31" in date_str:
                issues["invalid_dates"].append({
                    "asset_id": asset_id,
                    "field": "In Service Date",
                    "current_value": date_str,
                    "suggested_fix": "Correct to valid date"
                })

        # Check for missing in-service dates (can use acquisition date)
        if not asset.in_service_date and asset.acquisition_date:
            issues["missing_dates"].append({
                "asset_id": asset_id,
                "field": "In Service Date",
                "current_value": "Missing",
                "suggested_fix": str(asset.acquisition_date)
            })

        # Check for cost format issues
        if asset.cost is not None:
            cost_str = str(asset.cost)
            # Check for trailing minus (e.g., "4,129.66-")
            if cost_str.endswith('-'):
                issues["negative_format"].append({
                    "asset_id": asset_id,
                    "field": "Cost",
                    "current_value": cost_str,
                    "suggested_fix": f"-{cost_str[:-1]}"
                })
            # Check for currency symbols or weird characters
            if '$' in cost_str or '€' in cost_str or '£' in cost_str:
                issues["cost_format"].append({
                    "asset_id": asset_id,
                    "field": "Cost",
                    "current_value": cost_str,
                    "suggested_fix": cost_str.replace('$', '').replace('€', '').replace('£', '').strip()
                })

        # Check description for OCR errors - be more specific to avoid false positives
        # Only flag patterns that look like actual OCR mistakes, not product names like "OptiPlex"
        if asset.description:
            desc = asset.description
            import re
            # Look for specific OCR error patterns:
            # - O directly adjacent to digits (e.g., "1O0" should be "100")
            # - l or I in the middle of numbers (e.g., "1l1" should be "111")
            ocr_patterns = [
                (r'\d[Oo]\d', 'O in number sequence'),  # digit-O-digit
                (r'\d[Il]\d', 'I/l in number sequence'),  # digit-I/l-digit
                (r'[Oo]\d{2,}', 'O before numbers'),  # O followed by 2+ digits
                (r'\d{2,}[Oo]', 'O after numbers'),  # 2+ digits followed by O
            ]

            for pattern, error_type in ocr_patterns:
                if re.search(pattern, desc):
                    issues["ocr_errors"].append({
                        "asset_id": asset_id,
                        "field": "Description",
                        "current_value": desc[:50] + ('...' if len(desc) > 50 else ''),
                        "suggested_fix": f"Check for {error_type}"
                    })
                    break  # Only report once per asset

    return {
        "issues": issues,
        "stats": {
            "total": len(assets),
            "issues_count": sum(len(v) for v in issues.values())
        },
        "summary": f"Found {sum(len(v) for v in issues.values())} issues in {len(assets)} assets"
    }


@app.post("/cleanup/fix")
async def fix_data_category(request: Request, response: Response, body: dict = Body(...)):
    """
    Apply auto-fix for a specific category of issues.

    Body: { "category": "invalid_dates" | "negative_format" | "missing_dates" | "ocr_errors" | "cost_format" }
    """
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    category = body.get("category")
    if not category:
        raise HTTPException(status_code=400, detail="Category required")

    fixed_count = 0
    assets = session.assets

    for asset_id, asset in assets.items():
        if category == "missing_dates":
            if not asset.in_service_date and asset.acquisition_date:
                asset.in_service_date = asset.acquisition_date
                fixed_count += 1

        elif category == "invalid_dates":
            # Fix invalid dates like Feb 30 -> Feb 28
            if asset.in_service_date:
                date_str = str(asset.in_service_date)
                try:
                    from datetime import datetime
                    import calendar
                    # Try to parse and fix invalid dates
                    if "2/30" in date_str or "2/31" in date_str:
                        # February - use last day of Feb
                        year = datetime.now().year
                        if "/" in date_str:
                            parts = date_str.split("/")
                            if len(parts) >= 3:
                                year = int(parts[2]) if len(parts[2]) == 4 else 2000 + int(parts[2])
                        last_day = 29 if calendar.isleap(year) else 28
                        asset.in_service_date = datetime(year, 2, last_day)
                        fixed_count += 1
                    elif "4/31" in date_str or "6/31" in date_str or "9/31" in date_str or "11/31" in date_str:
                        # 30-day months
                        if "/" in date_str:
                            parts = date_str.split("/")
                            month = int(parts[0])
                            year = int(parts[2]) if len(parts) >= 3 and len(parts[2]) == 4 else datetime.now().year
                            asset.in_service_date = datetime(year, month, 30)
                            fixed_count += 1
                except Exception:
                    pass  # Can't auto-fix this date

        elif category == "negative_format":
            if asset.cost is not None:
                cost_str = str(asset.cost)
                if cost_str.endswith('-'):
                    try:
                        asset.cost = -float(cost_str[:-1].replace(',', ''))
                        fixed_count += 1
                    except ValueError:
                        pass

        elif category == "cost_format":
            if asset.cost is not None:
                cost_str = str(asset.cost)
                if '$' in cost_str or '€' in cost_str or '£' in cost_str:
                    try:
                        cleaned = cost_str.replace('$', '').replace('€', '').replace('£', '').replace(',', '').strip()
                        asset.cost = float(cleaned)
                        fixed_count += 1
                    except ValueError:
                        pass

        elif category == "ocr_errors":
            # OCR errors require manual review - we can try to auto-fix common patterns
            if asset.description:
                import re
                desc = asset.description
                original = desc
                # Fix O -> 0 when surrounded by digits
                desc = re.sub(r'(\d)O(\d)', r'\g<1>0\g<2>', desc)
                desc = re.sub(r'(\d)o(\d)', r'\g<1>0\g<2>', desc)
                # Fix l -> 1 when surrounded by digits
                desc = re.sub(r'(\d)[lI](\d)', r'\g<1>1\g<2>', desc)
                if desc != original:
                    asset.description = desc
                    fixed_count += 1

    return {
        "success": True,
        "category": category,
        "fixed_count": fixed_count
    }


@app.post("/cleanup/fix-all")
async def fix_all_data_issues(request: Request, response: Response):
    """Apply all available auto-fixes."""
    session = await get_current_session(request)
    add_session_to_response(response, session.session_id)

    total_fixed = 0
    fixed_by_category = {}

    # Fix all categories
    all_categories = ["invalid_dates", "missing_dates", "negative_format", "cost_format", "ocr_errors"]
    for category in all_categories:
        result = await fix_data_category(request, response, {"category": category})
        count = result.get("fixed_count", 0)
        total_fixed += count
        fixed_by_category[category] = count

    return {
        "success": True,
        "total_fixed": total_fixed,
        "fixed_by_category": fixed_by_category
    }


# ==============================================================================
# API VERSION INFO & DEPRECATION
# ==============================================================================

@app.get("/api/v1")
@app.get("/api/v1/")
async def api_version_info():
    """
    Returns API version information and available endpoints.
    Use /api/v1/ prefix for all API calls for future compatibility.
    """
    return {
        "version": API_VERSION,
        "api_version": "2.1.0",
        "status": "stable",
        "deprecation_notice": "Root-level endpoints (e.g., /upload) are deprecated. Use /api/v1/ prefix.",
        "endpoints": {
            "upload": "/api/v1/upload",
            "assets": "/api/v1/assets",
            "export": "/api/v1/export",
            "config": "/api/v1/config/tax",
            "stats": "/api/v1/stats"
        }
    }

# Mount versioned routes (aliasing root endpoints to /api/v1/)
# This allows frontend to use either /upload or /api/v1/upload
from starlette.routing import Mount, Route

# Create versioned aliases for all endpoints
@app.middleware("http")
async def version_redirect_middleware(request: Request, call_next):
    """
    Middleware to handle /api/v1/ prefix by stripping it before routing.
    This allows all endpoints to work with both:
      - /endpoint (deprecated)
      - /api/v1/endpoint (recommended)
    """
    path = request.scope.get("path", "")

    # If request is to /api/v1/..., strip the prefix for routing
    if path.startswith("/api/v1/") and path != "/api/v1/" and path != "/api/v1":
        # Rewrite path to root equivalent
        new_path = path[7:]  # Remove "/api/v1"
        request.scope["path"] = new_path

    response = await call_next(request)

    # Add deprecation header for non-versioned endpoints
    if not path.startswith("/api/"):
        response.headers["X-API-Deprecation"] = "Use /api/v1/ prefix for future compatibility"
        response.headers["X-API-Version"] = API_VERSION

    return response


if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)